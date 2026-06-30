import os
import io
from fastapi import Depends, FastAPI, File, Form, Header, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import Response
from pydantic import BaseModel
from sqlalchemy import func, text
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session, selectinload
from datetime import date, datetime

from . import models
from .database import Base, engine, get_db
from .seed import seed_database

app = FastAPI(title="SemiPLM API", version="0.1.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


PERMISSION_LABELS = {
    "all": "全部权限",
    "system": "系统设置",
    "organization": "组织管理",
    "user": "用户管理",
    "role": "角色权限",
    "workflow": "流程配置",
    "integration": "集成配置",
    "product": "产品主数据",
    "requirement": "需求规格",
    "material": "研发物料",
    "bom": "设计 BOM",
    "document": "文档管理",
    "process": "工艺路线",
    "change": "工程变更",
    "project": "项目管理",
    "quality": "质量闭环",
    "approval": "审批处理",
    "erp": "ERP 接口",
    "mes": "MES 接口",
}

USER_ROLE_FALLBACKS = {
    "admin": "ADMIN",
    "luofusen": "PE_ENGINEER",
    "yushuaibing": "PE_ENGINEER",
    "zhanghao": "PE_ENGINEER",
    "fanglei": "PM_MANAGER",
    "liangweiwei": "IT_ENGINEER",
}


def split_permissions(value: str | None) -> list[str]:
    if not value:
        return []
    return [item.strip() for item in value.replace("，", ",").split(",") if item.strip()]


def current_user_context(
    db: Session = Depends(get_db),
    username: str = Header("admin", alias="X-SemiPLM-User"),
) -> dict:
    user = db.query(models.User).filter(models.User.username == username).first()
    if not user:
        user = db.query(models.User).filter(models.User.username == "admin").first()
    if not user:
        raise HTTPException(status_code=401, detail="No active user")
    role = db.query(models.Role).filter(models.Role.name == user.role, models.Role.status == "启用").first()
    if not role:
        fallback_code = USER_ROLE_FALLBACKS.get(user.username)
        if fallback_code:
            role = db.query(models.Role).filter(models.Role.code == fallback_code, models.Role.status == "启用").first()
    if not role:
        role_keyword = user.role.replace("整合", "")
        role = (
            db.query(models.Role)
            .filter(models.Role.name.like(f"%{role_keyword}%"), models.Role.status == "启用")
            .first()
        )
    permissions = split_permissions(role.permissions if role else "")
    return {"user": user, "role": role, "permissions": permissions}


def has_permission(context: dict, required: str | list[str] | tuple[str, ...]) -> bool:
    permissions = set(context.get("permissions", []))
    if "all" in permissions:
        return True
    required_items = [required] if isinstance(required, str) else list(required)
    return any(item in permissions for item in required_items)


def require_permission(required: str | list[str] | tuple[str, ...]):
    def dependency(context: dict = Depends(current_user_context)) -> dict:
        if not has_permission(context, required):
            raise HTTPException(status_code=403, detail="Permission denied")
        return context

    return dependency


@app.on_event("startup")
def startup() -> None:
    Base.metadata.create_all(bind=engine)
    ensure_lightweight_schema()
    db = next(get_db())
    try:
        seed_database(db)
        normalize_single_org_master_data(db)
    finally:
        db.close()


def ensure_lightweight_schema() -> None:
    columns = {
        "bom_headers": {
            "source_bom_id": "INTEGER",
            "effective_date": "VARCHAR(30) DEFAULT ''",
            "expiry_date": "VARCHAR(30) DEFAULT ''",
            "effectivity_type": "VARCHAR(30) DEFAULT '日期'",
            "effective_batch": "VARCHAR(80) DEFAULT ''",
        },
        "bom_items": {
            "process_step_id": "INTEGER",
            "effective_date": "VARCHAR(30) DEFAULT ''",
            "expiry_date": "VARCHAR(30) DEFAULT ''",
            "effectivity_note": "VARCHAR(160) DEFAULT ''",
        },
        "process_routes": {
            "release_date": "VARCHAR(30) DEFAULT ''",
            "source_route_id": "INTEGER",
            "effective_batch": "VARCHAR(80) DEFAULT ''",
        },
        "integration_jobs": {
            "attempt_count": "INTEGER DEFAULT 0",
            "last_sync_at": "VARCHAR(30) DEFAULT ''",
            "response_message": "TEXT DEFAULT ''",
            "external_id": "VARCHAR(120) DEFAULT ''",
        },
        "change_actions": {
            "target_type": "VARCHAR(40) DEFAULT ''",
            "target_id": "INTEGER",
            "target_version": "VARCHAR(30) DEFAULT ''",
            "effectivity_type": "VARCHAR(30) DEFAULT '日期'",
            "effectivity_scope": "VARCHAR(80) DEFAULT ''",
            "effective_date": "VARCHAR(30) DEFAULT ''",
            "effective_batch": "VARCHAR(80) DEFAULT ''",
            "generated_object_no": "VARCHAR(160) DEFAULT ''",
        },
        "operation_logs": {
            "object_id": "INTEGER",
        },
        "project_deliverables": {
            "object_type": "VARCHAR(40) DEFAULT ''",
            "object_id": "INTEGER",
        },
        "documents": {
            "file_name": "VARCHAR(255) DEFAULT ''",
            "file_path": "VARCHAR(500) DEFAULT ''",
            "file_size": "INTEGER",
            "file_type": "VARCHAR(100) DEFAULT ''",
        },
        "project_tasks": {
            "start_date": "VARCHAR(30) DEFAULT ''",
            "parent_id": "INTEGER",
            "depends_on": "VARCHAR(200) DEFAULT ''",
        },
        "materials": {
            "supplier_id": "INTEGER",
        },
        "substitute_materials": {
            "material_id": "INTEGER",
            "substitute_material_id": "INTEGER",
        },
    }
    with engine.begin() as conn:
        for table, table_columns in columns.items():
            existing = {row[1] for row in conn.execute(text(f"PRAGMA table_info({table})")).fetchall()}
            for column, definition in table_columns.items():
                if column not in existing:
                    conn.execute(text(f"ALTER TABLE {table} ADD COLUMN {column} {definition}"))


def normalize_single_org_master_data(db: Session) -> None:
    role_rows = [
        ("ADMIN", "系统管理员", "系统配置、组织角色、接口配置、所有对象维护", "system,user,role,workflow,integration,all"),
        ("PE_ENGINEER", "工艺工程师", "工艺路线、工艺参数、变更影响评估", "process,material,change,approval"),
        ("PM_MANAGER", "项目经理", "NPI 项目、阶段门、计划和风险", "project,workflow,approval"),
        ("IT_ENGINEER", "IT工程师", "系统配置、组织用户角色、流程和接口维护", "organization,system,user,role,workflow,integration"),
    ]
    for code, name, description, permissions in role_rows:
        role = db.query(models.Role).filter(models.Role.code == code).first()
        if role:
            role.name = name
            role.description = description
            role.permissions = permissions
            role.status = "启用"
        else:
            db.add(models.Role(code=code, name=name, description=description, permissions=permissions, status="启用"))

    org = db.query(models.Organization).filter(models.Organization.code == "NZGD").first()
    if org:
        org.name = "南智光电"
        org.org_type = "公司"
        org.parent_code = ""
        org.manager = "梁维维"
        org.status = "启用"
        org.description = "公司主体"
    else:
        db.add(models.Organization(code="NZGD", name="南智光电", org_type="公司", parent_code="", manager="梁维维", status="启用", description="公司主体"))
    dept = db.query(models.Organization).filter(models.Organization.code == "PROD").first()
    if dept:
        dept.name = "生产部"
        dept.org_type = "部门"
        dept.parent_code = "NZGD"
        dept.manager = "房磊"
        dept.status = "启用"
        dept.description = "生产运营部门"
    else:
        db.add(models.Organization(code="PROD", name="生产部", org_type="部门", parent_code="NZGD", manager="房磊", status="启用", description="生产运营部门"))

    stale_users = {"rd01", "pe01", "qa01", "pm01"}
    db.query(models.User).filter(models.User.username.in_(stale_users)).delete(synchronize_session=False)
    user_rows = [
        ("admin", "系统管理员", "管理员", "生产部"),
        ("luofusen", "罗富森", "工艺工程师", "生产部"),
        ("yushuaibing", "于帅兵", "工艺工程师", "生产部"),
        ("zhanghao", "张昊", "工艺工程师", "生产部"),
        ("fanglei", "房磊", "项目经理", "生产部"),
        ("liangweiwei", "梁维维", "IT工程师", "生产部"),
    ]
    for username, display_name, role, department in user_rows:
        user = db.query(models.User).filter(models.User.username == username).first()
        if user:
            user.display_name = display_name
            user.role = role
            user.department = department
        else:
            db.add(models.User(username=username, display_name=display_name, role=role, department=department))

    name_map = {
        "林晨": "罗富森",
        "宋工": "罗富森",
        "许沐": "于帅兵",
        "周宁": "房磊",
        "陈知": "张昊",
        "罗远": "张昊",
    }
    department_map = {
        "平台管理": "生产部",
        "光电工艺部": "生产部",
        "光刻刻蚀镀膜工程部": "生产部",
        "工艺部": "生产部",
        "质量部": "生产部",
        "研发中心": "生产部",
        "工艺工程部": "生产部",
        "制造部": "生产部",
        "文控中心": "生产部",
        "项目管理部": "生产部",
        "IT部": "生产部",
        "项目/制造/质量": "生产部",
    }

    def normalize_text(value: str) -> str:
        for old, new in {**name_map, **department_map}.items():
            value = value.replace(old, new)
        return value

    for model, columns in [
        (models.Product, ["owner"]),
        (models.BomHeader, ["owner"]),
        (models.Document, ["owner"]),
        (models.ProcessRoute, ["owner"]),
        (models.ProcessStep, ["owner"]),
        (models.Change, ["owner", "reason", "before_desc", "after_desc"]),
        (models.ChangeImpact, ["target", "action"]),
        (models.Approval, ["approver", "comment"]),
        (models.ChangeAction, ["department", "owner", "result"]),
        (models.IntegrationEndpoint, ["owner"]),
        (models.Project, ["owner"]),
        (models.ProjectTask, ["owner"]),
        (models.QualityIssue, ["owner"]),
        (models.Requirement, ["owner"]),
        (models.ProductBaseline, ["created_by"]),
        (models.BaselineItem, ["object_no", "title"]),
        (models.CodingRule, ["owner"]),
        (models.WorkflowTask, ["assignee", "acted_by", "comment"]),
        (models.WorkflowInstance, ["started_by"]),
        (models.WorkflowLog, ["actor", "comment"]),
        (models.IntegrationJob, ["triggered_by", "message", "response_message"]),
    ]:
        for row in db.query(model).all():
            for column in columns:
                value = getattr(row, column, "")
                if isinstance(value, str) and value:
                    setattr(row, column, normalize_text(value))

    for action in db.query(models.ChangeAction).all():
        if not action.effectivity_type:
            action.effectivity_type = "日期"
        if not action.effective_date and "日期" in action.effectivity_type:
            action.effective_date = action.due_date or today_text()
        if not action.effective_batch and "LOT-" in action.target_object:
            action.effectivity_type = "日期+批次" if "日期" in action.effectivity_type else "批次"
            action.effective_batch = action.target_object[action.target_object.find("LOT-") :].split()[0]

    test_prefixes = ("ECR-API-VERIFY", "ECR-IMPACT-VERIFY", "ECR-CLOSE-FLOW")
    test_changes = db.query(models.Change).filter(
        (models.Change.change_no.like(f"{test_prefixes[0]}%"))
        | (models.Change.change_no.like(f"{test_prefixes[1]}%"))
        | (models.Change.change_no.like(f"{test_prefixes[2]}%"))
    ).all()
    test_change_ids = [row.id for row in test_changes]
    test_change_nos = [row.change_no for row in test_changes]
    if test_change_ids:
        db.query(models.ChangeAction).filter(models.ChangeAction.change_id.in_(test_change_ids)).delete(synchronize_session=False)
        db.query(models.ChangeImpact).filter(models.ChangeImpact.change_id.in_(test_change_ids)).delete(synchronize_session=False)
        db.query(models.Approval).filter(models.Approval.change_id.in_(test_change_ids)).delete(synchronize_session=False)
        db.query(models.Change).filter(models.Change.id.in_(test_change_ids)).delete(synchronize_session=False)
    if test_change_nos:
        workflow_ids = [
            row.id
            for row in db.query(models.WorkflowInstance.id).filter(models.WorkflowInstance.object_no.in_(test_change_nos)).all()
        ]
        if workflow_ids:
            db.query(models.WorkflowLog).filter(models.WorkflowLog.instance_id.in_(workflow_ids)).delete(synchronize_session=False)
            db.query(models.WorkflowTask).filter(models.WorkflowTask.instance_id.in_(workflow_ids)).delete(synchronize_session=False)
            db.query(models.WorkflowInstance).filter(models.WorkflowInstance.id.in_(workflow_ids)).delete(synchronize_session=False)
        db.query(models.IntegrationJob).filter(models.IntegrationJob.object_no.in_(test_change_nos)).delete(synchronize_session=False)
    db.commit()


def serialize_product(product: models.Product) -> dict:
    return {
        "id": product.id,
        "code": product.code,
        "model": product.model,
        "name": product.name,
        "product_type": product.product_type,
        "process_platform": product.process_platform,
        "wafer_size": product.wafer_size,
        "package_type": product.package_type,
        "temperature_grade": product.temperature_grade,
        "quality_grade": product.quality_grade,
        "application": product.application,
        "lifecycle": product.lifecycle,
        "owner": product.owner,
        "customer_part_no": product.customer_part_no,
        "internal_part_no": product.internal_part_no,
        "version": product.version,
        "readiness": product.readiness,
        "latest_release": product.latest_release,
    }


class ProductPayload(BaseModel):
    code: str
    model: str
    name: str
    product_type: str = "光电芯片"
    process_platform: str = ""
    wafer_size: str = ""
    package_type: str = ""
    temperature_grade: str = ""
    quality_grade: str = ""
    application: str = ""
    lifecycle: str = "设计中"
    owner: str = ""
    customer_part_no: str = ""
    internal_part_no: str = ""
    version: str = "A0"
    readiness: int = 0
    latest_release: str = ""


class ProductVersionPayload(BaseModel):
    version: str
    lifecycle: str = ""
    readiness: int = 0
    released_at: str = ""
    released_by: str = ""
    source_change_no: str = ""
    summary: str = ""


class ProductUpdatePayload(BaseModel):
    code: str | None = None
    model: str | None = None
    name: str | None = None
    product_type: str | None = None
    process_platform: str | None = None
    wafer_size: str | None = None
    package_type: str | None = None
    temperature_grade: str | None = None
    quality_grade: str | None = None
    application: str | None = None
    lifecycle: str | None = None
    owner: str | None = None
    customer_part_no: str | None = None
    internal_part_no: str | None = None
    version: str | None = None
    readiness: int | None = None
    latest_release: str | None = None


class MaterialPayload(BaseModel):
    code: str
    name: str
    category: str
    specification: str = ""
    supplier: str = ""
    supplier_id: int | None = None
    risk_level: str = "低"
    lifecycle: str = "有效"


class MaterialUpdatePayload(BaseModel):
    code: str | None = None
    name: str | None = None
    category: str | None = None
    specification: str | None = None
    supplier: str | None = None
    supplier_id: int | None = None
    risk_level: str | None = None
    lifecycle: str | None = None


class RequirementPayload(BaseModel):
    product_id: int
    req_no: str
    source: str
    category: str
    title: str
    priority: str = "中"
    status: str = "草稿"
    owner: str = ""
    acceptance_criteria: str = ""


class RequirementUpdatePayload(BaseModel):
    product_id: int | None = None
    req_no: str | None = None
    source: str | None = None
    category: str | None = None
    title: str | None = None
    priority: str | None = None
    status: str | None = None
    owner: str | None = None
    acceptance_criteria: str | None = None


class DocumentPayload(BaseModel):
    product_id: int
    doc_no: str
    title: str
    category: str
    version: str = "A0"
    status: str = "编制中"
    owner: str = ""
    approval_status: str = "未提交"
    updated_at: str = ""


class DocumentUpdatePayload(BaseModel):
    product_id: int | None = None
    doc_no: str | None = None
    title: str | None = None
    category: str | None = None
    version: str | None = None
    status: str | None = None
    owner: str | None = None
    approval_status: str | None = None
    updated_at: str | None = None


class BomHeaderPayload(BaseModel):
    product_id: int
    bom_type: str = "EBOM"
    version: str = "A0"
    status: str = "编制中"
    owner: str = ""
    release_date: str = ""
    source_bom_id: int | None = None
    effective_date: str = ""
    expiry_date: str = ""
    effectivity_type: str = "日期"
    effective_batch: str = ""


class BomHeaderUpdatePayload(BaseModel):
    product_id: int | None = None
    bom_type: str | None = None
    version: str | None = None
    status: str | None = None
    owner: str | None = None
    release_date: str | None = None
    source_bom_id: int | None = None
    effective_date: str | None = None
    expiry_date: str | None = None
    effectivity_type: str | None = None
    effective_batch: str | None = None


class BomItemPayload(BaseModel):
    material_code: str
    material_name: str
    category: str = ""
    specification: str = ""
    quantity: float = 1
    unit: str = "件"
    position: str = ""
    process_step_id: int | None = None
    process_step: str = ""
    substitute: str = ""
    status: str = "有效"
    effective_date: str = ""
    expiry_date: str = ""
    effectivity_note: str = ""


class BomItemUpdatePayload(BaseModel):
    material_code: str | None = None
    material_name: str | None = None
    category: str | None = None
    specification: str | None = None
    quantity: float | None = None
    unit: str | None = None
    position: str | None = None
    process_step_id: int | None = None
    process_step: str | None = None
    substitute: str | None = None
    status: str | None = None
    effective_date: str | None = None
    expiry_date: str | None = None
    effectivity_note: str | None = None


class BomTransformPayload(BaseModel):
    target_type: str = "PBOM"
    version: str = "A0"
    owner: str = ""
    effective_date: str = ""
    effectivity_type: str = "日期"
    effective_batch: str = ""


class ProcessRouteActionPayload(BaseModel):
    acted_by: str = "系统用户"
    comment: str = ""


class ProcessRoutePayload(BaseModel):
    product_id: int
    route_no: str
    name: str
    version: str = "A0"
    status: str = "编制中"
    owner: str = ""
    release_date: str = ""
    source_route_id: int | None = None
    effective_batch: str = ""


class ProcessRouteUpdatePayload(BaseModel):
    product_id: int | None = None
    route_no: str | None = None
    name: str | None = None
    version: str | None = None
    status: str | None = None
    owner: str | None = None
    release_date: str | None = None
    source_route_id: int | None = None
    effective_batch: str | None = None


class ProcessStepPayload(BaseModel):
    sequence: int
    stage: str
    operation: str
    key_params: str = ""
    owner: str = ""
    status: str = "有效"


class ProcessStepUpdatePayload(BaseModel):
    sequence: int | None = None
    stage: str | None = None
    operation: str | None = None
    key_params: str | None = None
    owner: str | None = None
    status: str | None = None


class IntegrationJobActionPayload(BaseModel):
    acted_by: str = "系统用户"
    response_message: str = ""
    external_id: str = ""


class ChangePayload(BaseModel):
    product_id: int
    change_no: str
    title: str
    change_type: str = "ECR"
    reason: str = ""
    status: str = "草稿"
    priority: str = "中"
    owner: str = ""
    submitted_at: str = ""
    before_desc: str = ""
    after_desc: str = ""


class ChangeUpdatePayload(BaseModel):
    product_id: int | None = None
    change_no: str | None = None
    title: str | None = None
    change_type: str | None = None
    reason: str | None = None
    status: str | None = None
    priority: str | None = None
    owner: str | None = None
    submitted_at: str | None = None
    before_desc: str | None = None
    after_desc: str | None = None


class ChangeImpactPayload(BaseModel):
    impact_type: str
    target: str
    risk: str = "中"
    action: str = ""


class ChangeImpactUpdatePayload(BaseModel):
    impact_type: str | None = None
    target: str | None = None
    risk: str | None = None
    action: str | None = None


class ChangeActionPayload(BaseModel):
    action_no: str = ""
    action_type: str = "资料更新"
    target_type: str = ""
    target_id: int | None = None
    target_version: str = ""
    target_object: str
    effectivity_type: str = "日期"
    effectivity_scope: str = ""
    effective_date: str = ""
    effective_batch: str = ""
    generated_object_no: str = ""
    department: str = ""
    owner: str = ""
    status: str = "待处理"
    due_date: str = ""
    result: str = ""


class ChangeActionUpdatePayload(BaseModel):
    action_no: str | None = None
    action_type: str | None = None
    target_type: str | None = None
    target_id: int | None = None
    target_version: str | None = None
    target_object: str | None = None
    effectivity_type: str | None = None
    effectivity_scope: str | None = None
    effective_date: str | None = None
    effective_batch: str | None = None
    generated_object_no: str | None = None
    department: str | None = None
    owner: str | None = None
    status: str | None = None
    due_date: str | None = None
    result: str | None = None


class ChangeActionClosePayload(BaseModel):
    acted_by: str = "系统用户"
    result: str = "已完成"


class ProblemReportPayload(BaseModel):
    pr_no: str = ""
    title: str
    problem_type: str = "设计问题"
    severity: str = "中"
    source: str = "内部"
    product_id: int | None = None
    product_model: str = ""
    description: str = ""
    suggested_action: str = ""
    status: str = "新建"
    reporter: str = ""
    reported_at: str = ""
    related_change_no: str = ""
    remark: str = ""


class ProblemReportUpdatePayload(BaseModel):
    title: str | None = None
    problem_type: str | None = None
    severity: str | None = None
    source: str | None = None
    product_id: int | None = None
    product_model: str | None = None
    description: str | None = None
    suggested_action: str | None = None
    status: str | None = None
    reporter: str | None = None
    reported_at: str | None = None
    related_change_no: str | None = None
    remark: str | None = None


class ProcessParameterPayload(BaseModel):
    param_code: str = ""
    param_name: str
    param_type: str = "CD"
    unit: str = ""
    category: str = ""
    default_value: str = ""
    min_value: str = ""
    max_value: str = ""
    description: str = ""
    status: str = "启用"


class ProcessParameterUpdatePayload(BaseModel):
    param_code: str | None = None
    param_name: str | None = None
    param_type: str | None = None
    unit: str | None = None
    category: str | None = None
    default_value: str | None = None
    min_value: str | None = None
    max_value: str | None = None
    description: str | None = None
    status: str | None = None


class OrganizationPayload(BaseModel):
    code: str
    name: str
    org_type: str = "部门"
    parent_code: str = "NZGD"
    manager: str = ""
    status: str = "启用"
    description: str = ""


class OrganizationUpdatePayload(BaseModel):
    code: str | None = None
    name: str | None = None
    org_type: str | None = None
    parent_code: str | None = None
    manager: str | None = None
    status: str | None = None
    description: str | None = None


class UserPayload(BaseModel):
    username: str
    display_name: str
    role: str
    department: str = "生产部"


class UserUpdatePayload(BaseModel):
    username: str | None = None
    display_name: str | None = None
    role: str | None = None
    department: str | None = None


class RolePayload(BaseModel):
    code: str
    name: str
    description: str = ""
    permissions: str = ""
    status: str = "启用"


class RoleUpdatePayload(BaseModel):
    code: str | None = None
    name: str | None = None
    description: str | None = None
    permissions: str | None = None
    status: str | None = None


class CodingRulePayload(BaseModel):
    object_type: str
    code: str
    name: str
    prefix: str
    pattern: str
    current_no: int = 0
    sample: str = ""
    status: str = "启用"
    owner: str = ""


class CodingRuleUpdatePayload(BaseModel):
    object_type: str | None = None
    code: str | None = None
    name: str | None = None
    prefix: str | None = None
    pattern: str | None = None
    current_no: int | None = None
    sample: str | None = None
    status: str | None = None
    owner: str | None = None


class CategoryTemplatePayload(BaseModel):
    object_type: str
    code: str
    name: str
    parent_code: str = ""
    lifecycle_template: str = ""
    coding_rule: str = ""
    status: str = "启用"
    description: str = ""


class CategoryTemplateUpdatePayload(BaseModel):
    object_type: str | None = None
    code: str | None = None
    name: str | None = None
    parent_code: str | None = None
    lifecycle_template: str | None = None
    coding_rule: str | None = None
    status: str | None = None
    description: str | None = None


class AttributeTemplatePayload(BaseModel):
    name: str
    field_key: str
    data_type: str = "文本"
    required: str = "否"
    dictionary_code: str = ""
    default_value: str = ""
    sequence: int = 1


class AttributeTemplateUpdatePayload(BaseModel):
    name: str | None = None
    field_key: str | None = None
    data_type: str | None = None
    required: str | None = None
    dictionary_code: str | None = None
    default_value: str | None = None
    sequence: int | None = None


class LifecycleTemplatePayload(BaseModel):
    code: str
    name: str
    object_type: str
    status: str = "启用"
    description: str = ""


class LifecycleTemplateUpdatePayload(BaseModel):
    code: str | None = None
    name: str | None = None
    object_type: str | None = None
    status: str | None = None
    description: str | None = None


class LifecycleStatePayload(BaseModel):
    sequence: int
    name: str
    state_type: str = "中间态"
    allow_edit: str = "是"
    require_workflow: str = "否"
    next_states: str = ""


class LifecycleStateUpdatePayload(BaseModel):
    sequence: int | None = None
    name: str | None = None
    state_type: str | None = None
    allow_edit: str | None = None
    require_workflow: str | None = None
    next_states: str | None = None


class DictionaryItemPayload(BaseModel):
    dict_code: str
    dict_name: str
    item_value: str
    item_label: str
    object_scope: str = ""
    sequence: int = 1
    status: str = "启用"


class DictionaryItemUpdatePayload(BaseModel):
    dict_code: str | None = None
    dict_name: str | None = None
    item_value: str | None = None
    item_label: str | None = None
    object_scope: str | None = None
    sequence: int | None = None
    status: str | None = None


class WorkflowTemplatePayload(BaseModel):
    code: str
    name: str
    object_type: str
    project_type: str = ""
    status: str = "启用"
    description: str = ""


class WorkflowTemplateUpdatePayload(BaseModel):
    code: str | None = None
    name: str | None = None
    object_type: str | None = None
    project_type: str | None = None
    status: str | None = None
    description: str | None = None


class WorkflowNodePayload(BaseModel):
    sequence: int
    name: str
    role_name: str
    action_type: str = "审批"
    sla_hours: int = 24


class WorkflowNodeUpdatePayload(BaseModel):
    sequence: int | None = None
    name: str | None = None
    role_name: str | None = None
    action_type: str | None = None
    sla_hours: int | None = None


class IntegrationEndpointPayload(BaseModel):
    code: str
    name: str
    system_type: str
    base_url: str
    auth_type: str = "Token"
    direction: str = "双向"
    status: str = "启用"
    owner: str = ""
    object_scope: str = ""


class IntegrationEndpointUpdatePayload(BaseModel):
    code: str | None = None
    name: str | None = None
    system_type: str | None = None
    base_url: str | None = None
    auth_type: str | None = None
    direction: str | None = None
    status: str | None = None
    owner: str | None = None
    object_scope: str | None = None


class SystemParameterPayload(BaseModel):
    param_key: str
    param_value: str = ""
    param_group: str = "系统"
    description: str = ""


class SystemParameterUpdatePayload(BaseModel):
    param_key: str | None = None
    param_value: str | None = None
    param_group: str | None = None
    description: str | None = None


class ReportSnapshotPayload(BaseModel):
    report_type: str
    generated_by: str = "系统用户"
    schedule_key: str = "manual"


class SubstituteMaterialPayload(BaseModel):
    material_code: str
    material_name: str
    substitute_code: str
    substitute_name: str
    material_id: int | None = None
    substitute_material_id: int | None = None
    substitute_type: str = "功能替代"
    strategy: str = "一对一"
    risk_level: str = "中"
    status: str = "启用"
    effective_date: str = ""
    expiry_date: str = ""
    description: str = ""


class SubstituteMaterialUpdatePayload(BaseModel):
    material_code: str | None = None
    material_name: str | None = None
    substitute_code: str | None = None
    substitute_name: str | None = None
    material_id: int | None = None
    substitute_material_id: int | None = None
    substitute_type: str | None = None
    strategy: str | None = None
    risk_level: str | None = None
    status: str | None = None
    effective_date: str | None = None
    expiry_date: str | None = None
    description: str | None = None


class SupplierPayload(BaseModel):
    code: str
    name: str
    supplier_type: str = "材料"
    contact: str = ""
    phone: str = ""
    email: str = ""
    address: str = ""
    certification: str = ""
    risk_level: str = "中"
    status: str = "启用"
    description: str = ""


class SupplierUpdatePayload(BaseModel):
    code: str | None = None
    name: str | None = None
    supplier_type: str | None = None
    contact: str | None = None
    phone: str | None = None
    email: str | None = None
    address: str | None = None
    certification: str | None = None
    risk_level: str | None = None
    status: str | None = None
    description: str | None = None


class ProjectPayload(BaseModel):
    project_no: str
    name: str
    product_model: str = ""
    phase: str = "概念"
    progress: int = 0
    owner: str = ""
    start_date: str = ""
    end_date: str = ""
    risk_level: str = "低"


class ProjectUpdatePayload(BaseModel):
    name: str | None = None
    product_model: str | None = None
    phase: str | None = None
    progress: int | None = None
    owner: str | None = None
    start_date: str | None = None
    end_date: str | None = None
    risk_level: str | None = None


class ProjectTemplatePayload(BaseModel):
    code: str
    name: str
    description: str = ""
    stages: str = '["概念","设计","流片","验证","试产"]'
    status: str = "启用"


class ProjectTemplateUpdatePayload(BaseModel):
    name: str | None = None
    description: str | None = None
    stages: str | None = None
    status: str | None = None


class ProjectFromTemplatePayload(BaseModel):
    template_id: int
    project_no: str
    name: str
    product_model: str = ""
    owner: str = ""
    start_date: str = ""
    end_date: str = ""


class ProjectDeliverablePayload(BaseModel):
    name: str
    deliverable_type: str
    phase: str
    owner: str = ""
    status: str = "待处理"
    due_date: str = ""
    description: str = ""
    object_type: str = ""
    object_id: int | None = None


class ProjectDeliverableUpdatePayload(BaseModel):
    name: str | None = None
    deliverable_type: str | None = None
    phase: str | None = None
    owner: str | None = None
    status: str | None = None
    due_date: str | None = None
    completed_at: str | None = None
    description: str | None = None
    object_type: str | None = None
    object_id: int | None = None


class ProjectRiskPayload(BaseModel):
    risk_type: str = "技术"
    description: str = ""
    impact: str = "中"
    probability: str = "中"
    severity: str = "中"
    owner: str = ""
    status: str = "待处理"
    mitigation: str = ""


class ProjectRiskUpdatePayload(BaseModel):
    risk_type: str | None = None
    description: str | None = None
    impact: str | None = None
    probability: str | None = None
    severity: str | None = None
    owner: str | None = None
    status: str | None = None
    mitigation: str | None = None


class QualityCAPAPayload(BaseModel):
    capa_no: str = ""
    issue_id: int | None = None
    title: str
    source: str = "质量问题"
    root_cause: str = ""
    corrective_action: str = ""
    preventive_action: str = ""
    owner: str = ""
    status: str = "待处理"
    due_date: str = ""
    result: str = ""


class QualityCAPAUpdatePayload(BaseModel):
    title: str | None = None
    root_cause: str | None = None
    corrective_action: str | None = None
    preventive_action: str | None = None
    owner: str | None = None
    status: str | None = None
    due_date: str | None = None
    closed_at: str | None = None
    result: str | None = None


class QualityIssuePayload(BaseModel):
    issue_no: str = ""
    product_model: str = ""
    lot_no: str = ""
    title: str
    severity: str = "中"
    status: str = "新建"
    owner: str = ""
    root_cause: str = ""
    corrective_action: str = ""


class QualityIssueUpdatePayload(BaseModel):
    product_model: str | None = None
    lot_no: str | None = None
    title: str | None = None
    severity: str | None = None
    status: str | None = None
    owner: str | None = None
    root_cause: str | None = None
    corrective_action: str | None = None


class ProjectTaskPayload(BaseModel):
    name: str
    phase: str = ""
    owner: str = ""
    status: str = "待处理"
    due_date: str = ""
    start_date: str = ""
    parent_id: int | None = None
    depends_on: str = ""


class ProjectTaskUpdatePayload(BaseModel):
    name: str | None = None
    phase: str | None = None
    owner: str | None = None
    status: str | None = None
    due_date: str | None = None
    start_date: str | None = None
    parent_id: int | None = None
    depends_on: str | None = None


class ProjectPhaseGatePayload(BaseModel):
    acted_by: str = "系统用户"
    comment: str = ""


class QualityReportPayload(BaseModel):
    report_no: str = ""
    title: str
    report_type: str = "质量归档"
    product_model: str = ""
    issue_nos: str = ""
    capa_nos: str = ""
    summary: str = ""
    root_cause: str = ""
    corrective_action: str = ""
    preventive_action: str = ""
    owner: str = ""
    status: str = "已归档"


class QualityReportUpdatePayload(BaseModel):
    title: str | None = None
    report_type: str | None = None
    summary: str | None = None
    root_cause: str | None = None
    corrective_action: str | None = None
    preventive_action: str | None = None
    owner: str | None = None
    status: str | None = None


class QualityReportArchiveFromIssuePayload(BaseModel):
    issue_ids: list[int] = []
    title: str = ""
    owner: str = ""


class WorkflowTaskActionPayload(BaseModel):
    acted_by: str = "系统用户"
    comment: str = ""


class WorkflowRejectPayload(BaseModel):
    acted_by: str = "系统用户"
    comment: str = ""
    target_sequence: int | None = None


class WorkflowTransferPayload(BaseModel):
    acted_by: str = "系统用户"
    assignee: str
    comment: str = ""


class WorkflowWithdrawPayload(BaseModel):
    acted_by: str = "系统用户"
    comment: str = ""


def commit_or_409(db: Session, message: str) -> None:
    try:
        db.commit()
    except IntegrityError as exc:
        db.rollback()
        raise HTTPException(status_code=409, detail=message) from exc


def audit_log(db: Session, action: str, object_type: str, object_id: int | None, object_no: str, summary: str, operated_by: str) -> None:
    db.add(models.OperationLog(action=action, object_type=object_type, object_id=object_id, object_no=object_no, summary=summary, operated_by=operated_by, operated_at=today_text()))


def update_model(instance: object, payload: BaseModel) -> None:
    for key, value in payload.model_dump(exclude_unset=True).items():
        setattr(instance, key, value)


def ensure_product_exists(db: Session, product_id: int) -> None:
    if not db.query(models.Product.id).filter(models.Product.id == product_id).first():
        raise HTTPException(status_code=404, detail="Product not found")


def ensure_project_exists(db: Session, project_id: int) -> None:
    if not db.query(models.Project.id).filter(models.Project.id == project_id).first():
        raise HTTPException(status_code=404, detail="Project not found")


def today_text() -> str:
    return date.today().isoformat()


def day_before(value: str) -> str:
    if not value:
        return today_text()
    try:
        return date.fromordinal(date.fromisoformat(value).toordinal() - 1).isoformat()
    except ValueError:
        return today_text()


def is_current_effective_bom(row: models.BomHeader) -> bool:
    today = today_text()
    if row.status != "已发布":
        return False
    if row.effective_date and row.effective_date > today:
        return False
    if row.expiry_date and row.expiry_date < today:
        return False
    return True


def next_revision(value: str) -> str:
    text_value = (value or "").strip()
    if not text_value:
        return "A1"
    tail_digits = ""
    for char in reversed(text_value):
        if not char.isdigit():
            break
        tail_digits = char + tail_digits
    if tail_digits:
        prefix = text_value[: len(text_value) - len(tail_digits)]
        width = len(tail_digits)
        return f"{prefix}{int(tail_digits) + 1:0{width}d}"
    if len(text_value) == 1 and "A" <= text_value.upper() <= "Y":
        return chr(ord(text_value.upper()) + 1)
    return f"{text_value}-1"


def next_unique_bom_version(db: Session, source: models.BomHeader) -> str:
    version = next_revision(source.version)
    while db.query(models.BomHeader.id).filter(
        models.BomHeader.product_id == source.product_id,
        models.BomHeader.bom_type == source.bom_type,
        models.BomHeader.version == version,
    ).first():
        version = next_revision(version)
    return version


def next_unique_document_no(db: Session, source: models.Document, version: str) -> str:
    base_no = f"{source.doc_no}-R{version}"
    doc_no = base_no
    suffix = 1
    while db.query(models.Document.id).filter(models.Document.doc_no == doc_no).first():
        suffix += 1
        doc_no = f"{base_no}-{suffix}"
    return doc_no


def next_unique_process_version(db: Session, source: models.ProcessRoute) -> str:
    version = next_revision(source.version)
    while db.query(models.ProcessRoute.id).filter(
        models.ProcessRoute.product_id == source.product_id,
        models.ProcessRoute.version == version,
    ).first():
        version = next_revision(version)
    return version


def next_unique_route_no(db: Session, source: models.ProcessRoute, version: str) -> str:
    base_no = f"{source.route_no}-R{version}"
    route_no = base_no
    suffix = 1
    while db.query(models.ProcessRoute.id).filter(models.ProcessRoute.route_no == route_no).first():
        suffix += 1
        route_no = f"{base_no}-{suffix}"
    return route_no


def close_previous_effective_boms(db: Session, bom: models.BomHeader) -> list[dict]:
    cutoff = day_before(bom.effective_date or today_text())
    rows = (
        db.query(models.BomHeader)
        .filter(
            models.BomHeader.id != bom.id,
            models.BomHeader.product_id == bom.product_id,
            models.BomHeader.bom_type == bom.bom_type,
            models.BomHeader.status == "已发布",
        )
        .all()
    )
    closed = []
    for row in rows:
        if row.effective_date and bom.effective_date and row.effective_date > bom.effective_date:
            continue
        if row.expiry_date and row.expiry_date <= cutoff:
            continue
        row.expiry_date = cutoff
        closed.append({"id": row.id, "type": row.bom_type, "version": row.version, "expiry_date": row.expiry_date})
    return closed


def serialize_bom_item(item: models.BomItem) -> dict:
    return {
        "id": item.id,
        "parent_id": item.parent_id,
        "material_code": item.material_code,
        "material_name": item.material_name,
        "category": item.category,
        "specification": item.specification,
        "quantity": item.quantity,
        "unit": item.unit,
        "position": item.position,
        "process_step_id": item.process_step_id,
        "process_step": item.process_step,
        "substitute": item.substitute,
        "status": item.status,
        "effective_date": item.effective_date,
        "expiry_date": item.expiry_date,
        "effectivity_note": item.effectivity_note,
    }


def serialize_bom(row: models.BomHeader) -> dict:
    return {
        "id": row.id,
        "product_id": row.product_id,
        "product_model": row.product.model,
        "product_name": row.product.name,
        "bom_type": row.bom_type,
        "type": row.bom_type,
        "version": row.version,
        "status": row.status,
        "owner": row.owner,
        "release_date": row.release_date,
        "source_bom_id": row.source_bom_id,
        "effective_date": row.effective_date,
        "expiry_date": row.expiry_date,
        "effectivity_type": row.effectivity_type,
        "is_current": is_current_effective_bom(row),
        "effective_batch": row.effective_batch,
        "items": [serialize_bom_item(item) for item in sorted(row.items, key=lambda item: item.id)],
    }


def bom_item_compare_key(item: models.BomItem) -> tuple[str, str, str]:
    return (item.material_code, item.process_step or "", item.position or "")


def serialize_bom_compare_item(kind: str, item: models.BomItem | None, base: models.BomItem | None = None) -> dict:
    active = item or base
    assert active is not None
    return {
        "change_type": kind,
        "material_code": active.material_code,
        "material_name": active.material_name,
        "process_step": active.process_step,
        "from_quantity": base.quantity if base else None,
        "to_quantity": item.quantity if item else None,
        "from_status": base.status if base else "",
        "to_status": item.status if item else "",
        "from_effective_date": base.effective_date if base else "",
        "to_effective_date": item.effective_date if item else "",
    }


def serialize_process_route(row: models.ProcessRoute) -> dict:
    return {
        "id": row.id,
        "product_id": row.product_id,
        "route_no": row.route_no,
        "name": row.name,
        "product_model": row.product.model,
        "version": row.version,
        "status": row.status,
        "owner": row.owner,
        "release_date": row.release_date,
        "source_route_id": row.source_route_id,
        "effective_batch": row.effective_batch,
        "steps": [
            {
                "id": step.id,
                "sequence": step.sequence,
                "stage": step.stage,
                "operation": step.operation,
                "key_params": step.key_params,
                "owner": step.owner,
                "status": step.status,
            }
            for step in sorted(row.steps, key=lambda item: item.sequence)
        ],
    }


def ensure_route_editable(route: models.ProcessRoute) -> None:
    if route.status in ("已发布", "审批中"):
        raise HTTPException(status_code=409, detail=f"Process route in {route.status} status cannot be modified")


def validate_process_route_ready(route: models.ProcessRoute) -> None:
    if not route.steps:
        raise HTTPException(status_code=409, detail="Process route has no steps")
    sequences = [step.sequence for step in route.steps]
    if len(sequences) != len(set(sequences)):
        raise HTTPException(status_code=409, detail="Process route has duplicate step sequence")
    for step in route.steps:
        if not step.stage.strip() or not step.operation.strip() or not step.key_params.strip():
            raise HTTPException(status_code=409, detail="Process route step is incomplete")


def apply_bom_item_process_binding(db: Session, payload: BomItemPayload | BomItemUpdatePayload, product_id: int) -> dict:
    data = payload.model_dump(exclude_unset=True)
    step_id = data.get("process_step_id")
    if step_id:
        step = (
            db.query(models.ProcessStep)
            .join(models.ProcessRoute)
            .filter(models.ProcessStep.id == step_id, models.ProcessRoute.product_id == product_id)
            .first()
        )
        if not step:
            raise HTTPException(status_code=404, detail="Process step not found for this product")
        data["process_step"] = f"{step.sequence}-{step.stage}"
    return data


def create_integration_job(
    db: Session,
    target_system: str,
    object_type: str,
    object_no: str,
    product_model: str,
    triggered_by: str,
    message: str,
) -> None:
    count = db.query(models.IntegrationJob).count() + 1
    db.add(models.IntegrationJob(
        job_no=f"INT-{today_text().replace('-', '')}-{count:04d}",
        target_system=target_system,
        object_type=object_type,
        object_no=object_no,
        product_model=product_model,
        direction="下发",
        status="等待",
        triggered_by=triggered_by,
        triggered_at=today_text(),
        message=message,
    ))
    db.flush()


def serialize_integration_job(row: models.IntegrationJob) -> dict:
    return {
        "id": row.id,
        "job_no": row.job_no,
        "target_system": row.target_system,
        "object_type": row.object_type,
        "object_no": row.object_no,
        "product_model": row.product_model,
        "direction": row.direction,
        "status": row.status,
        "triggered_by": row.triggered_by,
        "triggered_at": row.triggered_at,
        "message": row.message,
        "attempt_count": row.attempt_count,
        "last_sync_at": row.last_sync_at,
        "response_message": row.response_message,
        "external_id": row.external_id,
    }


def serialize_change_action(row: models.ChangeAction) -> dict:
    return {
        "id": row.id,
        "action_no": row.action_no,
        "action_type": row.action_type,
        "target_type": row.target_type,
        "target_id": row.target_id,
        "target_version": row.target_version,
        "target_object": row.target_object,
        "effectivity_type": row.effectivity_type,
        "effectivity_scope": row.effectivity_scope,
        "effective_date": row.effective_date,
        "effective_batch": row.effective_batch,
        "generated_object_no": row.generated_object_no,
        "department": row.department,
        "owner": row.owner,
        "status": row.status,
        "due_date": row.due_date,
        "result": row.result,
    }


def serialize_change(row: models.Change, db: Session) -> dict:
    return {
        "id": row.id,
        "product_id": row.product_id,
        "change_no": row.change_no,
        "title": row.title,
        "product_model": row.product.model if row.product else "",
        "change_type": row.change_type,
        "reason": row.reason,
        "status": row.status,
        "priority": row.priority,
        "owner": row.owner,
        "submitted_at": row.submitted_at,
        "before_desc": row.before_desc,
        "after_desc": row.after_desc,
        "impacts": [{"id": impact.id, "type": impact.impact_type, "impact_type": impact.impact_type, "target": impact.target, "risk": impact.risk, "action": impact.action} for impact in row.impacts],
        "approvals": [{"step": approval.step_name, "approver": approval.approver, "status": approval.status, "comment": approval.comment, "approved_at": approval.approved_at} for approval in row.approvals],
        "actions": [
            serialize_change_action(action)
            for action in db.query(models.ChangeAction).filter(models.ChangeAction.change_id == row.id).order_by(models.ChangeAction.id).all()
        ],
    }


def analyze_change_impacts(db: Session, change: models.Change) -> None:
    existing_impacts = db.query(models.ChangeImpact).filter(models.ChangeImpact.change_id == change.id).count()
    existing_actions = db.query(models.ChangeAction).filter(models.ChangeAction.change_id == change.id).count()
    if existing_impacts or existing_actions:
        return

    product = change.product or db.query(models.Product).filter(models.Product.id == change.product_id).first()
    product_model = product.model if product else ""
    boms = (
        db.query(models.BomHeader)
        .filter(models.BomHeader.product_id == change.product_id)
        .order_by(models.BomHeader.bom_type, models.BomHeader.version.desc())
        .all()
    )
    routes = (
        db.query(models.ProcessRoute)
        .filter(models.ProcessRoute.product_id == change.product_id)
        .order_by(models.ProcessRoute.version.desc())
        .all()
    )
    documents = (
        db.query(models.Document)
        .filter(models.Document.product_id == change.product_id)
        .order_by(models.Document.category, models.Document.version.desc())
        .all()
    )

    impacts: list[tuple[str, str, str, str]] = [("产品", product_model, "中", "确认型号主数据、生命周期和版本基线是否需要更新。")]
    impacts.extend((f"{bom.bom_type} BOM", f"{bom.bom_type}-{product_model}-{bom.version}", "高", "评估物料结构、用量、替代料和有效期。") for bom in boms[:4])
    impacts.extend(("工艺路线", route.route_no, "高", "评估工序、工艺参数、站点控制和 MES 下发内容。") for route in routes[:3])
    impacts.extend(("文档", f"{doc.doc_no} / {doc.title}", "中", "评估规格书、流程卡、检验规范和文控发布。") for doc in documents[:4])
    for impact_type, target, risk, action in impacts:
        db.add(models.ChangeImpact(change_id=change.id, impact_type=impact_type, target=target, risk=risk, action=action))

    action_defs: list[dict] = []
    if boms:
        bom = boms[0]
        action_defs.append({
            "action_type": "BOM升版",
            "target_type": "BOM",
            "target_id": bom.id,
            "target_version": bom.version,
            "target_object": f"{bom.bom_type}-{product_model}-{bom.version}",
            "effectivity_type": "日期+批次",
            "effectivity_scope": "版本升版",
            "effective_date": today_text(),
            "effective_batch": f"LOT-{product_model}-2606",
            "department": "生产部",
            "owner": change.owner or "罗富森",
            "result": "根据影响分析生成下一版 BOM 草稿并维护有效期。",
        })
    if routes:
        route = routes[0]
        action_defs.append({
            "action_type": "工艺更新",
            "target_type": "工艺路线",
            "target_id": route.id,
            "target_version": route.version,
            "target_object": route.route_no,
            "effectivity_type": "日期+批次",
            "effectivity_scope": "工艺切换",
            "effective_date": today_text(),
            "effective_batch": f"LOT-{product_model}-2606",
            "department": "生产部",
            "owner": "罗富森",
            "result": "更新工艺路线、工序参数和站点控制。",
        })
    if documents:
        document = documents[0]
        action_defs.append({
            "action_type": "文档发布",
            "target_type": "文档",
            "target_id": document.id,
            "target_version": document.version,
            "target_object": f"{document.doc_no} / {document.title}",
            "effectivity_type": "日期",
            "effectivity_scope": "文档升版",
            "effective_date": today_text(),
            "effective_batch": "",
            "department": "生产部",
            "owner": "于帅兵",
            "result": "生成受影响文档下一版草稿并完成文控发布准备。",
        })
    action_defs.append({
        "action_type": "ECN通知",
        "target_type": "通知",
        "target_id": None,
        "target_version": "",
        "target_object": f"{change.change_no} 变更通知",
        "effectivity_type": "批次",
        "effectivity_scope": "签收跟踪",
        "effective_date": "",
        "effective_batch": f"LOT-{product_model}-2606",
        "department": "生产部",
        "owner": change.owner or "房磊",
        "result": "通知受影响部门并跟踪签收结果。",
    })
    for index, action_data in enumerate(action_defs, start=1):
        db.add(models.ChangeAction(
            change_id=change.id,
            action_no=f"ECA-{change.change_no}-{index:02d}",
            action_type=action_data["action_type"],
            target_type=action_data["target_type"],
            target_id=action_data["target_id"],
            target_version=action_data["target_version"],
            target_object=action_data["target_object"],
            effectivity_type=action_data["effectivity_type"],
            effectivity_scope=action_data["effectivity_scope"],
            effective_date=action_data["effective_date"],
            effective_batch=action_data["effective_batch"],
            department=action_data["department"],
            owner=action_data["owner"],
            status="待处理",
            due_date=today_text(),
            result=action_data["result"],
        ))


def create_change_release_jobs(db: Session, change: models.Change) -> None:
    existing_targets = {
        row.target_system
        for row in db.query(models.IntegrationJob)
        .filter(models.IntegrationJob.object_type == "ECN", models.IntegrationJob.object_no == change.change_no)
        .all()
    }
    if "ERP" not in existing_targets:
        create_integration_job(db, "ERP", "ECN", change.change_no, change.product.model, change.change_no, "工程变更执行完成，等待同步 ERP 物料/BOM 变更通知。")
    if "MES" not in existing_targets:
        create_integration_job(db, "MES", "ECN", change.change_no, change.product.model, change.change_no, "工程变更执行完成，等待同步 MES 工艺和制造切换要求。")


def validate_action_effectivity(action: models.ChangeAction) -> None:
    effectivity_type = action.effectivity_type or "日期"
    if "日期" in effectivity_type and not action.effective_date:
        raise HTTPException(status_code=409, detail="Date-based ECA action requires effective date")
    if "批次" in effectivity_type and not action.effective_batch:
        raise HTTPException(status_code=409, detail="Batch-based ECA action requires effective batch")
    if action.effective_date:
        try:
            date.fromisoformat(action.effective_date)
        except ValueError as exc:
            raise HTTPException(status_code=409, detail="Effective date must use YYYY-MM-DD") from exc


def validate_change_action_target(db: Session, action: models.ChangeAction) -> None:
    validate_action_effectivity(action)
    if action.target_type in {"BOM", "文档", "工艺路线"} and not action.target_id:
        raise HTTPException(status_code=409, detail="Revision ECA action requires target object")

    if action.target_type == "BOM" and action.target_id:
        source = (
            db.query(models.BomHeader)
            .options(selectinload(models.BomHeader.items))
            .filter(models.BomHeader.id == action.target_id)
            .first()
        )
        if not source:
            raise HTTPException(status_code=404, detail="Target BOM not found")
        if action.target_version and action.target_version != source.version:
            raise HTTPException(status_code=409, detail="ECA target version does not match current BOM version")
        if source.status != "已发布":
            raise HTTPException(status_code=409, detail="Only released BOM can be revised by ECA")
        if not source.items:
            raise HTTPException(status_code=409, detail="Target BOM has no items")
        return

    if action.target_type == "文档" and action.target_id:
        source = db.query(models.Document).filter(models.Document.id == action.target_id).first()
        if not source:
            raise HTTPException(status_code=404, detail="Target document not found")
        if action.target_version and action.target_version != source.version:
            raise HTTPException(status_code=409, detail="ECA target version does not match current document version")
        if source.status == "已废止":
            raise HTTPException(status_code=409, detail="Obsolete document cannot be revised by ECA")
        return

    if action.target_type == "工艺路线" and action.target_id:
        source = (
            db.query(models.ProcessRoute)
            .options(selectinload(models.ProcessRoute.steps))
            .filter(models.ProcessRoute.id == action.target_id)
            .first()
        )
        if not source:
            raise HTTPException(status_code=404, detail="Target process route not found")
        if action.target_version and action.target_version != source.version:
            raise HTTPException(status_code=409, detail="ECA target version does not match current process route version")
        if source.status not in {"已发布", "有效"}:
            raise HTTPException(status_code=409, detail="Only released process route can be revised by ECA")
        validate_process_route_ready(source)


def get_eca_generated_object_gate(db: Session, object_type: str, generated_object_no: str) -> dict | None:
    action = (
        db.query(models.ChangeAction)
        .filter(
            models.ChangeAction.target_type == object_type,
            models.ChangeAction.generated_object_no == generated_object_no,
        )
        .first()
    )
    if not action:
        return None
    change = db.query(models.Change).filter(models.Change.id == action.change_id).first()
    if action.status != "已完成":
        return {
            "ready": False,
            "action_no": action.action_no,
            "action_status": action.status,
            "change_no": change.change_no if change else "",
            "change_status": change.status if change else "",
            "message": f"ECA 动作 {action.action_no} 尚未关闭，生成对象不能提交发布。",
        }
    if change and change.status != "已关闭":
        return {
            "ready": False,
            "action_no": action.action_no,
            "action_status": action.status,
            "change_no": change.change_no,
            "change_status": change.status,
            "message": f"变更单 {change.change_no} 尚未关闭，需等待全部 ECA 动作完成后再提交发布生成对象。",
        }
    return {
        "ready": True,
        "action_no": action.action_no,
        "action_status": action.status,
        "change_no": change.change_no if change else "",
        "change_status": change.status if change else "",
        "message": "生成对象已满足提交发布条件。",
    }


def validate_eca_generated_object_ready(db: Session, object_type: str, object_id: int, generated_object_no: str) -> None:
    """校验 ECA 生成的对象草稿在提交/发布前已完成整张变更闭环。"""
    gate = get_eca_generated_object_gate(db, object_type, generated_object_no)
    if gate and not gate["ready"]:
        raise HTTPException(
            status_code=409,
            detail=gate["message"],
        )



def apply_change_action_revision(db: Session, action: models.ChangeAction) -> str:
    if action.generated_object_no:
        return action.generated_object_no

    validate_change_action_target(db, action)
    effective_date = action.effective_date or today_text()
    if action.target_type == "BOM" and action.target_id:
        source = db.query(models.BomHeader).options(selectinload(models.BomHeader.items), selectinload(models.BomHeader.product)).filter(models.BomHeader.id == action.target_id).first()
        if not source:
            return ""
        new_version = next_unique_bom_version(db, source)
        new_bom = models.BomHeader(
            product_id=source.product_id,
            bom_type=source.bom_type,
            version=new_version,
            status="编制中",
            owner=action.owner or source.owner,
            release_date="",
            source_bom_id=source.id,
            effective_date=effective_date,
            expiry_date="",
            effectivity_type=action.effectivity_type or "变更生效",
            effective_batch=action.effective_batch or "",
        )
        db.add(new_bom)
        db.flush()
        item_id_map: dict[int, int] = {}
        for source_item in sorted(source.items, key=lambda item: item.id):
            new_item = models.BomItem(
                bom_id=new_bom.id,
                parent_id=item_id_map.get(source_item.parent_id),
                material_code=source_item.material_code,
                material_name=source_item.material_name,
                category=source_item.category,
                specification=source_item.specification,
                quantity=source_item.quantity,
                unit=source_item.unit,
                position=source_item.position,
                process_step_id=source_item.process_step_id,
                process_step=source_item.process_step,
                substitute=source_item.substitute,
                status=source_item.status,
                effective_date=source_item.effective_date or effective_date,
                expiry_date=source_item.expiry_date,
                effectivity_note=source_item.effectivity_note or action.effective_batch,
            )
            db.add(new_item)
            db.flush()
            item_id_map[source_item.id] = new_item.id
        product_model = source.product.model if source.product else str(source.product_id)
        action.generated_object_no = f"{source.bom_type}-{product_model}-{new_version}"
        action.target_version = source.version
        return action.generated_object_no

    if action.target_type == "文档" and action.target_id:
        source = db.query(models.Document).filter(models.Document.id == action.target_id).first()
        if not source:
            return ""
        new_version = next_revision(source.version)
        new_doc_no = next_unique_document_no(db, source, new_version)
        db.add(models.Document(
            product_id=source.product_id,
            doc_no=new_doc_no,
            title=source.title,
            category=source.category,
            version=new_version,
            status="编制中",
            owner=action.owner or source.owner,
            approval_status="未提交",
            updated_at=effective_date,
        ))
        db.flush()
        action.generated_object_no = new_doc_no
        action.target_version = source.version
        return action.generated_object_no

    if action.target_type == "工艺路线" and action.target_id:
        source = (
            db.query(models.ProcessRoute)
            .options(selectinload(models.ProcessRoute.steps), selectinload(models.ProcessRoute.product))
            .filter(models.ProcessRoute.id == action.target_id)
            .first()
        )
        if not source:
            return ""
        new_version = next_unique_process_version(db, source)
        new_route_no = next_unique_route_no(db, source, new_version)
        new_route = models.ProcessRoute(
            product_id=source.product_id,
            route_no=new_route_no,
            name=source.name,
            version=new_version,
            status="编制中",
            owner=action.owner or source.owner,
            release_date="",
            source_route_id=source.id,
            effective_batch=action.effective_batch or "",
        )
        db.add(new_route)
        db.flush()
        for source_step in sorted(source.steps, key=lambda item: item.sequence):
            db.add(models.ProcessStep(
                route_id=new_route.id,
                sequence=source_step.sequence,
                stage=source_step.stage,
                operation=source_step.operation,
                key_params=source_step.key_params,
                owner=action.owner or source_step.owner,
                status=source_step.status,
            ))
        db.flush()
        action.generated_object_no = new_route_no
        action.target_version = source.version
        return action.generated_object_no

    return ""


def close_change_when_actions_done(db: Session, change_id: int, acted_by: str) -> bool:
    change = db.query(models.Change).options(selectinload(models.Change.product)).filter(models.Change.id == change_id).first()
    if not change or change.status == "已关闭":
        return False
    actions = db.query(models.ChangeAction).filter(models.ChangeAction.change_id == change_id).all()
    if not actions or any(action.status != "已完成" for action in actions):
        return False
    change.status = "已关闭"
    create_change_release_jobs(db, change)
    instance = (
        db.query(models.WorkflowInstance)
        .filter(models.WorkflowInstance.object_type == "变更", models.WorkflowInstance.object_id == change_id)
        .order_by(models.WorkflowInstance.id.desc())
        .first()
    )
    if instance:
        add_workflow_log(db, instance.id, None, "变更关闭", acted_by, "全部 ECA 执行动作已关闭", "执行中", "已关闭")
    return True


def add_workflow_log(
    db: Session,
    instance_id: int,
    task_id: int | None,
    action: str,
    actor: str,
    comment: str,
    from_status: str,
    to_status: str,
) -> None:
    db.add(models.WorkflowLog(
        instance_id=instance_id,
        task_id=task_id,
        action=action,
        actor=actor,
        acted_at=today_text(),
        comment=comment,
        from_status=from_status,
        to_status=to_status,
    ))


def serialize_workflow_instance(row: models.WorkflowInstance) -> dict:
    return {
        "id": row.id,
        "template_id": row.template_id,
        "template_name": row.template.name if row.template else "",
        "object_type": row.object_type,
        "object_id": row.object_id,
        "object_no": row.object_no,
        "title": row.title,
        "product_model": row.product_model,
        "status": row.status,
        "started_by": row.started_by,
        "started_at": row.started_at,
        "completed_at": row.completed_at,
        "tasks": [
            {
                "id": task.id,
                "sequence": task.sequence,
                "node_name": task.node_name,
                "role_name": task.role_name,
                "action_type": task.action_type,
                "status": task.status,
                "assignee": task.assignee,
                "acted_by": task.acted_by,
                "acted_at": task.acted_at,
                "comment": task.comment,
                "sla_hours": task.sla_hours,
            }
            for task in sorted(row.tasks, key=lambda item: item.sequence)
        ],
        "logs": [
            {
                "id": log.id,
                "task_id": log.task_id,
                "action": log.action,
                "actor": log.actor,
                "acted_at": log.acted_at,
                "comment": log.comment,
                "from_status": log.from_status,
                "to_status": log.to_status,
            }
            for log in sorted(getattr(row, "logs", []), key=lambda item: item.id)
        ],
    }


def serialize_coding_rule(row: models.CodingRule) -> dict:
    return {
        "id": row.id,
        "object_type": row.object_type,
        "code": row.code,
        "name": row.name,
        "prefix": row.prefix,
        "pattern": row.pattern,
        "current_no": row.current_no,
        "sample": row.sample,
        "status": row.status,
        "owner": row.owner,
    }


def serialize_category_template(row: models.CategoryTemplate) -> dict:
    return {
        "id": row.id,
        "object_type": row.object_type,
        "code": row.code,
        "name": row.name,
        "parent_code": row.parent_code,
        "lifecycle_template": row.lifecycle_template,
        "coding_rule": row.coding_rule,
        "status": row.status,
        "description": row.description,
        "attributes": [
            {
                "id": item.id,
                "name": item.name,
                "field_key": item.field_key,
                "data_type": item.data_type,
                "required": item.required,
                "dictionary_code": item.dictionary_code,
                "default_value": item.default_value,
                "sequence": item.sequence,
            }
            for item in sorted(row.attributes, key=lambda attr: attr.sequence)
        ],
    }


def serialize_lifecycle_template(row: models.LifecycleTemplate) -> dict:
    return {
        "id": row.id,
        "code": row.code,
        "name": row.name,
        "object_type": row.object_type,
        "status": row.status,
        "description": row.description,
        "states": [
            {
                "id": item.id,
                "sequence": item.sequence,
                "name": item.name,
                "state_type": item.state_type,
                "allow_edit": item.allow_edit,
                "require_workflow": item.require_workflow,
                "next_states": item.next_states,
            }
            for item in sorted(row.states, key=lambda state: state.sequence)
        ],
    }


def serialize_dictionary_item(row: models.DictionaryItem) -> dict:
    return {
        "id": row.id,
        "dict_code": row.dict_code,
        "dict_name": row.dict_name,
        "item_value": row.item_value,
        "item_label": row.item_label,
        "object_scope": row.object_scope,
        "sequence": row.sequence,
        "status": row.status,
    }


def start_workflow(
    db: Session,
    template_code: str,
    object_type: str,
    object_id: int,
    object_no: str,
    title: str,
    product_model: str,
    started_by: str,
) -> models.WorkflowInstance:
    existing = (
        db.query(models.WorkflowInstance)
        .filter(
            models.WorkflowInstance.object_type == object_type,
            models.WorkflowInstance.object_id == object_id,
            models.WorkflowInstance.status == "运行中",
        )
        .first()
    )
    if existing:
        return existing
    template = (
        db.query(models.WorkflowTemplate)
        .options(selectinload(models.WorkflowTemplate.nodes))
        .filter(models.WorkflowTemplate.code == template_code, models.WorkflowTemplate.status == "启用")
        .first()
    )
    if not template:
        raise HTTPException(status_code=409, detail=f"Workflow template {template_code} is not enabled")
    instance = models.WorkflowInstance(
        template_id=template.id,
        object_type=object_type,
        object_id=object_id,
        object_no=object_no,
        title=title,
        product_model=product_model,
        status="运行中",
        started_by=started_by,
        started_at=today_text(),
    )
    db.add(instance)
    db.flush()
    for index, node in enumerate(sorted(template.nodes, key=lambda item: item.sequence)):
        db.add(models.WorkflowTask(
            instance_id=instance.id,
            sequence=node.sequence,
            node_name=node.name,
            role_name=node.role_name,
            action_type=node.action_type,
            status="待处理" if index == 0 else "未开始",
            sla_hours=node.sla_hours,
        ))
    return instance


def complete_business_object(db: Session, instance: models.WorkflowInstance) -> None:
    instance.status = "已完成"
    instance.completed_at = today_text()
    if instance.object_type == "BOM":
        bom = db.query(models.BomHeader).options(selectinload(models.BomHeader.product)).filter(models.BomHeader.id == instance.object_id).first()
        if bom:
            bom.status = "已发布"
            bom.release_date = today_text()
            create_integration_job(db, "ERP", "BOM", f"{bom.bom_type}-{bom.product.model}-{bom.version}", bom.product.model, instance.object_no, "BOM 流程已完成，等待同步 ERP 物料结构和用量。")
            auto_complete_project_deliverable(db, "BOM", bom.id)
    elif instance.object_type == "文档":
        document = db.query(models.Document).options(selectinload(models.Document.product)).filter(models.Document.id == instance.object_id).first()
        if document:
            document.status = "已发布"
            document.approval_status = "已签核"
            document.updated_at = today_text()
            create_integration_job(db, "QMS", "文档", document.doc_no, document.product.model, document.doc_no, "文档流程已完成，等待同步 QMS/文控归档。")
            auto_complete_project_deliverable(db, "文档", document.id)
    elif instance.object_type == "变更":
        change = db.query(models.Change).options(selectinload(models.Change.product)).filter(models.Change.id == instance.object_id).first()
        if change:
            change.status = "执行中"
            if not change.submitted_at:
                change.submitted_at = today_text()
            close_change_when_actions_done(db, change.id, "系统")


def auto_complete_project_deliverable(db: Session, object_type: str, object_id: int) -> None:
    """对象审批发布后，自动完成关联的项目交付物。"""
    deliverables = (
        db.query(models.ProjectDeliverable)
        .filter(
            models.ProjectDeliverable.object_type == object_type,
            models.ProjectDeliverable.object_id == object_id,
            models.ProjectDeliverable.status.notin_(["已完成", "已关闭"]),
        )
        .all()
    )
    for d in deliverables:
        d.status = "已完成"
        d.completed_at = today_text()


def withdraw_business_object(db: Session, instance: models.WorkflowInstance) -> None:
    if instance.object_type == "BOM":
        bom = db.query(models.BomHeader).filter(models.BomHeader.id == instance.object_id).first()
        if bom and bom.status != "已发布":
            bom.status = "编制中"
    elif instance.object_type == "文档":
        document = db.query(models.Document).filter(models.Document.id == instance.object_id).first()
        if document and document.status != "已发布":
            document.status = "编制中"
            document.approval_status = "未提交"
            document.updated_at = today_text()
    elif instance.object_type == "变更":
        change = db.query(models.Change).filter(models.Change.id == instance.object_id).first()
        if change and change.status != "已关闭":
            change.status = "草稿"


@app.get("/api/health")
def health() -> dict:
    return {"status": "ok"}


@app.get("/api/session/current")
def session_current(context: dict = Depends(current_user_context)) -> dict:
    user = context["user"]
    role = context["role"]
    permissions = context["permissions"]
    return {
        "user": {
            "id": user.id,
            "username": user.username,
            "display_name": user.display_name,
            "role": user.role,
            "department": user.department,
        },
        "role": {
            "id": role.id if role else None,
            "code": role.code if role else "",
            "name": role.name if role else user.role,
            "description": role.description if role else "",
            "status": role.status if role else "未配置",
        },
        "permissions": permissions,
        "permission_labels": {key: PERMISSION_LABELS.get(key, key) for key in permissions},
    }


@app.get("/api/admin/roles")
def roles(page: int = 1, page_size: int = 20, keyword: str = "", db: Session = Depends(get_db)) -> dict:
    q = db.query(models.Role)
    if keyword:
        kw = f"%{keyword}%"
        q = q.filter(models.Role.code.ilike(kw) | models.Role.name.ilike(kw))
    total = q.count()
    rows = q.order_by(models.Role.id).offset((page - 1) * page_size).limit(page_size).all()
    return {
        "items": [
            {"id": row.id, "code": row.code, "name": row.name, "description": row.description, "permissions": row.permissions, "status": row.status}
            for row in rows
        ],
        "total": total,
        "page": page,
        "page_size": page_size,
    }


@app.get("/api/admin/organizations")
def organizations(page: int = 1, page_size: int = 20, keyword: str = "", db: Session = Depends(get_db)) -> dict:
    q = db.query(models.Organization)
    if keyword:
        kw = f"%{keyword}%"
        q = q.filter(models.Organization.code.ilike(kw) | models.Organization.name.ilike(kw))
    total = q.count()
    rows = q.order_by(models.Organization.org_type, models.Organization.id).offset((page - 1) * page_size).limit(page_size).all()
    return {
        "items": [
            {"id": row.id, "code": row.code, "name": row.name, "org_type": row.org_type, "parent_code": row.parent_code, "manager": row.manager, "status": row.status, "description": row.description}
            for row in rows
        ],
        "total": total,
        "page": page,
        "page_size": page_size,
    }


@app.post("/api/admin/organizations", status_code=201)
def create_organization(payload: OrganizationPayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("organization"))) -> dict:
    row = models.Organization(**payload.model_dump())
    db.add(row)
    commit_or_409(db, "Organization code already exists")
    db.refresh(row)
    return {"id": row.id, "code": row.code, "name": row.name, "org_type": row.org_type, "parent_code": row.parent_code, "manager": row.manager, "status": row.status, "description": row.description}


@app.put("/api/admin/organizations/{org_id}")
def update_organization(org_id: int, payload: OrganizationUpdatePayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("organization"))) -> dict:
    row = db.query(models.Organization).filter(models.Organization.id == org_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Organization not found")
    update_model(row, payload)
    commit_or_409(db, "Organization code already exists")
    db.refresh(row)
    return {"id": row.id, "code": row.code, "name": row.name, "org_type": row.org_type, "parent_code": row.parent_code, "manager": row.manager, "status": row.status, "description": row.description}


@app.delete("/api/admin/organizations/{org_id}")
def delete_organization(org_id: int, db: Session = Depends(get_db), _: dict = Depends(require_permission("organization"))) -> dict:
    row = db.query(models.Organization).filter(models.Organization.id == org_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Organization not found")
    if row.code in {"NZGD", "PROD"}:
        raise HTTPException(status_code=409, detail="Built-in single organization cannot be deleted")
    db.delete(row)
    db.commit()
    return {"deleted": True}


@app.post("/api/admin/roles", status_code=201)
def create_role(payload: RolePayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("role"))) -> dict:
    role = models.Role(**payload.model_dump())
    db.add(role)
    commit_or_409(db, "Role code already exists")
    db.refresh(role)
    return {"id": role.id, "code": role.code, "name": role.name, "description": role.description, "permissions": role.permissions, "status": role.status}


@app.put("/api/admin/roles/{role_id}")
def update_role(role_id: int, payload: RoleUpdatePayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("role"))) -> dict:
    role = db.query(models.Role).filter(models.Role.id == role_id).first()
    if not role:
        raise HTTPException(status_code=404, detail="Role not found")
    update_model(role, payload)
    commit_or_409(db, "Role code already exists")
    db.refresh(role)
    return {"id": role.id, "code": role.code, "name": role.name, "description": role.description, "permissions": role.permissions, "status": role.status}


@app.delete("/api/admin/roles/{role_id}")
def delete_role(role_id: int, db: Session = Depends(get_db), _: dict = Depends(require_permission("role"))) -> dict:
    role = db.query(models.Role).filter(models.Role.id == role_id).first()
    if not role:
        raise HTTPException(status_code=404, detail="Role not found")
    if db.query(models.User).filter(models.User.role == role.name).count():
        raise HTTPException(status_code=409, detail="Role is used by users")
    db.delete(role)
    db.commit()
    return {"deleted": True}


@app.get("/api/admin/foundation/coding-rules")
def coding_rules(page: int = 1, page_size: int = 20, keyword: str = "", db: Session = Depends(get_db)) -> dict:
    q = db.query(models.CodingRule)
    if keyword:
        kw = f"%{keyword}%"
        q = q.filter(models.CodingRule.code.ilike(kw) | models.CodingRule.name.ilike(kw))
    total = q.count()
    rows = q.order_by(models.CodingRule.object_type, models.CodingRule.id).offset((page - 1) * page_size).limit(page_size).all()
    return {"items": [serialize_coding_rule(row) for row in rows], "total": total, "page": page, "page_size": page_size}


@app.post("/api/admin/foundation/coding-rules", status_code=201)
def create_coding_rule(payload: CodingRulePayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("system"))) -> dict:
    row = models.CodingRule(**payload.model_dump())
    db.add(row)
    commit_or_409(db, "Coding rule code already exists")
    db.refresh(row)
    return serialize_coding_rule(row)


@app.put("/api/admin/foundation/coding-rules/{rule_id}")
def update_coding_rule(rule_id: int, payload: CodingRuleUpdatePayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("system"))) -> dict:
    row = db.query(models.CodingRule).filter(models.CodingRule.id == rule_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Coding rule not found")
    update_model(row, payload)
    commit_or_409(db, "Coding rule code already exists")
    db.refresh(row)
    return serialize_coding_rule(row)


@app.delete("/api/admin/foundation/coding-rules/{rule_id}")
def delete_coding_rule(rule_id: int, db: Session = Depends(get_db), _: dict = Depends(require_permission("system"))) -> dict:
    row = db.query(models.CodingRule).filter(models.CodingRule.id == rule_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Coding rule not found")
    db.delete(row)
    db.commit()
    return {"deleted": True}


@app.get("/api/admin/foundation/categories")
def category_templates(page: int = 1, page_size: int = 20, keyword: str = "", db: Session = Depends(get_db)) -> dict:
    q = db.query(models.CategoryTemplate).options(selectinload(models.CategoryTemplate.attributes))
    if keyword:
        kw = f"%{keyword}%"
        q = q.filter(models.CategoryTemplate.code.ilike(kw) | models.CategoryTemplate.name.ilike(kw))
    total = q.count()
    rows = q.order_by(models.CategoryTemplate.object_type, models.CategoryTemplate.id).offset((page - 1) * page_size).limit(page_size).all()
    return {"items": [serialize_category_template(row) for row in rows], "total": total, "page": page, "page_size": page_size}


@app.post("/api/admin/foundation/categories", status_code=201)
def create_category_template(payload: CategoryTemplatePayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("system"))) -> dict:
    row = models.CategoryTemplate(**payload.model_dump())
    db.add(row)
    commit_or_409(db, "Category code already exists")
    db.refresh(row)
    return serialize_category_template(row)


@app.put("/api/admin/foundation/categories/{category_id}")
def update_category_template(category_id: int, payload: CategoryTemplateUpdatePayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("system"))) -> dict:
    row = (
        db.query(models.CategoryTemplate)
        .options(selectinload(models.CategoryTemplate.attributes))
        .filter(models.CategoryTemplate.id == category_id)
        .first()
    )
    if not row:
        raise HTTPException(status_code=404, detail="Category not found")
    update_model(row, payload)
    commit_or_409(db, "Category code already exists")
    db.refresh(row)
    return serialize_category_template(row)


@app.delete("/api/admin/foundation/categories/{category_id}")
def delete_category_template(category_id: int, db: Session = Depends(get_db), _: dict = Depends(require_permission("system"))) -> dict:
    row = db.query(models.CategoryTemplate).filter(models.CategoryTemplate.id == category_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Category not found")
    db.delete(row)
    db.commit()
    return {"deleted": True}


@app.post("/api/admin/foundation/categories/{category_id}/attributes", status_code=201)
def create_attribute_template(category_id: int, payload: AttributeTemplatePayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("system"))) -> dict:
    if not db.query(models.CategoryTemplate.id).filter(models.CategoryTemplate.id == category_id).first():
        raise HTTPException(status_code=404, detail="Category not found")
    row = models.AttributeTemplate(category_id=category_id, **payload.model_dump())
    db.add(row)
    db.commit()
    db.refresh(row)
    return {"id": row.id, "name": row.name, "field_key": row.field_key, "data_type": row.data_type, "required": row.required, "dictionary_code": row.dictionary_code, "default_value": row.default_value, "sequence": row.sequence}


@app.put("/api/admin/foundation/attributes/{attribute_id}")
def update_attribute_template(attribute_id: int, payload: AttributeTemplateUpdatePayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("system"))) -> dict:
    row = db.query(models.AttributeTemplate).filter(models.AttributeTemplate.id == attribute_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Attribute not found")
    update_model(row, payload)
    db.commit()
    db.refresh(row)
    return {"id": row.id, "name": row.name, "field_key": row.field_key, "data_type": row.data_type, "required": row.required, "dictionary_code": row.dictionary_code, "default_value": row.default_value, "sequence": row.sequence}


@app.delete("/api/admin/foundation/attributes/{attribute_id}")
def delete_attribute_template(attribute_id: int, db: Session = Depends(get_db), _: dict = Depends(require_permission("system"))) -> dict:
    row = db.query(models.AttributeTemplate).filter(models.AttributeTemplate.id == attribute_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Attribute not found")
    db.delete(row)
    db.commit()
    return {"deleted": True}


@app.get("/api/admin/foundation/lifecycles")
def lifecycle_templates(page: int = 1, page_size: int = 20, keyword: str = "", db: Session = Depends(get_db)) -> dict:
    q = db.query(models.LifecycleTemplate).options(selectinload(models.LifecycleTemplate.states))
    if keyword:
        kw = f"%{keyword}%"
        q = q.filter(models.LifecycleTemplate.code.ilike(kw) | models.LifecycleTemplate.name.ilike(kw))
    total = q.count()
    rows = q.order_by(models.LifecycleTemplate.object_type, models.LifecycleTemplate.id).offset((page - 1) * page_size).limit(page_size).all()
    return {"items": [serialize_lifecycle_template(row) for row in rows], "total": total, "page": page, "page_size": page_size}


@app.post("/api/admin/foundation/lifecycles", status_code=201)
def create_lifecycle_template(payload: LifecycleTemplatePayload, db: Session = Depends(get_db)) -> dict:
    row = models.LifecycleTemplate(**payload.model_dump())
    db.add(row)
    commit_or_409(db, "Lifecycle code already exists")
    db.refresh(row)
    return serialize_lifecycle_template(row)


@app.put("/api/admin/foundation/lifecycles/{template_id}")
def update_lifecycle_template(template_id: int, payload: LifecycleTemplateUpdatePayload, db: Session = Depends(get_db)) -> dict:
    row = (
        db.query(models.LifecycleTemplate)
        .options(selectinload(models.LifecycleTemplate.states))
        .filter(models.LifecycleTemplate.id == template_id)
        .first()
    )
    if not row:
        raise HTTPException(status_code=404, detail="Lifecycle not found")
    update_model(row, payload)
    commit_or_409(db, "Lifecycle code already exists")
    db.refresh(row)
    return serialize_lifecycle_template(row)


@app.delete("/api/admin/foundation/lifecycles/{template_id}")
def delete_lifecycle_template(template_id: int, db: Session = Depends(get_db)) -> dict:
    row = db.query(models.LifecycleTemplate).filter(models.LifecycleTemplate.id == template_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Lifecycle not found")
    db.delete(row)
    db.commit()
    return {"deleted": True}


@app.post("/api/admin/foundation/lifecycles/{template_id}/states", status_code=201)
def create_lifecycle_state(template_id: int, payload: LifecycleStatePayload, db: Session = Depends(get_db)) -> dict:
    if not db.query(models.LifecycleTemplate.id).filter(models.LifecycleTemplate.id == template_id).first():
        raise HTTPException(status_code=404, detail="Lifecycle not found")
    row = models.LifecycleState(template_id=template_id, **payload.model_dump())
    db.add(row)
    db.commit()
    db.refresh(row)
    return {"id": row.id, "sequence": row.sequence, "name": row.name, "state_type": row.state_type, "allow_edit": row.allow_edit, "require_workflow": row.require_workflow, "next_states": row.next_states}


@app.put("/api/admin/foundation/lifecycle-states/{state_id}")
def update_lifecycle_state(state_id: int, payload: LifecycleStateUpdatePayload, db: Session = Depends(get_db)) -> dict:
    row = db.query(models.LifecycleState).filter(models.LifecycleState.id == state_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Lifecycle state not found")
    update_model(row, payload)
    db.commit()
    db.refresh(row)
    return {"id": row.id, "sequence": row.sequence, "name": row.name, "state_type": row.state_type, "allow_edit": row.allow_edit, "require_workflow": row.require_workflow, "next_states": row.next_states}


@app.delete("/api/admin/foundation/lifecycle-states/{state_id}")
def delete_lifecycle_state(state_id: int, db: Session = Depends(get_db)) -> dict:
    row = db.query(models.LifecycleState).filter(models.LifecycleState.id == state_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Lifecycle state not found")
    db.delete(row)
    db.commit()
    return {"deleted": True}


@app.get("/api/admin/foundation/dictionaries")
def dictionary_items(page: int = 1, page_size: int = 20, keyword: str = "", db: Session = Depends(get_db)) -> dict:
    q = db.query(models.DictionaryItem)
    if keyword:
        kw = f"%{keyword}%"
        q = q.filter(
            models.DictionaryItem.dict_code.ilike(kw)
            | models.DictionaryItem.dict_name.ilike(kw)
            | models.DictionaryItem.item_value.ilike(kw)
            | models.DictionaryItem.item_label.ilike(kw)
        )
    total = q.count()
    rows = q.order_by(models.DictionaryItem.dict_code, models.DictionaryItem.sequence, models.DictionaryItem.id).offset((page - 1) * page_size).limit(page_size).all()
    return {"items": [serialize_dictionary_item(row) for row in rows], "total": total, "page": page, "page_size": page_size}


@app.post("/api/admin/foundation/dictionaries", status_code=201)
def create_dictionary_item(payload: DictionaryItemPayload, db: Session = Depends(get_db)) -> dict:
    row = models.DictionaryItem(**payload.model_dump())
    db.add(row)
    db.commit()
    db.refresh(row)
    return serialize_dictionary_item(row)


@app.put("/api/admin/foundation/dictionaries/{item_id}")
def update_dictionary_item(item_id: int, payload: DictionaryItemUpdatePayload, db: Session = Depends(get_db)) -> dict:
    row = db.query(models.DictionaryItem).filter(models.DictionaryItem.id == item_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Dictionary item not found")
    update_model(row, payload)
    db.commit()
    db.refresh(row)
    return serialize_dictionary_item(row)


@app.delete("/api/admin/foundation/dictionaries/{item_id}")
def delete_dictionary_item(item_id: int, db: Session = Depends(get_db)) -> dict:
    row = db.query(models.DictionaryItem).filter(models.DictionaryItem.id == item_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Dictionary item not found")
    db.delete(row)
    db.commit()
    return {"deleted": True}


@app.get("/api/admin/users")
def users(page: int = 1, page_size: int = 20, keyword: str = "", db: Session = Depends(get_db)) -> dict:
    q = db.query(models.User)
    if keyword:
        kw = f"%{keyword}%"
        q = q.filter(
            models.User.username.ilike(kw)
            | models.User.display_name.ilike(kw)
            | models.User.role.ilike(kw)
        )
    total = q.count()
    rows = q.order_by(models.User.id).offset((page - 1) * page_size).limit(page_size).all()
    return {
        "items": [
            {"id": row.id, "username": row.username, "display_name": row.display_name, "role": row.role, "department": row.department}
            for row in rows
        ],
        "total": total,
        "page": page,
        "page_size": page_size,
    }


@app.post("/api/admin/users", status_code=201)
def create_user(payload: UserPayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("user"))) -> dict:
    user = models.User(**payload.model_dump())
    db.add(user)
    commit_or_409(db, "Username already exists")
    db.refresh(user)
    return {"id": user.id, "username": user.username, "display_name": user.display_name, "role": user.role, "department": user.department}


@app.put("/api/admin/users/{user_id}")
def update_user(user_id: int, payload: UserUpdatePayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("user"))) -> dict:
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    update_model(user, payload)
    commit_or_409(db, "Username already exists")
    db.refresh(user)
    return {"id": user.id, "username": user.username, "display_name": user.display_name, "role": user.role, "department": user.department}


@app.delete("/api/admin/users/{user_id}")
def delete_user(user_id: int, db: Session = Depends(get_db), _: dict = Depends(require_permission("user"))) -> dict:
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    if user.username == "admin":
        raise HTTPException(status_code=409, detail="Built-in admin cannot be deleted")
    db.delete(user)
    db.commit()
    return {"deleted": True}


@app.get("/api/admin/workflows")
def workflow_templates(page: int = 1, page_size: int = 20, keyword: str = "", db: Session = Depends(get_db)) -> dict:
    q = db.query(models.WorkflowTemplate).options(selectinload(models.WorkflowTemplate.nodes))
    if keyword:
        kw = f"%{keyword}%"
        q = q.filter(models.WorkflowTemplate.code.ilike(kw) | models.WorkflowTemplate.name.ilike(kw))
    total = q.count()
    rows = q.order_by(models.WorkflowTemplate.id).offset((page - 1) * page_size).limit(page_size).all()
    return {
        "items": [
            {
                "id": row.id,
                "code": row.code,
                "name": row.name,
                "object_type": row.object_type,
                "project_type": row.project_type,
                "status": row.status,
                "description": row.description,
                "nodes": [
                    {"id": node.id, "sequence": node.sequence, "name": node.name, "role_name": node.role_name, "action_type": node.action_type, "sla_hours": node.sla_hours}
                    for node in sorted(row.nodes, key=lambda item: item.sequence)
                ],
            }
            for row in rows
        ],
        "total": total,
        "page": page,
        "page_size": page_size,
    }


@app.post("/api/admin/workflows", status_code=201)
def create_workflow_template(payload: WorkflowTemplatePayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("workflow"))) -> dict:
    template = models.WorkflowTemplate(**payload.model_dump())
    db.add(template)
    commit_or_409(db, "Workflow code already exists")
    db.refresh(template)
    return {"id": template.id, "code": template.code, "name": template.name, "object_type": template.object_type, "project_type": template.project_type, "status": template.status, "description": template.description, "nodes": []}


@app.put("/api/admin/workflows/{workflow_id}")
def update_workflow_template(workflow_id: int, payload: WorkflowTemplateUpdatePayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("workflow"))) -> dict:
    template = db.query(models.WorkflowTemplate).filter(models.WorkflowTemplate.id == workflow_id).first()
    if not template:
        raise HTTPException(status_code=404, detail="Workflow not found")
    update_model(template, payload)
    commit_or_409(db, "Workflow code already exists")
    db.refresh(template)
    return {"id": template.id, "code": template.code, "name": template.name, "object_type": template.object_type, "project_type": template.project_type, "status": template.status, "description": template.description}


@app.delete("/api/admin/workflows/{workflow_id}")
def delete_workflow_template(workflow_id: int, db: Session = Depends(get_db), _: dict = Depends(require_permission("workflow"))) -> dict:
    template = db.query(models.WorkflowTemplate).filter(models.WorkflowTemplate.id == workflow_id).first()
    if not template:
        raise HTTPException(status_code=404, detail="Workflow not found")
    db.delete(template)
    db.commit()
    return {"deleted": True}


@app.post("/api/admin/workflows/{workflow_id}/nodes", status_code=201)
def create_workflow_node(workflow_id: int, payload: WorkflowNodePayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("workflow"))) -> dict:
    if not db.query(models.WorkflowTemplate.id).filter(models.WorkflowTemplate.id == workflow_id).first():
        raise HTTPException(status_code=404, detail="Workflow not found")
    node = models.WorkflowNode(template_id=workflow_id, **payload.model_dump())
    db.add(node)
    db.commit()
    db.refresh(node)
    return {"id": node.id, "sequence": node.sequence, "name": node.name, "role_name": node.role_name, "action_type": node.action_type, "sla_hours": node.sla_hours}


@app.put("/api/admin/workflow-nodes/{node_id}")
def update_workflow_node(node_id: int, payload: WorkflowNodeUpdatePayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("workflow"))) -> dict:
    node = db.query(models.WorkflowNode).filter(models.WorkflowNode.id == node_id).first()
    if not node:
        raise HTTPException(status_code=404, detail="Workflow node not found")
    update_model(node, payload)
    db.commit()
    db.refresh(node)
    return {"id": node.id, "sequence": node.sequence, "name": node.name, "role_name": node.role_name, "action_type": node.action_type, "sla_hours": node.sla_hours}


@app.delete("/api/admin/workflow-nodes/{node_id}")
def delete_workflow_node(node_id: int, db: Session = Depends(get_db), _: dict = Depends(require_permission("workflow"))) -> dict:
    node = db.query(models.WorkflowNode).filter(models.WorkflowNode.id == node_id).first()
    if not node:
        raise HTTPException(status_code=404, detail="Workflow node not found")
    db.delete(node)
    db.commit()
    return {"deleted": True}


@app.get("/api/admin/integration-endpoints")
def integration_endpoints(page: int = 1, page_size: int = 20, keyword: str = "", db: Session = Depends(get_db)) -> dict:
    q = db.query(models.IntegrationEndpoint)
    if keyword:
        kw = f"%{keyword}%"
        q = q.filter(models.IntegrationEndpoint.code.ilike(kw) | models.IntegrationEndpoint.name.ilike(kw))
    total = q.count()
    rows = q.order_by(models.IntegrationEndpoint.id).offset((page - 1) * page_size).limit(page_size).all()
    return {
        "items": [
            {
                "id": row.id,
                "code": row.code,
                "name": row.name,
                "system_type": row.system_type,
                "base_url": row.base_url,
                "auth_type": row.auth_type,
                "direction": row.direction,
                "status": row.status,
                "owner": row.owner,
                "object_scope": row.object_scope,
            }
            for row in rows
        ],
        "total": total,
        "page": page,
        "page_size": page_size,
    }


@app.post("/api/admin/integration-endpoints", status_code=201)
def create_integration_endpoint(payload: IntegrationEndpointPayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("integration"))) -> dict:
    endpoint = models.IntegrationEndpoint(**payload.model_dump())
    db.add(endpoint)
    commit_or_409(db, "Endpoint code already exists")
    db.refresh(endpoint)
    return {"id": endpoint.id, "code": endpoint.code, "name": endpoint.name, "system_type": endpoint.system_type, "base_url": endpoint.base_url, "auth_type": endpoint.auth_type, "direction": endpoint.direction, "status": endpoint.status, "owner": endpoint.owner, "object_scope": endpoint.object_scope}


@app.put("/api/admin/integration-endpoints/{endpoint_id}")
def update_integration_endpoint(endpoint_id: int, payload: IntegrationEndpointUpdatePayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("integration"))) -> dict:
    endpoint = db.query(models.IntegrationEndpoint).filter(models.IntegrationEndpoint.id == endpoint_id).first()
    if not endpoint:
        raise HTTPException(status_code=404, detail="Endpoint not found")
    update_model(endpoint, payload)
    commit_or_409(db, "Endpoint code already exists")
    db.refresh(endpoint)
    return {"id": endpoint.id, "code": endpoint.code, "name": endpoint.name, "system_type": endpoint.system_type, "base_url": endpoint.base_url, "auth_type": endpoint.auth_type, "direction": endpoint.direction, "status": endpoint.status, "owner": endpoint.owner, "object_scope": endpoint.object_scope}


@app.delete("/api/admin/integration-endpoints/{endpoint_id}")
def delete_integration_endpoint(endpoint_id: int, db: Session = Depends(get_db), _: dict = Depends(require_permission("integration"))) -> dict:
    endpoint = db.query(models.IntegrationEndpoint).filter(models.IntegrationEndpoint.id == endpoint_id).first()
    if not endpoint:
        raise HTTPException(status_code=404, detail="Endpoint not found")
    db.delete(endpoint)
    db.commit()
    return {"deleted": True}


@app.get("/api/workflow-instances")
def workflow_instances(page: int = 1, page_size: int = 20, keyword: str = "", db: Session = Depends(get_db)) -> dict:
    q = (
        db.query(models.WorkflowInstance)
        .options(
            selectinload(models.WorkflowInstance.template),
            selectinload(models.WorkflowInstance.tasks),
            selectinload(models.WorkflowInstance.logs),
        )
    )
    if keyword:
        kw = f"%{keyword}%"
        q = q.filter(models.WorkflowInstance.object_no.ilike(kw) | models.WorkflowInstance.title.ilike(kw))
    total = q.count()
    rows = q.order_by(models.WorkflowInstance.id.desc()).offset((page - 1) * page_size).limit(page_size).all()
    return {"items": [serialize_workflow_instance(row) for row in rows], "total": total, "page": page, "page_size": page_size}


@app.get("/api/workflow-tasks")
def workflow_tasks(page: int = 1, page_size: int = 20, keyword: str = "", db: Session = Depends(get_db)) -> dict:
    q = (
        db.query(models.WorkflowTask)
        .join(models.WorkflowInstance)
        .options(selectinload(models.WorkflowTask.instance).selectinload(models.WorkflowInstance.template))
    )
    if keyword:
        kw = f"%{keyword}%"
        q = q.filter(
            models.WorkflowTask.node_name.ilike(kw)
            | models.WorkflowTask.assignee.ilike(kw)
            | models.WorkflowInstance.object_no.ilike(kw)
            | models.WorkflowInstance.title.ilike(kw)
        )
    total = q.count()
    rows = q.order_by(models.WorkflowTask.status.desc(), models.WorkflowTask.id).offset((page - 1) * page_size).limit(page_size).all()
    return {
        "items": [
            {
                "id": row.id,
                "instance_id": row.instance_id,
                "object_type": row.instance.object_type,
                "object_no": row.instance.object_no,
                "title": row.instance.title,
                "product_model": row.instance.product_model,
                "template_name": row.instance.template.name if row.instance.template else "",
                "sequence": row.sequence,
                "node_name": row.node_name,
                "role_name": row.role_name,
                "action_type": row.action_type,
                "status": row.status,
                "assignee": row.assignee,
                "acted_by": row.acted_by,
                "acted_at": row.acted_at,
                "comment": row.comment,
                "sla_hours": row.sla_hours,
            }
            for row in rows
        ],
        "total": total,
        "page": page,
        "page_size": page_size,
    }


@app.post("/api/workflow-tasks/{task_id}/approve")
def approve_workflow_task(task_id: int, payload: WorkflowTaskActionPayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("approval"))) -> dict:
    task = (
        db.query(models.WorkflowTask)
        .options(selectinload(models.WorkflowTask.instance).selectinload(models.WorkflowInstance.tasks))
        .filter(models.WorkflowTask.id == task_id)
        .first()
    )
    if not task:
        raise HTTPException(status_code=404, detail="Workflow task not found")
    if task.status != "待处理":
        raise HTTPException(status_code=409, detail="Only pending workflow task can be approved")
    old_status = task.status
    task.status = "已通过"
    task.acted_by = payload.acted_by
    task.acted_at = today_text()
    task.comment = payload.comment
    add_workflow_log(db, task.instance_id, task.id, "通过", payload.acted_by, payload.comment, old_status, task.status)

    next_task = (
        db.query(models.WorkflowTask)
        .filter(
            models.WorkflowTask.instance_id == task.instance_id,
            models.WorkflowTask.sequence > task.sequence,
            models.WorkflowTask.status == "未开始",
        )
        .order_by(models.WorkflowTask.sequence)
        .first()
    )
    if next_task:
        old_next_status = next_task.status
        next_task.status = "待处理"
        add_workflow_log(db, next_task.instance_id, next_task.id, "激活节点", payload.acted_by, f"上一节点 {task.node_name} 已通过", old_next_status, next_task.status)
    else:
        old_instance_status = task.instance.status
        complete_business_object(db, task.instance)
        add_workflow_log(db, task.instance_id, None, "流程完成", payload.acted_by, "全部节点已通过", old_instance_status, task.instance.status)

    db.commit()
    db.refresh(task)
    return {"id": task.id, "status": task.status, "instance_status": task.instance.status}


@app.post("/api/workflow-tasks/{task_id}/reject")
def reject_workflow_task(task_id: int, payload: WorkflowRejectPayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("approval"))) -> dict:
    task = (
        db.query(models.WorkflowTask)
        .options(selectinload(models.WorkflowTask.instance).selectinload(models.WorkflowInstance.tasks))
        .filter(models.WorkflowTask.id == task_id)
        .first()
    )
    if not task:
        raise HTTPException(status_code=404, detail="Workflow task not found")
    if task.status != "待处理":
        raise HTTPException(status_code=409, detail="Only pending workflow task can be rejected")
    old_status = task.status
    task.status = "已驳回"
    task.acted_by = payload.acted_by
    task.acted_at = today_text()
    task.comment = payload.comment
    add_workflow_log(db, task.instance_id, task.id, "驳回", payload.acted_by, payload.comment, old_status, task.status)

    if payload.target_sequence:
        target = (
            db.query(models.WorkflowTask)
            .filter(
                models.WorkflowTask.instance_id == task.instance_id,
                models.WorkflowTask.sequence == payload.target_sequence,
            )
            .first()
        )
        if not target:
            raise HTTPException(status_code=404, detail="Reject target node not found")
        old_target_status = target.status
        target.status = "待处理"
        add_workflow_log(db, task.instance_id, target.id, "退回节点", payload.acted_by, f"从 {task.node_name} 驳回", old_target_status, target.status)
    else:
        old_instance_status = task.instance.status
        task.instance.status = "已驳回"
        withdraw_business_object(db, task.instance)
        add_workflow_log(db, task.instance_id, None, "流程驳回结束", payload.acted_by, payload.comment, old_instance_status, task.instance.status)
    db.commit()
    return {"id": task.id, "status": task.status, "instance_status": task.instance.status}


@app.post("/api/workflow-tasks/{task_id}/transfer")
def transfer_workflow_task(task_id: int, payload: WorkflowTransferPayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("approval"))) -> dict:
    task = db.query(models.WorkflowTask).filter(models.WorkflowTask.id == task_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="Workflow task not found")
    if task.status != "待处理":
        raise HTTPException(status_code=409, detail="Only pending workflow task can be transferred")
    old_assignee = task.assignee
    task.assignee = payload.assignee
    add_workflow_log(db, task.instance_id, task.id, "转办", payload.acted_by, payload.comment or f"转办给 {payload.assignee}", old_assignee, payload.assignee)
    db.commit()
    return {"id": task.id, "assignee": task.assignee, "status": task.status}


@app.post("/api/workflow-instances/{instance_id}/withdraw")
def withdraw_workflow_instance(instance_id: int, payload: WorkflowWithdrawPayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("approval"))) -> dict:
    instance = (
        db.query(models.WorkflowInstance)
        .options(selectinload(models.WorkflowInstance.tasks))
        .filter(models.WorkflowInstance.id == instance_id)
        .first()
    )
    if not instance:
        raise HTTPException(status_code=404, detail="Workflow instance not found")
    if instance.status != "运行中":
        raise HTTPException(status_code=409, detail="Only running workflow can be withdrawn")
    old_status = instance.status
    instance.status = "已撤回"
    instance.completed_at = today_text()
    for task in instance.tasks:
        if task.status in ["待处理", "未开始"]:
            task.status = "已取消"
    withdraw_business_object(db, instance)
    add_workflow_log(db, instance.id, None, "撤回", payload.acted_by, payload.comment, old_status, instance.status)
    db.commit()
    return {"id": instance.id, "status": instance.status}


@app.get("/api/dashboard")
def dashboard(db: Session = Depends(get_db)) -> dict:
    product_count = db.query(models.Product).count()
    active_projects = db.query(models.Project).filter(models.Project.phase != "量产导入").count()
    pending_changes = db.query(models.Change).filter(models.Change.status.in_(["审批中", "执行中"])).count()
    docs = db.query(models.Document).count()
    signed_docs = db.query(models.Document).filter(models.Document.approval_status == "已签核").count()
    bom_total = db.query(models.BomHeader).count()
    bom_ready = db.query(models.BomHeader).filter(models.BomHeader.status == "已发布").count()

    lifecycle_rows = db.query(models.Product.lifecycle, func.count(models.Product.id)).group_by(models.Product.lifecycle).all()
    changes_rows = db.query(models.Change.change_type, func.count(models.Change.id)).group_by(models.Change.change_type).all()
    quality_rows = db.query(models.QualityLot.tested_at, func.avg(models.QualityLot.cp_yield), func.avg(models.QualityLot.ft_yield)).group_by(models.QualityLot.tested_at).order_by(models.QualityLot.tested_at).all()

    # 新增聚合指标
    integration_pending = db.query(models.IntegrationJob).filter(models.IntegrationJob.status.in_(["待处理", "处理中", "失败"])).count()
    integration_failed = db.query(models.IntegrationJob).filter(models.IntegrationJob.status == "失败").count()
    quality_open_issues = db.query(models.QualityIssue).filter(models.QualityIssue.status != "已关闭").count()
    capa_open = db.query(models.QualityCAPA).filter(models.QualityCAPA.status != "已关闭").count()
    deliverable_pending = db.query(models.ProjectDeliverable).filter(models.ProjectDeliverable.status.notin_(["已完成", "已关闭"])).count()
    recent_reports = db.query(models.QualityReport).count()

    project_rows = db.query(models.Project.phase, func.count(models.Project.id)).group_by(models.Project.phase).all()
    integration_status_rows = db.query(models.IntegrationJob.status, func.count(models.IntegrationJob.id)).group_by(models.IntegrationJob.status).all()
    change_status_rows = db.query(models.Change.status, func.count(models.Change.id)).group_by(models.Change.status).all()

    return {
        "metrics": {
            "products": product_count,
            "active_projects": active_projects,
            "pending_changes": pending_changes,
            "document_readiness": round((signed_docs / docs) * 100) if docs else 0,
            "bom_readiness": round((bom_ready / bom_total) * 100) if bom_total else 0,
            "integration_pending": integration_pending,
            "integration_failed": integration_failed,
            "quality_open_issues": quality_open_issues,
            "capa_open": capa_open,
            "deliverable_pending": deliverable_pending,
            "quality_reports": recent_reports,
        },
        "lifecycle": [{"name": name, "value": value} for name, value in lifecycle_rows],
        "changes": [{"name": name, "value": value} for name, value in changes_rows],
        "quality_trend": [{"date": date, "cp": round(cp, 1), "ft": round(ft, 1)} for date, cp, ft in quality_rows],
        "project_phases": [{"name": name, "value": value} for name, value in project_rows],
        "integration_status": [{"name": name, "value": value} for name, value in integration_status_rows],
        "change_status": [{"name": name, "value": value} for name, value in change_status_rows],
        "recent_tasks": [
            {
                "name": task.name,
                "phase": task.phase,
                "owner": task.owner,
                "status": task.status,
                "due_date": task.due_date,
            }
            for task in db.query(models.ProjectTask).order_by(models.ProjectTask.due_date).limit(6)
        ],
    }


@app.get("/api/products")
def products(page: int = 1, page_size: int = 20, keyword: str = "", db: Session = Depends(get_db)) -> dict:
    q = db.query(models.Product)
    if keyword:
        kw = f"%{keyword}%"
        q = q.filter(
            models.Product.code.ilike(kw)
            | models.Product.model.ilike(kw)
            | models.Product.name.ilike(kw)
            | models.Product.product_type.ilike(kw)
        )
    total = q.count()
    rows = q.order_by(models.Product.id).offset((page - 1) * page_size).limit(page_size).all()
    return {
        "items": [serialize_product(p) for p in rows],
        "total": total,
        "page": page,
        "page_size": page_size,
    }


@app.post("/api/products", status_code=201)
def create_product(payload: ProductPayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("product"))) -> dict:
    product = models.Product(**payload.model_dump())
    db.add(product)
    commit_or_409(db, "Product code or model already exists")
    db.refresh(product)
    return serialize_product(product)


@app.get("/api/products/{product_id}")
def product_detail(product_id: int, db: Session = Depends(get_db)) -> dict:
    product = (
        db.query(models.Product)
        .options(
            selectinload(models.Product.bom_headers),
            selectinload(models.Product.documents),
            selectinload(models.Product.process_routes).selectinload(models.ProcessRoute.steps),
            selectinload(models.Product.changes).selectinload(models.Change.impacts),
            selectinload(models.Product.quality_lots),
            selectinload(models.Product.requirements),
        )
        .filter(models.Product.id == product_id)
        .first()
    )
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    data = serialize_product(product)
    # 项目通过 product_model 字符串关联，非外键，单独查询
    linked_projects = (
        db.query(models.Project)
        .filter(models.Project.product_model == product.model)
        .order_by(models.Project.id.desc())
        .all()
    )
    data.update(
        {
            "boms": [{"id": bom.id, "type": bom.bom_type, "version": bom.version, "status": bom.status, "owner": bom.owner, "release_date": bom.release_date} for bom in product.bom_headers],
            "documents": [{"id": doc.id, "doc_no": doc.doc_no, "title": doc.title, "category": doc.category, "version": doc.version, "status": doc.status, "approval_status": doc.approval_status} for doc in product.documents],
            "routes": [{"id": route.id, "route_no": route.route_no, "name": route.name, "version": route.version, "status": route.status, "steps": len(route.steps)} for route in product.process_routes],
            "changes": [{"id": change.id, "change_no": change.change_no, "title": change.title, "status": change.status, "priority": change.priority} for change in product.changes],
            "quality": [{"lot_no": lot.lot_no, "stage": lot.stage, "cp_yield": lot.cp_yield, "ft_yield": lot.ft_yield, "status": lot.status} for lot in product.quality_lots],
            "requirements": [{"id": req.id, "req_no": req.req_no, "title": req.title, "source": req.source, "category": req.category, "priority": req.priority, "status": req.status, "owner": req.owner} for req in product.requirements],
            "projects": [{"id": proj.id, "project_no": proj.project_no, "name": proj.name, "phase": proj.phase, "progress": proj.progress, "owner": proj.owner, "risk_level": proj.risk_level} for proj in linked_projects],
        }
    )
    return data


@app.put("/api/products/{product_id}")
def update_product(product_id: int, payload: ProductUpdatePayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("product"))) -> dict:
    product = db.query(models.Product).filter(models.Product.id == product_id).first()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    update_model(product, payload)
    commit_or_409(db, "Product code or model already exists")
    db.refresh(product)
    return serialize_product(product)


@app.delete("/api/products/{product_id}")
def delete_product(product_id: int, db: Session = Depends(get_db), _: dict = Depends(require_permission("product"))) -> dict:
    product = db.query(models.Product).filter(models.Product.id == product_id).first()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    linked_count = (
        db.query(models.BomHeader).filter(models.BomHeader.product_id == product_id).count()
        + db.query(models.Document).filter(models.Document.product_id == product_id).count()
        + db.query(models.ProcessRoute).filter(models.ProcessRoute.product_id == product_id).count()
        + db.query(models.Change).filter(models.Change.product_id == product_id).count()
        + db.query(models.QualityLot).filter(models.QualityLot.product_id == product_id).count()
        + db.query(models.Requirement).filter(models.Requirement.product_id == product_id).count()
        + db.query(models.ProductBaseline).filter(models.ProductBaseline.product_id == product_id).count()
    )
    if linked_count:
        raise HTTPException(status_code=409, detail="Product has linked PLM data and cannot be deleted")
    db.delete(product)
    db.commit()
    return {"deleted": True}


@app.get("/api/products/{product_id}/versions")
def product_versions(product_id: int, db: Session = Depends(get_db)) -> list[dict]:
    rows = db.query(models.ProductVersion).filter(models.ProductVersion.product_id == product_id).order_by(models.ProductVersion.id.desc()).all()
    return [
        {"id": r.id, "version": r.version, "lifecycle": r.lifecycle, "readiness": r.readiness, "released_at": r.released_at, "released_by": r.released_by, "source_change_no": r.source_change_no, "summary": r.summary}
        for r in rows
    ]


@app.post("/api/products/{product_id}/versions", status_code=201)
def create_product_version(product_id: int, payload: ProductVersionPayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("product"))) -> dict:
    ensure_product_exists(db, product_id)
    row = models.ProductVersion(product_id=product_id, **payload.model_dump())
    db.add(row)
    db.commit()
    db.refresh(row)
    return {"id": row.id, "version": row.version}


@app.get("/api/boms")
def boms(page: int = 1, page_size: int = 20, keyword: str = "", db: Session = Depends(get_db)) -> dict:
    q = db.query(models.BomHeader).options(selectinload(models.BomHeader.product), selectinload(models.BomHeader.items))
    if keyword:
        kw = f"%{keyword}%"
        q = q.join(models.Product).filter(models.BomHeader.bom_type.ilike(kw) | models.BomHeader.version.ilike(kw) | models.Product.model.ilike(kw))
    total = q.count()
    rows = q.order_by(models.BomHeader.id).offset((page - 1) * page_size).limit(page_size).all()
    return {"items": [serialize_bom(row) for row in rows], "total": total, "page": page, "page_size": page_size}


@app.post("/api/boms", status_code=201)
def create_bom(payload: BomHeaderPayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("bom"))) -> dict:
    ensure_product_exists(db, payload.product_id)
    exists = (
        db.query(models.BomHeader.id)
        .filter(
            models.BomHeader.product_id == payload.product_id,
            models.BomHeader.bom_type == payload.bom_type,
            models.BomHeader.version == payload.version,
        )
        .first()
    )
    if exists:
        raise HTTPException(status_code=409, detail="BOM version already exists for this product and type")
    bom = models.BomHeader(**payload.model_dump())
    db.add(bom)
    db.commit()
    db.refresh(bom)
    bom = db.query(models.BomHeader).options(selectinload(models.BomHeader.product), selectinload(models.BomHeader.items)).filter(models.BomHeader.id == bom.id).first()
    return serialize_bom(bom)


@app.put("/api/boms/{bom_id}")
def update_bom(bom_id: int, payload: BomHeaderUpdatePayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("bom"))) -> dict:
    bom = db.query(models.BomHeader).filter(models.BomHeader.id == bom_id).first()
    if not bom:
        raise HTTPException(status_code=404, detail="BOM not found")
    if bom.status in ("审批中", "已发布"):
        raise HTTPException(status_code=409, detail=f"BOM in {bom.status} status cannot be edited")
    if payload.product_id is not None:
        ensure_product_exists(db, payload.product_id)
    next_product_id = payload.product_id if payload.product_id is not None else bom.product_id
    next_type = payload.bom_type if payload.bom_type is not None else bom.bom_type
    next_version = payload.version if payload.version is not None else bom.version
    exists = (
        db.query(models.BomHeader.id)
        .filter(
            models.BomHeader.id != bom.id,
            models.BomHeader.product_id == next_product_id,
            models.BomHeader.bom_type == next_type,
            models.BomHeader.version == next_version,
        )
        .first()
    )
    if exists:
        raise HTTPException(status_code=409, detail="BOM version already exists for this product and type")
    update_model(bom, payload)
    db.commit()
    db.refresh(bom)
    bom = db.query(models.BomHeader).options(selectinload(models.BomHeader.product), selectinload(models.BomHeader.items)).filter(models.BomHeader.id == bom.id).first()
    return serialize_bom(bom)


@app.delete("/api/boms/{bom_id}")
def delete_bom(bom_id: int, db: Session = Depends(get_db), _: dict = Depends(require_permission("bom"))) -> dict:
    bom = db.query(models.BomHeader).filter(models.BomHeader.id == bom_id).first()
    if not bom:
        raise HTTPException(status_code=404, detail="BOM not found")
    if bom.status == "已发布":
        raise HTTPException(status_code=409, detail="Released BOM cannot be deleted")
    if bom.status == "审批中":
        raise HTTPException(status_code=409, detail="BOM in review cannot be deleted")
    db.delete(bom)
    db.commit()
    return {"deleted": True}


@app.post("/api/boms/{bom_id}/items", status_code=201)
def create_bom_item(bom_id: int, payload: BomItemPayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("bom"))) -> dict:
    bom = db.query(models.BomHeader).filter(models.BomHeader.id == bom_id).first()
    if not bom:
        raise HTTPException(status_code=404, detail="BOM not found")
    if bom.status in ("审批中", "已发布"):
        raise HTTPException(status_code=409, detail=f"BOM in {bom.status} status cannot be edited")
    data = apply_bom_item_process_binding(db, payload, bom.product_id)
    item = models.BomItem(bom_id=bom_id, parent_id=None, **data)
    db.add(item)
    db.commit()
    db.refresh(item)
    return serialize_bom_item(item)


@app.put("/api/bom-items/{item_id}")
def update_bom_item(item_id: int, payload: BomItemUpdatePayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("bom"))) -> dict:
    item = db.query(models.BomItem).options(selectinload(models.BomItem.bom)).filter(models.BomItem.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="BOM item not found")
    if item.bom.status in ("审批中", "已发布"):
        raise HTTPException(status_code=409, detail=f"BOM in {item.bom.status} status cannot be edited")
    data = apply_bom_item_process_binding(db, payload, item.bom.product_id)
    for key, value in data.items():
        setattr(item, key, value)
    db.commit()
    db.refresh(item)
    return serialize_bom_item(item)


@app.post("/api/boms/{bom_id}/transform", status_code=201)
def transform_bom(bom_id: int, payload: BomTransformPayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("bom"))) -> dict:
    source = (
        db.query(models.BomHeader)
        .options(selectinload(models.BomHeader.product), selectinload(models.BomHeader.items))
        .filter(models.BomHeader.id == bom_id)
        .first()
    )
    if not source:
        raise HTTPException(status_code=404, detail="Source BOM not found")
    if source.bom_type not in ["EBOM", "PBOM"] and payload.target_type in ["PBOM", "MBOM"]:
        raise HTTPException(status_code=409, detail="Only EBOM/PBOM can be transformed to downstream BOM")
    exists = (
        db.query(models.BomHeader.id)
        .filter(
            models.BomHeader.product_id == source.product_id,
            models.BomHeader.bom_type == payload.target_type,
            models.BomHeader.version == payload.version,
        )
        .first()
    )
    if exists:
        raise HTTPException(status_code=409, detail="Target BOM version already exists for this product and type")
    target = models.BomHeader(
        product_id=source.product_id,
        bom_type=payload.target_type,
        version=payload.version,
        status="编制中",
        owner=payload.owner or source.owner,
        release_date="",
        source_bom_id=source.id,
        effective_date=payload.effective_date,
        expiry_date="",
        effectivity_type=payload.effectivity_type,
        effective_batch=payload.effective_batch or source.effective_batch or "",
    )
    db.add(target)
    db.flush()
    for item in source.items:
        db.add(models.BomItem(
            bom_id=target.id,
            parent_id=None,
            material_code=item.material_code,
            material_name=item.material_name,
            category=item.category,
            specification=item.specification,
            quantity=item.quantity,
            unit=item.unit,
            position=item.position,
            process_step_id=item.process_step_id,
            process_step=item.process_step or "待工艺分配",
            substitute=item.substitute,
            status=item.status,
            effective_date=payload.effective_date or item.effective_date,
            expiry_date=item.expiry_date,
            effectivity_note=f"由 {source.bom_type} {source.version} 转换",
        ))
    db.commit()
    created = (
        db.query(models.BomHeader)
        .options(selectinload(models.BomHeader.product), selectinload(models.BomHeader.items))
        .filter(models.BomHeader.id == target.id)
        .first()
    )
    return serialize_bom(created)


@app.get("/api/boms/{bom_id}/compare/{target_bom_id}")
def compare_bom(bom_id: int, target_bom_id: int, db: Session = Depends(get_db)) -> dict:
    base = db.query(models.BomHeader).options(selectinload(models.BomHeader.product), selectinload(models.BomHeader.items)).filter(models.BomHeader.id == bom_id).first()
    target = db.query(models.BomHeader).options(selectinload(models.BomHeader.product), selectinload(models.BomHeader.items)).filter(models.BomHeader.id == target_bom_id).first()
    if not base or not target:
        raise HTTPException(status_code=404, detail="BOM not found")
    base_map = {bom_item_compare_key(item): item for item in base.items}
    target_map = {bom_item_compare_key(item): item for item in target.items}
    changes = []
    for key in sorted(set(base_map) | set(target_map)):
        base_item = base_map.get(key)
        target_item = target_map.get(key)
        if base_item and not target_item:
            changes.append(serialize_bom_compare_item("删除", None, base_item))
        elif target_item and not base_item:
            changes.append(serialize_bom_compare_item("新增", target_item, None))
        elif base_item and target_item and (
            base_item.quantity != target_item.quantity
            or base_item.status != target_item.status
            or base_item.effective_date != target_item.effective_date
            or base_item.expiry_date != target_item.expiry_date
        ):
            changes.append(serialize_bom_compare_item("变更", target_item, base_item))
    return {
        "base": {"id": base.id, "type": base.bom_type, "version": base.version, "product_model": base.product.model},
        "target": {"id": target.id, "type": target.bom_type, "version": target.version, "product_model": target.product.model},
        "summary": {
            "added": len([item for item in changes if item["change_type"] == "新增"]),
            "removed": len([item for item in changes if item["change_type"] == "删除"]),
            "changed": len([item for item in changes if item["change_type"] == "变更"]),
        },
        "changes": changes,
    }


@app.get("/api/boms/where-used/{material_code}")
def bom_where_used(material_code: str, db: Session = Depends(get_db)) -> list[dict]:
    rows = (
        db.query(models.BomItem)
        .join(models.BomHeader)
        .options(selectinload(models.BomItem.bom).selectinload(models.BomHeader.product))
        .filter(models.BomItem.material_code == material_code)
        .order_by(models.BomHeader.product_id, models.BomHeader.bom_type, models.BomHeader.version)
        .all()
    )
    return [
        {
            "bom_id": item.bom_id,
            "product_model": item.bom.product.model,
            "product_name": item.bom.product.name,
            "bom_type": item.bom.bom_type,
            "version": item.bom.version,
            "bom_status": item.bom.status,
            "material_code": item.material_code,
            "material_name": item.material_name,
            "quantity": item.quantity,
            "unit": item.unit,
            "process_step": item.process_step,
            "effective_date": item.effective_date,
            "expiry_date": item.expiry_date,
        }
        for item in rows
    ]


@app.delete("/api/bom-items/{item_id}")
def delete_bom_item(item_id: int, db: Session = Depends(get_db), _: dict = Depends(require_permission("bom"))) -> dict:
    item = db.query(models.BomItem).options(selectinload(models.BomItem.bom)).filter(models.BomItem.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="BOM item not found")
    if item.bom.status in ("审批中", "已发布"):
        raise HTTPException(status_code=409, detail=f"BOM in {item.bom.status} status cannot be edited")
    db.delete(item)
    db.commit()
    return {"deleted": True}


@app.post("/api/boms/{bom_id}/submit")
def submit_bom(bom_id: int, db: Session = Depends(get_db), _: dict = Depends(require_permission("bom"))) -> dict:
    bom = db.query(models.BomHeader).options(selectinload(models.BomHeader.product), selectinload(models.BomHeader.items)).filter(models.BomHeader.id == bom_id).first()
    if not bom:
        raise HTTPException(status_code=404, detail="BOM not found")
    if not bom.items:
        raise HTTPException(status_code=409, detail="BOM has no items")
    if bom.status == "已发布":
        raise HTTPException(status_code=409, detail="Released BOM cannot be submitted")
    if bom.status == "审批中":
        raise HTTPException(status_code=409, detail="BOM is already in review")
    # ECA 生成对象校验：如果 BOM 有 source_bom_id，说明是 ECA 升版生成的草案
    if bom.source_bom_id:
        generated_no = f"{bom.bom_type}-{bom.product.model}-{bom.version}"
        validate_eca_generated_object_ready(db, "BOM", bom.id, generated_no)
    bom.status = "审批中"
    instance = start_workflow(
        db,
        template_code="WF-BOM-REL",
        object_type="BOM",
        object_id=bom.id,
        object_no=f"{bom.bom_type}-{bom.product.model}-{bom.version}",
        title=f"{bom.product.model} {bom.bom_type} {bom.version} 发布",
        product_model=bom.product.model,
        started_by=bom.owner,
    )
    db.commit()
    return {"id": bom.id, "status": bom.status, "workflow_id": instance.id}


@app.post("/api/boms/{bom_id}/approve")
def approve_bom(bom_id: int, db: Session = Depends(get_db), _: dict = Depends(require_permission(["approval", "bom"]))) -> dict:
    bom = db.query(models.BomHeader).options(selectinload(models.BomHeader.product), selectinload(models.BomHeader.items)).filter(models.BomHeader.id == bom_id).first()
    if not bom:
        raise HTTPException(status_code=404, detail="BOM not found")
    if not bom.items:
        raise HTTPException(status_code=409, detail="BOM has no items")
    if bom.status not in ("审批中", "已发布"):
        raise HTTPException(status_code=409, detail="BOM must be submitted for review before approval")
    if bom.source_bom_id:
        generated_no = f"{bom.bom_type}-{bom.product.model}-{bom.version}"
        validate_eca_generated_object_ready(db, "BOM", bom.id, generated_no)
    if bom.status == "已发布" and is_current_effective_bom(bom):
        return {"id": bom.id, "status": bom.status, "release_date": bom.release_date, "effective_date": bom.effective_date, "expiry_date": bom.expiry_date, "closed_versions": []}
    if not bom.effective_date:
        bom.effective_date = today_text()
    bom.status = "已发布"
    bom.release_date = today_text()
    bom.expiry_date = ""
    closed_versions = close_previous_effective_boms(db, bom)
    create_integration_job(
        db,
        target_system="ERP",
        object_type="BOM",
        object_no=f"{bom.bom_type}-{bom.product.model}-{bom.version}",
        product_model=bom.product.model,
        triggered_by=f"BOM-{bom.id}",
        message="BOM 已发布，等待同步 ERP 物料结构和用量。",
    )
    db.commit()
    return {"id": bom.id, "status": bom.status, "release_date": bom.release_date, "effective_date": bom.effective_date, "expiry_date": bom.expiry_date, "closed_versions": closed_versions}


@app.get("/api/boms/{bom_id}/export")
def export_bom_excel(bom_id: int, db: Session = Depends(get_db)):
    from openpyxl import Workbook
    bom = db.query(models.BomHeader).options(selectinload(models.BomHeader.product), selectinload(models.BomHeader.items)).filter(models.BomHeader.id == bom_id).first()
    if not bom:
        raise HTTPException(status_code=404, detail="BOM not found")
    wb = Workbook()
    ws = wb.active
    ws.title = "BOM"
    ws.append(["BOM编号", f"{bom.bom_type}-{bom.product.model}-{bom.version}"])
    ws.append(["产品型号", bom.product.model])
    ws.append(["BOM类型", bom.bom_type])
    ws.append(["版本", bom.version])
    ws.append(["状态", bom.status])
    ws.append(["负责人", bom.owner])
    ws.append(["生效日期", bom.effective_date])
    ws.append(["失效日期", bom.expiry_date])
    ws.append([])
    ws.append(["物料编码", "物料名称", "分类", "规格", "数量", "单位", "位号", "工序", "替代料", "状态", "生效日期", "失效日期", "有效性备注"])
    for item in bom.items:
        ws.append([
            item.material_code, item.material_name, item.category, item.specification,
            item.quantity, item.unit, item.position, item.process_step,
            item.substitute, item.status, item.effective_date, item.expiry_date, item.effectivity_note,
        ])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"{bom.bom_type}-{bom.product.model}-{bom.version}.xlsx"
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.post("/api/boms/{bom_id}/import")
async def import_bom_excel(bom_id: int, file: UploadFile = File(...), db: Session = Depends(get_db), _: dict = Depends(require_permission("bom"))) -> dict:
    from openpyxl import load_workbook
    bom = db.query(models.BomHeader).filter(models.BomHeader.id == bom_id).first()
    if not bom:
        raise HTTPException(status_code=404, detail="BOM not found")
    if bom.status == "已发布":
        raise HTTPException(status_code=409, detail="Released BOM cannot be modified")
    content = await file.read()
    wb = load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active
    imported_count = 0
    for row in ws.iter_rows(min_row=1, values_only=True):
        if not row or not row[0]:
            continue
        if str(row[0]).strip() in ("物料编码", "BOM编号", "产品型号", "BOM类型", "版本", "状态", "负责人", "生效日期", "失效日期"):
            continue
        if len(row) < 6:
            continue
        material_code = str(row[0]).strip()
        material_name = str(row[1]).strip() if row[1] else ""
        category = str(row[2]).strip() if row[2] else ""
        specification = str(row[3]).strip() if row[3] else ""
        try:
            quantity = float(row[4]) if row[4] else 0
        except (ValueError, TypeError):
            quantity = 0
        unit = str(row[5]).strip() if row[5] else ""
        position = str(row[6]).strip() if len(row) > 6 and row[6] else ""
        process_step = str(row[7]).strip() if len(row) > 7 and row[7] else ""
        substitute = str(row[8]).strip() if len(row) > 8 and row[8] else ""
        status = str(row[9]).strip() if len(row) > 9 and row[9] else "有效"
        effective_date = str(row[10]).strip() if len(row) > 10 and row[10] else ""
        expiry_date = str(row[11]).strip() if len(row) > 11 and row[11] else ""
        effectivity_note = str(row[12]).strip() if len(row) > 12 and row[12] else ""
        item = models.BomItem(
            bom_id=bom_id, material_code=material_code, material_name=material_name,
            category=category, specification=specification, quantity=quantity, unit=unit,
            position=position, process_step=process_step, substitute=substitute,
            status=status, effective_date=effective_date, expiry_date=expiry_date,
            effectivity_note=effectivity_note,
        )
        db.add(item)
        imported_count += 1
    db.commit()
    return {"imported": imported_count, "bom_id": bom_id}


@app.get("/api/materials")
def materials(page: int = 1, page_size: int = 20, keyword: str = "", db: Session = Depends(get_db)) -> dict:
    q = db.query(models.Material)
    if keyword:
        kw = f"%{keyword}%"
        q = q.filter(models.Material.code.ilike(kw) | models.Material.name.ilike(kw) | models.Material.category.ilike(kw))
    total = q.count()
    rows = q.order_by(models.Material.category, models.Material.code).offset((page - 1) * page_size).limit(page_size).all()
    return {
        "items": [
            {"id": row.id, "code": row.code, "name": row.name, "category": row.category, "specification": row.specification, "supplier": row.supplier, "supplier_id": row.supplier_id, "risk_level": row.risk_level, "lifecycle": row.lifecycle}
            for row in rows
        ],
        "total": total, "page": page, "page_size": page_size,
    }


@app.post("/api/materials", status_code=201)
def create_material(payload: MaterialPayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("material"))) -> dict:
    material = models.Material(**payload.model_dump())
    db.add(material)
    commit_or_409(db, "Material code already exists")
    db.refresh(material)
    return {
        "id": material.id,
        "code": material.code,
        "name": material.name,
        "category": material.category,
        "specification": material.specification,
        "supplier": material.supplier,
        "supplier_id": material.supplier_id,
        "risk_level": material.risk_level,
        "lifecycle": material.lifecycle,
    }


@app.put("/api/materials/{material_id}")
def update_material(material_id: int, payload: MaterialUpdatePayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("material"))) -> dict:
    material = db.query(models.Material).filter(models.Material.id == material_id).first()
    if not material:
        raise HTTPException(status_code=404, detail="Material not found")
    update_model(material, payload)
    commit_or_409(db, "Material code already exists")
    db.refresh(material)
    return {
        "id": material.id,
        "code": material.code,
        "name": material.name,
        "category": material.category,
        "specification": material.specification,
        "supplier": material.supplier,
        "supplier_id": material.supplier_id,
        "risk_level": material.risk_level,
        "lifecycle": material.lifecycle,
    }


@app.delete("/api/materials/{material_id}")
def delete_material(material_id: int, db: Session = Depends(get_db), _: dict = Depends(require_permission("material"))) -> dict:
    material = db.query(models.Material).filter(models.Material.id == material_id).first()
    if not material:
        raise HTTPException(status_code=404, detail="Material not found")
    bom_refs = db.query(models.BomItem).filter(models.BomItem.material_code == material.code).count()
    if bom_refs:
        raise HTTPException(status_code=409, detail="Material is used by BOM and cannot be deleted")
    db.delete(material)
    db.commit()
    return {"deleted": True}


@app.get("/api/documents")
def documents(page: int = 1, page_size: int = 20, keyword: str = "", db: Session = Depends(get_db)) -> dict:
    q = db.query(models.Document).options(selectinload(models.Document.product))
    if keyword:
        kw = f"%{keyword}%"
        q = q.filter(models.Document.doc_no.ilike(kw) | models.Document.title.ilike(kw) | models.Document.category.ilike(kw))
    total = q.count()
    rows = q.order_by(models.Document.id).offset((page - 1) * page_size).limit(page_size).all()
    return {
        "items": [
            {"id": row.id, "doc_no": row.doc_no, "title": row.title, "category": row.category, "version": row.version, "status": row.status, "approval_status": row.approval_status, "owner": row.owner, "updated_at": row.updated_at, "product_model": row.product.model, "product_id": row.product_id, "file_name": row.file_name, "file_size": row.file_size, "file_type": row.file_type}
            for row in rows
        ],
        "total": total, "page": page, "page_size": page_size,
    }


@app.post("/api/documents", status_code=201)
def create_document(payload: DocumentPayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("document"))) -> dict:
    ensure_product_exists(db, payload.product_id)
    document = models.Document(**payload.model_dump())
    db.add(document)
    commit_or_409(db, "Document number already exists")
    db.refresh(document)
    return {
        "id": document.id,
        "doc_no": document.doc_no,
        "title": document.title,
        "category": document.category,
        "version": document.version,
        "status": document.status,
        "approval_status": document.approval_status,
        "owner": document.owner,
        "updated_at": document.updated_at,
        "product_model": document.product.model,
        "product_id": document.product_id,
        "file_name": document.file_name,
        "file_size": document.file_size,
        "file_type": document.file_type,
    }


@app.put("/api/documents/{document_id}")
def update_document(document_id: int, payload: DocumentUpdatePayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("document"))) -> dict:
    document = db.query(models.Document).options(selectinload(models.Document.product)).filter(models.Document.id == document_id).first()
    if not document:
        raise HTTPException(status_code=404, detail="Document not found")
    if document.status in ("审批中", "已发布"):
        raise HTTPException(status_code=409, detail=f"Document in {document.status} status cannot be edited")
    if payload.product_id is not None:
        ensure_product_exists(db, payload.product_id)
    update_model(document, payload)
    commit_or_409(db, "Document number already exists")
    db.refresh(document)
    return {
        "id": document.id,
        "doc_no": document.doc_no,
        "title": document.title,
        "category": document.category,
        "version": document.version,
        "status": document.status,
        "approval_status": document.approval_status,
        "owner": document.owner,
        "updated_at": document.updated_at,
        "product_model": document.product.model,
        "product_id": document.product_id,
        "file_name": document.file_name,
        "file_size": document.file_size,
        "file_type": document.file_type,
    }


@app.delete("/api/documents/{document_id}")
def delete_document(document_id: int, db: Session = Depends(get_db), _: dict = Depends(require_permission("document"))) -> dict:
    document = db.query(models.Document).filter(models.Document.id == document_id).first()
    if not document:
        raise HTTPException(status_code=404, detail="Document not found")
    if document.status == "已发布":
        raise HTTPException(status_code=409, detail="Released document cannot be deleted")
    if document.status == "审批中":
        raise HTTPException(status_code=409, detail="Document in review cannot be deleted")
    db.delete(document)
    db.commit()
    return {"deleted": True}


@app.post("/api/documents/{document_id}/submit")
def submit_document(document_id: int, db: Session = Depends(get_db), _: dict = Depends(require_permission("document"))) -> dict:
    document = db.query(models.Document).options(selectinload(models.Document.product)).filter(models.Document.id == document_id).first()
    if not document:
        raise HTTPException(status_code=404, detail="Document not found")
    if document.status == "已发布":
        raise HTTPException(status_code=409, detail="Released document cannot be submitted")
    if document.status == "审批中":
        raise HTTPException(status_code=409, detail="Document is already in review")
    # ECA 生成对象校验：检查文档是否由 ECA 动作生成
    validate_eca_generated_object_ready(db, "文档", document.id, document.doc_no)
    document.status = "审批中"
    document.approval_status = "流转中"
    document.updated_at = today_text()
    instance = start_workflow(
        db,
        template_code="WF-DOC-STD",
        object_type="文档",
        object_id=document.id,
        object_no=document.doc_no,
        title=document.title,
        product_model=document.product.model,
        started_by=document.owner,
    )
    db.commit()
    return {"id": document.id, "status": document.status, "approval_status": document.approval_status, "workflow_id": instance.id}


@app.post("/api/documents/{document_id}/approve")
def approve_document(document_id: int, db: Session = Depends(get_db), _: dict = Depends(require_permission(["approval", "document"]))) -> dict:
    document = db.query(models.Document).options(selectinload(models.Document.product)).filter(models.Document.id == document_id).first()
    if not document:
        raise HTTPException(status_code=404, detail="Document not found")
    if document.status not in ("审批中", "已发布"):
        raise HTTPException(status_code=409, detail="Document must be submitted for review before approval")
    validate_eca_generated_object_ready(db, "文档", document.id, document.doc_no)
    document.status = "已发布"
    document.approval_status = "已签核"
    document.updated_at = today_text()
    create_integration_job(
        db,
        target_system="QMS",
        object_type="文档",
        object_no=document.doc_no,
        product_model=document.product.model,
        triggered_by=document.doc_no,
        message="文档已签核发布，等待同步 QMS/文控归档。",
    )
    db.commit()
    return {"id": document.id, "status": document.status, "approval_status": document.approval_status, "updated_at": document.updated_at}


FILE_UPLOAD_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "data", "files")


@app.post("/api/documents/{document_id}/upload")
async def upload_document_file(document_id: int, file: UploadFile = File(...), db: Session = Depends(get_db), _: dict = Depends(require_permission("document"))) -> dict:
    document = db.query(models.Document).filter(models.Document.id == document_id).first()
    if not document:
        raise HTTPException(status_code=404, detail="Document not found")
    if document.status == "已发布":
        raise HTTPException(status_code=409, detail="Released document cannot be modified")
    if document.status == "审批中":
        raise HTTPException(status_code=409, detail="Document in review cannot be modified")
    os.makedirs(FILE_UPLOAD_DIR, exist_ok=True)
    file_content = await file.read()
    file_size = len(file_content)
    safe_name = f"{document_id}_{file.filename}"
    file_path = os.path.join(FILE_UPLOAD_DIR, safe_name)
    with open(file_path, "wb") as f:
        f.write(file_content)
    document.file_name = file.filename
    document.file_path = file_path
    document.file_size = file_size
    document.file_type = file.content_type or ""
    document.updated_at = today_text()
    db.commit()
    return {"id": document.id, "file_name": document.file_name, "file_size": document.file_size, "file_type": document.file_type}


@app.get("/api/documents/{document_id}/download")
def download_document_file(document_id: int, db: Session = Depends(get_db)):
    document = db.query(models.Document).filter(models.Document.id == document_id).first()
    if not document:
        raise HTTPException(status_code=404, detail="Document not found")
    if not document.file_path or not os.path.exists(document.file_path):
        raise HTTPException(status_code=404, detail="File not found")
    with open(document.file_path, "rb") as f:
        content = f.read()
    return Response(
        content=content,
        media_type=document.file_type or "application/octet-stream",
        headers={"Content-Disposition": f'attachment; filename="{document.file_name}"'},
    )


@app.get("/api/requirements")
def requirements(page: int = 1, page_size: int = 20, keyword: str = "", db: Session = Depends(get_db)) -> dict:
    q = db.query(models.Requirement).options(selectinload(models.Requirement.product))
    if keyword:
        kw = f"%{keyword}%"
        q = q.filter(models.Requirement.req_no.ilike(kw) | models.Requirement.title.ilike(kw) | models.Requirement.category.ilike(kw))
    total = q.count()
    rows = q.order_by(models.Requirement.id).offset((page - 1) * page_size).limit(page_size).all()
    return {
        "items": [
            {"id": row.id, "req_no": row.req_no, "source": row.source, "category": row.category, "title": row.title, "priority": row.priority, "status": row.status, "owner": row.owner, "acceptance_criteria": row.acceptance_criteria, "product_model": row.product.model, "product_id": row.product_id}
            for row in rows
        ],
        "total": total, "page": page, "page_size": page_size,
    }


@app.post("/api/requirements", status_code=201)
def create_requirement(payload: RequirementPayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("requirement"))) -> dict:
    ensure_product_exists(db, payload.product_id)
    requirement = models.Requirement(**payload.model_dump())
    db.add(requirement)
    commit_or_409(db, "Requirement number already exists")
    db.refresh(requirement)
    return {
        "id": requirement.id,
        "req_no": requirement.req_no,
        "source": requirement.source,
        "category": requirement.category,
        "title": requirement.title,
        "priority": requirement.priority,
        "status": requirement.status,
        "owner": requirement.owner,
        "acceptance_criteria": requirement.acceptance_criteria,
        "product_model": requirement.product.model,
        "product_id": requirement.product_id,
    }


@app.put("/api/requirements/{requirement_id}")
def update_requirement(requirement_id: int, payload: RequirementUpdatePayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("requirement"))) -> dict:
    requirement = db.query(models.Requirement).options(selectinload(models.Requirement.product)).filter(models.Requirement.id == requirement_id).first()
    if not requirement:
        raise HTTPException(status_code=404, detail="Requirement not found")
    if payload.product_id is not None:
        ensure_product_exists(db, payload.product_id)
    update_model(requirement, payload)
    commit_or_409(db, "Requirement number already exists")
    db.refresh(requirement)
    return {
        "id": requirement.id,
        "req_no": requirement.req_no,
        "source": requirement.source,
        "category": requirement.category,
        "title": requirement.title,
        "priority": requirement.priority,
        "status": requirement.status,
        "owner": requirement.owner,
        "acceptance_criteria": requirement.acceptance_criteria,
        "product_model": requirement.product.model,
        "product_id": requirement.product_id,
    }


@app.delete("/api/requirements/{requirement_id}")
def delete_requirement(requirement_id: int, db: Session = Depends(get_db), _: dict = Depends(require_permission("requirement"))) -> dict:
    requirement = db.query(models.Requirement).filter(models.Requirement.id == requirement_id).first()
    if not requirement:
        raise HTTPException(status_code=404, detail="Requirement not found")
    if requirement.status in ["已确认", "已发布"]:
        raise HTTPException(status_code=409, detail="Confirmed requirement cannot be deleted")
    db.delete(requirement)
    db.commit()
    return {"deleted": True}


@app.get("/api/requirements/{requirement_id}/trace")
def requirement_trace(requirement_id: int, db: Session = Depends(get_db)) -> dict:
    """需求追溯链路：需求 -> 产品 -> BOM/变更/项目/文档/工艺"""
    requirement = (
        db.query(models.Requirement)
        .options(selectinload(models.Requirement.product))
        .filter(models.Requirement.id == requirement_id)
        .first()
    )
    if not requirement:
        raise HTTPException(status_code=404, detail="Requirement not found")
    product = requirement.product
    if not product:
        return {"requirement": {"req_no": requirement.req_no, "title": requirement.title}, "product": None, "boms": [], "changes": [], "projects": [], "documents": [], "routes": []}
    # 聚合该产品下游对象
    boms = (
        db.query(models.BomHeader)
        .filter(models.BomHeader.product_id == product.id)
        .order_by(models.BomHeader.id.desc())
        .all()
    )
    changes = (
        db.query(models.Change)
        .filter(models.Change.product_id == product.id)
        .order_by(models.Change.id.desc())
        .limit(20)
        .all()
    )
    documents = (
        db.query(models.Document)
        .filter(models.Document.product_id == product.id)
        .order_by(models.Document.id.desc())
        .limit(20)
        .all()
    )
    routes = (
        db.query(models.ProcessRoute)
        .filter(models.ProcessRoute.product_id == product.id)
        .order_by(models.ProcessRoute.id.desc())
        .all()
    )
    # 项目通过 product_model 字符串关联
    projects = (
        db.query(models.Project)
        .filter(models.Project.product_model == product.model)
        .order_by(models.Project.id.desc())
        .all()
    )
    return {
        "requirement": {
            "id": requirement.id,
            "req_no": requirement.req_no,
            "title": requirement.title,
            "source": requirement.source,
            "category": requirement.category,
            "priority": requirement.priority,
            "status": requirement.status,
            "owner": requirement.owner,
            "acceptance_criteria": requirement.acceptance_criteria,
        },
        "product": {
            "id": product.id,
            "model": product.model,
            "name": product.name,
            "lifecycle": product.lifecycle,
            "version": product.version,
            "readiness": product.readiness,
        },
        "boms": [{"id": b.id, "type": b.bom_type, "version": b.version, "status": b.status, "owner": b.owner, "release_date": b.release_date} for b in boms],
        "changes": [{"id": c.id, "change_no": c.change_no, "title": c.title, "status": c.status, "priority": c.priority} for c in changes],
        "documents": [{"id": d.id, "doc_no": d.doc_no, "title": d.title, "category": d.category, "version": d.version, "status": d.status} for d in documents],
        "routes": [{"id": r.id, "route_no": r.route_no, "name": r.name, "version": r.version, "status": r.status} for r in routes],
        "projects": [{"id": p.id, "project_no": p.project_no, "name": p.name, "phase": p.phase, "progress": p.progress, "owner": p.owner, "risk_level": p.risk_level} for p in projects],
    }


@app.get("/api/baselines")
def baselines(page: int = 1, page_size: int = 20, keyword: str = "", db: Session = Depends(get_db)) -> dict:
    q = (
        db.query(models.ProductBaseline)
        .options(selectinload(models.ProductBaseline.product), selectinload(models.ProductBaseline.items))
    )
    if keyword:
        kw = f"%{keyword}%"
        q = q.filter(models.ProductBaseline.baseline_no.ilike(kw) | models.ProductBaseline.name.ilike(kw))
    total = q.count()
    rows = q.order_by(models.ProductBaseline.id).offset((page - 1) * page_size).limit(page_size).all()
    return {
        "items": [
            {
                "id": row.id,
                "baseline_no": row.baseline_no,
                "name": row.name,
                "product_model": row.product.model,
                "version": row.version,
                "status": row.status,
                "created_by": row.created_by,
                "released_at": row.released_at,
                "items": [
                    {
                        "id": item.id,
                        "item_type": item.item_type,
                        "item_no": item.item_no,
                        "title": item.title,
                        "version": item.version,
                        "status": item.status,
                    }
                    for item in row.items
                ],
            }
            for row in rows
        ],
        "total": total,
        "page": page,
        "page_size": page_size,
    }


@app.get("/api/workbench")
def workbench(db: Session = Depends(get_db), context: dict = Depends(current_user_context)) -> dict:
    user = context["user"]
    user_display = user.display_name

    approvals = (
        db.query(models.Approval)
        .join(models.Change)
        .join(models.Product)
        .filter(models.Approval.status.in_(["处理中", "待处理", "流转中"]))
        .order_by(models.Approval.id)
        .limit(12)
        .all()
    )
    gate_rows = db.query(models.Product).order_by(models.Product.readiness).limit(8).all()

    my_changes = (
        db.query(models.Change)
        .join(models.Product)
        .filter(models.Change.owner == user_display)
        .order_by(models.Change.id.desc())
        .limit(10)
        .all()
    )

    my_projects = (
        db.query(models.Project)
        .filter(models.Project.owner == user_display)
        .order_by(models.Project.id.desc())
        .limit(10)
        .all()
    )

    my_tasks = (
        db.query(models.ProjectTask)
        .join(models.Project)
        .filter(models.ProjectTask.owner == user_display, models.ProjectTask.status != "已完成")
        .order_by(models.ProjectTask.due_date)
        .limit(10)
        .all()
    )

    my_quality_issues = (
        db.query(models.QualityIssue)
        .filter(models.QualityIssue.owner == user_display, models.QualityIssue.status != "已关闭")
        .order_by(models.QualityIssue.id.desc())
        .limit(10)
        .all()
    )

    recent_changes = (
        db.query(models.Change)
        .join(models.Product)
        .filter(models.Change.status.in_(["审批中", "执行中", "已关闭"]))
        .order_by(models.Change.id.desc())
        .limit(8)
        .all()
    )

    pending_workflow_tasks = (
        db.query(models.WorkflowTask)
        .join(models.WorkflowInstance)
        .filter(models.WorkflowTask.status == "待处理")
        .order_by(models.WorkflowTask.id.desc())
        .limit(8)
        .all()
    )

    return {
        "todo_approvals": [
            {
                "id": row.id,
                "change_no": row.change.change_no,
                "product_model": row.change.product.model,
                "step_name": row.step_name,
                "approver": row.approver,
                "status": row.status,
                "priority": row.change.priority,
                "submitted_at": row.change.submitted_at,
            }
            for row in approvals
        ],
        "stage_gates": [
            {
                "id": product.id,
                "model": product.model,
                "lifecycle": product.lifecycle,
                "owner": product.owner,
                "readiness": product.readiness,
                "next_gate": "量产发布" if product.readiness >= 80 else "资料齐套评审",
                "blocker": "可靠性报告/变更签核未闭环" if product.readiness < 80 else "客户承认书归档",
            }
            for product in gate_rows
        ],
        "my_changes": [
            {
                "id": c.id,
                "change_no": c.change_no,
                "title": c.title,
                "product_model": c.product.model,
                "status": c.status,
                "priority": c.priority,
                "submitted_at": c.submitted_at,
            }
            for c in my_changes
        ],
        "my_projects": [
            {
                "id": p.id,
                "project_no": p.project_no,
                "name": p.name,
                "product_model": p.product_model,
                "phase": p.phase,
                "progress": p.progress,
                "end_date": p.end_date,
                "risk_level": p.risk_level,
            }
            for p in my_projects
        ],
        "my_tasks": [
            {
                "id": t.id,
                "name": t.name,
                "phase": t.phase,
                "status": t.status,
                "due_date": t.due_date,
                "project_no": t.project.project_no if t.project else "",
            }
            for t in my_tasks
        ],
        "my_quality_issues": [
            {
                "id": i.id,
                "issue_no": i.issue_no,
                "title": i.title,
                "product_model": i.product_model,
                "severity": i.severity,
                "status": i.status,
            }
            for i in my_quality_issues
        ],
        "recent_changes": [
            {
                "id": c.id,
                "change_no": c.change_no,
                "title": c.title,
                "product_model": c.product.model,
                "status": c.status,
                "priority": c.priority,
                "submitted_at": c.submitted_at,
            }
            for c in recent_changes
        ],
        "pending_workflow_tasks": [
            {
                "id": t.id,
                "instance_id": t.instance_id,
                "node_name": t.node_name,
                "role_name": t.role_name,
                "status": t.status,
                "object_type": t.instance.object_type if t.instance else "",
                "object_no": t.instance.object_no if t.instance else "",
                "title": t.instance.title if t.instance else "",
            }
            for t in pending_workflow_tasks
        ],
    }


@app.get("/api/workbench/calendar")
def workbench_calendar(month: str, db: Session = Depends(get_db), context: dict = Depends(current_user_context)) -> dict:
    """工作台任务日历：聚合当前用户在该月内截止的项目任务/ECA动作/项目交付物"""
    user = context["user"]
    user_display = user.display_name
    # month 格式 YYYY-MM，构造当月起止
    try:
        year, mon = month.split("-")
        year_i, mon_i = int(year), int(mon)
        start = f"{year_i:04d}-{mon_i:02d}-01"
        if mon_i == 12:
            end = f"{year_i + 1:04d}-01-01"
        else:
            end = f"{year_i:04d}-{mon_i + 1:02d}-01"
    except Exception:
        raise HTTPException(status_code=400, detail="month must be YYYY-MM")

    items: list[dict] = []

    # 项目任务
    tasks = (
        db.query(models.ProjectTask, models.Project)
        .join(models.Project, models.ProjectTask.project_id == models.Project.id)
        .filter(
            models.ProjectTask.owner == user_display,
            models.ProjectTask.due_date >= start,
            models.ProjectTask.due_date < end,
        )
        .order_by(models.ProjectTask.due_date)
        .all()
    )
    for t, proj in tasks:
        items.append({
            "type": "项目任务",
            "no": t.name,
            "title": t.name,
            "owner": t.owner,
            "status": t.status,
            "due_date": t.due_date,
            "source_no": proj.project_no if proj else "",
            "source_name": proj.name if proj else "",
        })

    # ECA 动作
    actions = (
        db.query(models.ChangeAction)
        .filter(
            models.ChangeAction.owner == user_display,
            models.ChangeAction.due_date >= start,
            models.ChangeAction.due_date < end,
            models.ChangeAction.due_date != "",
        )
        .order_by(models.ChangeAction.due_date)
        .all()
    )
    for a in actions:
        change = db.query(models.Change).filter(models.Change.id == a.change_id).first()
        items.append({
            "type": "ECA动作",
            "no": a.action_no,
            "title": a.target_object or a.action_no,
            "owner": a.owner,
            "status": a.status,
            "due_date": a.due_date,
            "source_no": change.change_no if change else "",
            "source_name": change.title if change else "",
        })

    # 项目交付物
    deliverables = (
        db.query(models.ProjectDeliverable, models.Project)
        .join(models.Project, models.ProjectDeliverable.project_id == models.Project.id)
        .filter(
            models.ProjectDeliverable.owner == user_display,
            models.ProjectDeliverable.due_date >= start,
            models.ProjectDeliverable.due_date < end,
            models.ProjectDeliverable.due_date != "",
        )
        .order_by(models.ProjectDeliverable.due_date)
        .all()
    )
    for d, proj in deliverables:
        items.append({
            "type": "项目交付物",
            "no": d.name,
            "title": d.name,
            "owner": d.owner,
            "status": d.status,
            "due_date": d.due_date,
            "source_no": proj.project_no if proj else "",
            "source_name": proj.name if proj else "",
        })

    # 统计
    by_type = {}
    by_status = {}
    overdue = 0
    today_str = datetime.now().strftime("%Y-%m-%d")
    for it in items:
        by_type[it["type"]] = by_type.get(it["type"], 0) + 1
        by_status[it["status"]] = by_status.get(it["status"], 0) + 1
        if it["due_date"] < today_str and it["status"] not in ["已完成", "已关闭"]:
            overdue += 1

    return {
        "month": month,
        "items": items,
        "summary": {
            "total": len(items),
            "overdue": overdue,
            "by_type": [{"name": k, "value": v} for k, v in by_type.items()],
            "by_status": [{"name": k, "value": v} for k, v in by_status.items()],
        },
    }


@app.get("/api/workbench/notifications")
def workbench_notifications(action: str | None = None, limit: int = 50, db: Session = Depends(get_db)) -> dict:
    """工作台消息通知：基于操作日志聚合关键事件（发布/关闭/驳回/提交/删除/失败）"""
    # 关键动作白名单
    key_actions = ["发布", "关闭", "驳回", "提交", "删除", "新增", "编辑"]
    q = db.query(models.OperationLog).filter(models.OperationLog.action.in_(key_actions))
    if action:
        q = q.filter(models.OperationLog.action == action)
    rows = q.order_by(models.OperationLog.id.desc()).limit(min(limit, 200)).all()
    items = [
        {
            "id": row.id,
            "action": row.action,
            "object_type": row.object_type,
            "object_no": row.object_no,
            "summary": row.summary,
            "operated_by": row.operated_by,
            "operated_at": row.operated_at,
            "level": _notify_level(row.action),
        }
        for row in rows
    ]
    # 按动作统计
    by_action = {}
    for it in items:
        by_action[it["action"]] = by_action.get(it["action"], 0) + 1
    return {
        "items": items,
        "total": len(items),
        "by_action": [{"name": k, "value": v} for k, v in by_action.items()],
    }


def _notify_level(action: str) -> str:
    """通知级别：danger/warning/info/success"""
    if action in ["删除", "驳回"]:
        return "danger"
    if action in ["关闭", "发布"]:
        return "success"
    if action in ["提交"]:
        return "warning"
    return "info"


@app.get("/api/workbench/closure-check")
def workbench_closure_check(db: Session = Depends(get_db)) -> dict:
    """业务闭环验证：按产品检查 9 个环节的数据完整性，定位断点"""
    products = db.query(models.Product).order_by(models.Product.id).all()
    # 环节定义：(key, label)
    stages = [
        ("requirement", "需求规格"),
        ("product_version", "产品版本"),
        ("bom", "BOM"),
        ("process_route", "工艺路线"),
        ("document", "文档"),
        ("change", "工程变更"),
        ("project", "项目"),
        ("quality", "质量追溯"),
        ("integration", "下游同步"),
    ]

    product_rows = []
    full_closed = 0
    total_breakpoints = 0

    for p in products:
        req_count = db.query(models.Requirement).filter(models.Requirement.product_id == p.id).count()
        version_count = db.query(models.ProductVersion).filter(models.ProductVersion.product_id == p.id).count()
        bom_count = db.query(models.BomHeader).filter(models.BomHeader.product_id == p.id).count()
        route_count = db.query(models.ProcessRoute).filter(models.ProcessRoute.product_id == p.id).count()
        doc_count = db.query(models.Document).filter(models.Document.product_id == p.id).count()
        change_count = db.query(models.Change).filter(models.Change.product_id == p.id).count()
        # 项目通过 product_model 关联
        project_count = db.query(models.Project).filter(models.Project.product_model == p.model).count()
        quality_count = db.query(models.QualityLot).filter(models.QualityLot.product_id == p.id).count()
        # 集成：该产品相关对象产生的集成任务
        integration_count = db.query(models.IntegrationJob).filter(models.IntegrationJob.product_model == p.model).count()

        counts = {
            "requirement": req_count,
            "product_version": version_count,
            "bom": bom_count,
            "process_route": route_count,
            "document": doc_count,
            "change": change_count,
            "project": project_count,
            "quality": quality_count,
            "integration": integration_count,
        }
        # 状态：有数据=ok，无数据=gap，有数据但发布率低=warn
        stage_status = {}
        breakpoints = 0
        for key, _label in stages:
            cnt = counts[key]
            if cnt > 0:
                stage_status[key] = "ok"
            else:
                stage_status[key] = "gap"
                breakpoints += 1

        if breakpoints == 0:
            full_closed += 1
        total_breakpoints += breakpoints

        product_rows.append({
            "id": p.id,
            "model": p.model,
            "name": p.name,
            "lifecycle": p.lifecycle,
            "owner": p.owner,
            "readiness": p.readiness,
            "counts": counts,
            "stage_status": stage_status,
            "breakpoints": breakpoints,
            "closed": breakpoints == 0,
        })

    # 环节维度的覆盖率
    stage_coverage = []
    for key, label in stages:
        ok_count = sum(1 for r in product_rows if r["stage_status"][key] == "ok")
        stage_coverage.append({
            "key": key,
            "label": label,
            "ok_count": ok_count,
            "total": len(products),
            "rate": round((ok_count / len(products)) * 100) if products else 0,
        })

    return {
        "summary": {
            "product_total": len(products),
            "full_closed": full_closed,
            "closure_rate": round((full_closed / len(products)) * 100) if products else 0,
            "total_breakpoints": total_breakpoints,
        },
        "stages": [{"key": k, "label": l} for k, l in stages],
        "stage_coverage": stage_coverage,
        "products": product_rows,
    }


@app.get("/api/process-routes")
def process_routes(page: int = 1, page_size: int = 20, keyword: str = "", db: Session = Depends(get_db)) -> dict:
    q = db.query(models.ProcessRoute).options(selectinload(models.ProcessRoute.product), selectinload(models.ProcessRoute.steps))
    if keyword:
        kw = f"%{keyword}%"
        q = q.filter(models.ProcessRoute.route_no.ilike(kw) | models.ProcessRoute.name.ilike(kw))
    total = q.count()
    rows = q.order_by(models.ProcessRoute.id).offset((page - 1) * page_size).limit(page_size).all()
    return {"items": [serialize_process_route(row) for row in rows], "total": total, "page": page, "page_size": page_size}


@app.post("/api/process-routes", status_code=201)
def create_process_route(payload: ProcessRoutePayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("process"))) -> dict:
    ensure_product_exists(db, payload.product_id)
    if payload.status == "已发布":
        raise HTTPException(status_code=409, detail="Process route must be released through approval")
    route = models.ProcessRoute(**payload.model_dump())
    db.add(route)
    commit_or_409(db, "Process route number already exists")
    db.refresh(route)
    route = (
        db.query(models.ProcessRoute)
        .options(selectinload(models.ProcessRoute.product), selectinload(models.ProcessRoute.steps))
        .filter(models.ProcessRoute.id == route.id)
        .first()
    )
    assert route is not None
    return serialize_process_route(route)


@app.put("/api/process-routes/{route_id}")
def update_process_route(route_id: int, payload: ProcessRouteUpdatePayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("process"))) -> dict:
    route = (
        db.query(models.ProcessRoute)
        .options(selectinload(models.ProcessRoute.product), selectinload(models.ProcessRoute.steps))
        .filter(models.ProcessRoute.id == route_id)
        .first()
    )
    if not route:
        raise HTTPException(status_code=404, detail="Process route not found")
    ensure_route_editable(route)
    if payload.status == "已发布":
        raise HTTPException(status_code=409, detail="Process route must be released through approval")
    if payload.product_id is not None:
        ensure_product_exists(db, payload.product_id)
    update_model(route, payload)
    commit_or_409(db, "Process route number already exists")
    route = (
        db.query(models.ProcessRoute)
        .options(selectinload(models.ProcessRoute.product), selectinload(models.ProcessRoute.steps))
        .filter(models.ProcessRoute.id == route_id)
        .first()
    )
    assert route is not None
    return serialize_process_route(route)


@app.delete("/api/process-routes/{route_id}")
def delete_process_route(route_id: int, db: Session = Depends(get_db), _: dict = Depends(require_permission("process"))) -> dict:
    route = db.query(models.ProcessRoute).filter(models.ProcessRoute.id == route_id).first()
    if not route:
        raise HTTPException(status_code=404, detail="Process route not found")
    ensure_route_editable(route)
    db.delete(route)
    db.commit()
    return {"ok": True}


@app.post("/api/process-routes/{route_id}/steps", status_code=201)
def create_process_step(route_id: int, payload: ProcessStepPayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("process"))) -> dict:
    route = (
        db.query(models.ProcessRoute)
        .options(selectinload(models.ProcessRoute.product), selectinload(models.ProcessRoute.steps))
        .filter(models.ProcessRoute.id == route_id)
        .first()
    )
    if not route:
        raise HTTPException(status_code=404, detail="Process route not found")
    ensure_route_editable(route)
    if any(step.sequence == payload.sequence for step in route.steps):
        raise HTTPException(status_code=409, detail="Process step sequence already exists")
    step = models.ProcessStep(route_id=route_id, **payload.model_dump())
    db.add(step)
    db.commit()
    db.refresh(step)
    return {"id": step.id, "sequence": step.sequence, "stage": step.stage, "operation": step.operation, "key_params": step.key_params, "owner": step.owner, "status": step.status}


@app.put("/api/process-steps/{step_id}")
def update_process_step(step_id: int, payload: ProcessStepUpdatePayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("process"))) -> dict:
    step = (
        db.query(models.ProcessStep)
        .options(selectinload(models.ProcessStep.route))
        .filter(models.ProcessStep.id == step_id)
        .first()
    )
    if not step:
        raise HTTPException(status_code=404, detail="Process step not found")
    ensure_route_editable(step.route)
    if payload.sequence is not None:
        exists = (
            db.query(models.ProcessStep.id)
            .filter(models.ProcessStep.route_id == step.route_id, models.ProcessStep.sequence == payload.sequence, models.ProcessStep.id != step.id)
            .first()
        )
        if exists:
            raise HTTPException(status_code=409, detail="Process step sequence already exists")
    update_model(step, payload)
    db.commit()
    db.refresh(step)
    return {"id": step.id, "sequence": step.sequence, "stage": step.stage, "operation": step.operation, "key_params": step.key_params, "owner": step.owner, "status": step.status}


@app.delete("/api/process-steps/{step_id}")
def delete_process_step(step_id: int, db: Session = Depends(get_db), _: dict = Depends(require_permission("process"))) -> dict:
    step = (
        db.query(models.ProcessStep)
        .options(selectinload(models.ProcessStep.route))
        .filter(models.ProcessStep.id == step_id)
        .first()
    )
    if not step:
        raise HTTPException(status_code=404, detail="Process step not found")
    ensure_route_editable(step.route)
    db.delete(step)
    db.commit()
    return {"ok": True}


@app.post("/api/process-routes/{route_id}/submit")
def submit_process_route(route_id: int, payload: ProcessRouteActionPayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("process"))) -> dict:
    route = (
        db.query(models.ProcessRoute)
        .options(selectinload(models.ProcessRoute.product), selectinload(models.ProcessRoute.steps))
        .filter(models.ProcessRoute.id == route_id)
        .first()
    )
    if not route:
        raise HTTPException(status_code=404, detail="Process route not found")
    validate_process_route_ready(route)
    if route.status == "已发布":
        raise HTTPException(status_code=409, detail="Released process route cannot be submitted")
    if route.status == "审批中":
        raise HTTPException(status_code=409, detail="Process route is already in review")
    # ECA 生成对象校验：如果工艺路线有 source_route_id，说明是 ECA 升版生成的草案
    if route.source_route_id:
        validate_eca_generated_object_ready(db, "工艺路线", route.id, route.route_no)
    route.status = "审批中"
    db.commit()
    db.refresh(route)
    return serialize_process_route(route)


@app.post("/api/process-routes/{route_id}/approve")
def approve_process_route(route_id: int, payload: ProcessRouteActionPayload, db: Session = Depends(get_db), _: dict = Depends(require_permission(["approval", "process"]))) -> dict:
    route = (
        db.query(models.ProcessRoute)
        .options(selectinload(models.ProcessRoute.product), selectinload(models.ProcessRoute.steps))
        .filter(models.ProcessRoute.id == route_id)
        .first()
    )
    if not route:
        raise HTTPException(status_code=404, detail="Process route not found")
    validate_process_route_ready(route)
    if route.status not in ("审批中", "已发布"):
        raise HTTPException(status_code=409, detail="Process route must be submitted for review before approval")
    if route.source_route_id:
        validate_eca_generated_object_ready(db, "工艺路线", route.id, route.route_no)
    route.status = "已发布"
    route.release_date = today_text()
    create_integration_job(
        db,
        target_system="MES",
        object_type="工艺路线",
        object_no=route.route_no,
        product_model=route.product.model,
        triggered_by=route.route_no,
        message=f"工艺路线已发布，等待同步 MES 工艺流程、工序参数和站点控制。发布人：{payload.acted_by}",
    )
    db.commit()
    db.refresh(route)
    return serialize_process_route(route)


@app.get("/api/products/{product_id}/process-steps")
def product_process_steps(product_id: int, db: Session = Depends(get_db)) -> list[dict]:
    ensure_product_exists(db, product_id)
    rows = (
        db.query(models.ProcessStep)
        .join(models.ProcessRoute)
        .filter(models.ProcessRoute.product_id == product_id)
        .order_by(models.ProcessRoute.version.desc(), models.ProcessStep.sequence)
        .all()
    )
    return [
        {
            "id": step.id,
            "route_id": step.route_id,
            "sequence": step.sequence,
            "stage": step.stage,
            "operation": step.operation,
            "key_params": step.key_params,
            "owner": step.owner,
            "status": step.status,
            "label": f"{step.sequence}-{step.stage} / {step.operation}",
        }
        for step in rows
    ]


@app.get("/api/changes")
def changes(page: int = 1, page_size: int = 20, keyword: str = "", db: Session = Depends(get_db)) -> dict:
    q = db.query(models.Change).options(selectinload(models.Change.product), selectinload(models.Change.impacts), selectinload(models.Change.approvals))
    if keyword:
        kw = f"%{keyword}%"
        q = q.filter(models.Change.change_no.ilike(kw) | models.Change.title.ilike(kw))
    total = q.count()
    rows = q.order_by(models.Change.id).offset((page - 1) * page_size).limit(page_size).all()
    return {"items": [serialize_change(row, db) for row in rows], "total": total, "page": page, "page_size": page_size}


@app.post("/api/changes", status_code=201)
def create_change(payload: ChangePayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("change"))) -> dict:
    if not db.query(models.Product.id).filter(models.Product.id == payload.product_id).first():
        raise HTTPException(status_code=404, detail="Product not found")
    change = models.Change(**payload.model_dump())
    if not change.submitted_at and change.status != "草稿":
        change.submitted_at = today_text()
    db.add(change)
    commit_or_409(db, "Change number already exists")
    db.refresh(change)
    analyze_change_impacts(db, change)
    db.commit()
    change = db.query(models.Change).options(selectinload(models.Change.product), selectinload(models.Change.impacts), selectinload(models.Change.approvals)).filter(models.Change.id == change.id).first()
    return serialize_change(change, db)


@app.put("/api/changes/{change_id}")
def update_change(change_id: int, payload: ChangeUpdatePayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("change"))) -> dict:
    change = db.query(models.Change).filter(models.Change.id == change_id).first()
    if not change:
        raise HTTPException(status_code=404, detail="Change not found")
    if change.status not in {"草稿", "已驳回"}:
        raise HTTPException(status_code=409, detail="Only draft or rejected change can be edited")
    data = payload.model_dump(exclude_unset=True)
    if "product_id" in data and not db.query(models.Product.id).filter(models.Product.id == data["product_id"]).first():
        raise HTTPException(status_code=404, detail="Product not found")
    for key, value in data.items():
        setattr(change, key, value)
    commit_or_409(db, "Change number already exists")
    change = db.query(models.Change).options(selectinload(models.Change.product), selectinload(models.Change.impacts), selectinload(models.Change.approvals)).filter(models.Change.id == change_id).first()
    return serialize_change(change, db)


@app.delete("/api/changes/{change_id}")
def delete_change(change_id: int, db: Session = Depends(get_db), _: dict = Depends(require_permission("change"))) -> dict:
    change = db.query(models.Change).filter(models.Change.id == change_id).first()
    if not change:
        raise HTTPException(status_code=404, detail="Change not found")
    if change.status != "草稿":
        raise HTTPException(status_code=409, detail="Only draft changes can be deleted")
    db.delete(change)
    db.commit()
    return {"ok": True}


@app.post("/api/changes/{change_id}/submit")
def submit_change(change_id: int, db: Session = Depends(get_db), _: dict = Depends(require_permission("change"))) -> dict:
    change = db.query(models.Change).options(selectinload(models.Change.product), selectinload(models.Change.impacts), selectinload(models.Change.approvals)).filter(models.Change.id == change_id).first()
    if not change:
        raise HTTPException(status_code=404, detail="Change not found")
    if change.status not in ("草稿", "已驳回"):
        raise HTTPException(status_code=409, detail=f"Change in {change.status} status cannot be submitted")
    analyze_change_impacts(db, change)
    change.status = "审批中"
    change.submitted_at = change.submitted_at or today_text()
    start_workflow(
        db,
        template_code="WF-ECR-ECN",
        object_type="变更",
        object_id=change.id,
        object_no=change.change_no,
        title=change.title,
        product_model=change.product.model,
        started_by=change.owner or "系统用户",
    )
    db.commit()
    db.refresh(change)
    return serialize_change(change, db)


@app.post("/api/changes/{change_id}/analyze")
def analyze_change(change_id: int, db: Session = Depends(get_db), _: dict = Depends(require_permission("change"))) -> dict:
    change = db.query(models.Change).options(selectinload(models.Change.product), selectinload(models.Change.impacts), selectinload(models.Change.approvals)).filter(models.Change.id == change_id).first()
    if not change:
        raise HTTPException(status_code=404, detail="Change not found")
    analyze_change_impacts(db, change)
    db.commit()
    change = db.query(models.Change).options(selectinload(models.Change.product), selectinload(models.Change.impacts), selectinload(models.Change.approvals)).filter(models.Change.id == change_id).first()
    return serialize_change(change, db)


@app.post("/api/changes/{change_id}/impacts", status_code=201)
def create_change_impact(change_id: int, payload: ChangeImpactPayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("change"))) -> dict:
    if not db.query(models.Change.id).filter(models.Change.id == change_id).first():
        raise HTTPException(status_code=404, detail="Change not found")
    impact = models.ChangeImpact(change_id=change_id, **payload.model_dump())
    db.add(impact)
    db.commit()
    db.refresh(impact)
    return {"id": impact.id, "type": impact.impact_type, "impact_type": impact.impact_type, "target": impact.target, "risk": impact.risk, "action": impact.action}


@app.put("/api/change-impacts/{impact_id}")
def update_change_impact(impact_id: int, payload: ChangeImpactUpdatePayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("change"))) -> dict:
    impact = db.query(models.ChangeImpact).filter(models.ChangeImpact.id == impact_id).first()
    if not impact:
        raise HTTPException(status_code=404, detail="Change impact not found")
    update_model(impact, payload)
    db.commit()
    db.refresh(impact)
    return {"id": impact.id, "type": impact.impact_type, "impact_type": impact.impact_type, "target": impact.target, "risk": impact.risk, "action": impact.action}


@app.delete("/api/change-impacts/{impact_id}")
def delete_change_impact(impact_id: int, db: Session = Depends(get_db), _: dict = Depends(require_permission("change"))) -> dict:
    impact = db.query(models.ChangeImpact).filter(models.ChangeImpact.id == impact_id).first()
    if not impact:
        raise HTTPException(status_code=404, detail="Change impact not found")
    db.delete(impact)
    db.commit()
    return {"ok": True}


@app.post("/api/changes/{change_id}/actions", status_code=201)
def create_change_action(change_id: int, payload: ChangeActionPayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("change"))) -> dict:
    if not db.query(models.Change.id).filter(models.Change.id == change_id).first():
        raise HTTPException(status_code=404, detail="Change not found")
    count = db.query(models.ChangeAction).filter(models.ChangeAction.change_id == change_id).count() + 1
    action_no = payload.action_no or f"ECA-{change_id}-{count:02d}"
    action = models.ChangeAction(change_id=change_id, **{**payload.model_dump(), "action_no": action_no})
    validate_change_action_target(db, action)
    db.add(action)
    commit_or_409(db, "Change action number already exists")
    db.refresh(action)
    return serialize_change_action(action)


@app.get("/api/changes/{change_id}/revision-archive")
def change_revision_archive(change_id: int, db: Session = Depends(get_db)) -> list[dict]:
    if not db.query(models.Change.id).filter(models.Change.id == change_id).first():
        raise HTTPException(status_code=404, detail="Change not found")
    rows = (
        db.query(models.ChangeAction)
        .filter(models.ChangeAction.change_id == change_id, models.ChangeAction.generated_object_no != "")
        .order_by(models.ChangeAction.id)
        .all()
    )
    archive_rows = []
    for row in rows:
        gate = get_eca_generated_object_gate(db, row.target_type, row.generated_object_no) or {
            "ready": True,
            "action_no": row.action_no,
            "action_status": row.status,
            "change_no": "",
            "change_status": "",
            "message": "生成对象已满足提交发布条件。",
        }
        archive_rows.append({
            "action_no": row.action_no,
            "action_type": row.action_type,
            "target_type": row.target_type,
            "target_id": row.target_id,
            "source_object": row.target_object,
            "source_version": row.target_version,
            "effectivity_type": row.effectivity_type,
            "effectivity_scope": row.effectivity_scope,
            "effective_date": row.effective_date,
            "effective_batch": row.effective_batch,
            "generated_object_no": row.generated_object_no,
            "owner": row.owner,
            "status": row.status,
            "change_no": gate["change_no"],
            "change_status": gate["change_status"],
            "release_gate_status": "可提交" if gate["ready"] else "待变更闭环",
            "release_gate_message": gate["message"],
            "target_url": (
                f"/boms?highlight={row.target_id}" if row.target_type == "BOM"
                else f"/documents?highlight={row.target_id}" if row.target_type == "文档"
                else f"/process?highlight={row.target_id}" if row.target_type == "工艺路线"
                else ""
            ),
            "generated_url": (
                f"/boms?highlight={row.generated_object_no}" if "BOM" in (row.generated_object_no or "")
                else f"/documents?highlight={row.generated_object_no}" if "DOC-" in (row.generated_object_no or "")
                else f"/process?highlight={row.generated_object_no}" if "ROUTE" in (row.generated_object_no or "")
                else ""
            ),
        })
    return archive_rows


@app.put("/api/change-actions/{action_id}")
def update_change_action(action_id: int, payload: ChangeActionUpdatePayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("change"))) -> dict:
    action = db.query(models.ChangeAction).filter(models.ChangeAction.id == action_id).first()
    if not action:
        raise HTTPException(status_code=404, detail="Change action not found")
    if action.status == "已完成":
        raise HTTPException(status_code=409, detail="Completed action cannot be edited")
    update_model(action, payload)
    validate_change_action_target(db, action)
    commit_or_409(db, "Change action number already exists")
    db.refresh(action)
    return serialize_change_action(action)


@app.post("/api/change-actions/{action_id}/close")
def close_change_action(action_id: int, payload: ChangeActionClosePayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("change"))) -> dict:
    action = db.query(models.ChangeAction).filter(models.ChangeAction.id == action_id).first()
    if not action:
        raise HTTPException(status_code=404, detail="Change action not found")
    if action.status == "已完成":
        raise HTTPException(status_code=409, detail="ECA action is already closed")

    change = db.query(models.Change).filter(models.Change.id == action.change_id).first()
    if change and change.status != "执行中":
        raise HTTPException(status_code=409, detail=f"Change is in {change.status} status, actions can only be closed when change is executing")

    # 批次一致性校验：同一变更中相同批次的所有 ECA 动作必须使用一致格式
    if "批次" in (action.effectivity_type or ""):
        if not action.effective_batch:
            raise HTTPException(status_code=409, detail="Batch-based ECA action requires effective batch")
        siblings = (
            db.query(models.ChangeAction)
            .filter(
                models.ChangeAction.change_id == action.change_id,
                models.ChangeAction.id != action.id,
                models.ChangeAction.effective_batch != "",
            )
            .all()
        )
        for sibling in siblings:
            if sibling.effective_batch != action.effective_batch and "批次" in (sibling.effectivity_type or ""):
                raise HTTPException(
                    status_code=409,
                    detail=f"Batch inconsistency: action {action.action_no} batch '{action.effective_batch}' "
                    f"does not match sibling {sibling.action_no} batch '{sibling.effective_batch}'",
                )

    generated_object_no = apply_change_action_revision(db, action)
    action.status = "已完成"
    generated_note = f"，生成对象：{generated_object_no}" if generated_object_no else ""
    action.result = f"{payload.result}（关闭人：{payload.acted_by}，日期：{today_text()}{generated_note}）"
    closed_change = close_change_when_actions_done(db, action.change_id, payload.acted_by)
    db.commit()
    db.refresh(action)
    result = serialize_change_action(action)
    result["closed_change"] = closed_change
    return result


@app.get("/api/change-actions")
def change_actions(page: int = 1, page_size: int = 20, keyword: str = "", db: Session = Depends(get_db)) -> dict:
    q = db.query(models.ChangeAction).join(models.Change).join(models.Product)
    if keyword:
        kw = f"%{keyword}%"
        q = q.filter(models.ChangeAction.action_no.ilike(kw) | models.ChangeAction.target_object.ilike(kw))
    total = q.count()
    rows = q.order_by(models.ChangeAction.id).offset((page - 1) * page_size).limit(page_size).all()
    return {
        "items": [
            {
                "id": row.id,
                "action_no": row.action_no,
                "change_no": row.change.change_no,
                "product_model": row.change.product.model,
                "action_type": row.action_type,
                "target_type": row.target_type,
                "target_id": row.target_id,
                "target_version": row.target_version,
                "target_object": row.target_object,
                "effectivity_type": row.effectivity_type,
                "effectivity_scope": row.effectivity_scope,
                "effective_date": row.effective_date,
                "effective_batch": row.effective_batch,
                "generated_object_no": row.generated_object_no,
                "department": row.department,
                "owner": row.owner,
                "status": row.status,
                "due_date": row.due_date,
                "result": row.result,
            }
            for row in rows
        ],
        "total": total,
        "page": page,
        "page_size": page_size,
    }


@app.get("/api/products/{product_id}/effectivity-batches")
def product_effectivity_batches(product_id: int, db: Session = Depends(get_db)) -> list[dict]:
    """按产品聚合所有 ECA 动作的生效批次，用于批次控制总览。"""
    product = db.query(models.Product).filter(models.Product.id == product_id).first()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")

    actions = (
        db.query(models.ChangeAction)
        .join(models.Change)
        .filter(models.Change.product_id == product_id)
        .order_by(models.ChangeAction.id)
        .all()
    )

    batch_map: dict[str, dict] = {}
    no_batch_rows: list[dict] = []
    for action in actions:
        change = db.query(models.Change).filter(models.Change.id == action.change_id).first()
        change_no = change.change_no if change else ""
        change_status = change.status if change else ""
        gate = get_eca_generated_object_gate(db, action.target_type, action.generated_object_no) if action.generated_object_no else None
        row = {
            "action_no": action.action_no,
            "action_type": action.action_type,
            "target_type": action.target_type,
            "target_object": action.target_object,
            "target_version": action.target_version,
            "generated_object_no": action.generated_object_no,
            "effectivity_type": action.effectivity_type,
            "effectivity_scope": action.effectivity_scope,
            "effective_date": action.effective_date,
            "effective_batch": action.effective_batch,
            "owner": action.owner,
            "status": action.status,
            "due_date": action.due_date,
            "change_no": change_no,
            "change_status": change_status,
            "release_gate_status": (gate["ready"] if gate else True),
            "release_gate_message": (gate["message"] if gate else ""),
        }
        batch_key = action.effective_batch or ""
        if "批次" in (action.effectivity_type or "") and batch_key:
            if batch_key not in batch_map:
                batch_map[batch_key] = {
                    "effective_batch": batch_key,
                    "effectivity_type": action.effectivity_type,
                    "effective_date": action.effective_date,
                    "actions": [],
                    "change_nos": set(),
                    "pending_count": 0,
                    "done_count": 0,
                }
            batch_map[batch_key]["actions"].append(row)
            batch_map[batch_key]["change_nos"].add(change_no)
            if action.status == "已完成":
                batch_map[batch_key]["done_count"] += 1
            else:
                batch_map[batch_key]["pending_count"] += 1
            if not batch_map[batch_key]["effective_date"] and action.effective_date:
                batch_map[batch_key]["effective_date"] = action.effective_date
        else:
            no_batch_rows.append(row)

    batches = []
    for batch_key, info in batch_map.items():
        all_done = info["pending_count"] == 0 and info["done_count"] > 0
        batches.append({
            "effective_batch": info["effective_batch"],
            "effectivity_type": info["effectivity_type"],
            "effective_date": info["effective_date"],
            "change_nos": sorted(info["change_nos"]),
            "action_count": len(info["actions"]),
            "pending_count": info["pending_count"],
            "done_count": info["done_count"],
            "batch_status": "已生效" if all_done else "执行中" if info["done_count"] > 0 else "待执行",
            "actions": info["actions"],
        })
    batches.sort(key=lambda b: b["effective_batch"])

    return {
        "product_id": product_id,
        "product_model": product.model,
        "batches": batches,
        "no_batch_actions": no_batch_rows,
    }


@app.get("/api/boms/{bom_id}/version-history")
def bom_version_history(bom_id: int, db: Session = Depends(get_db)) -> list[dict]:
    """追溯 BOM 完整版本链路：来源版本、生成变更单、ECA 动作和发布门状态。"""
    bom = db.query(models.BomHeader).options(selectinload(models.BomHeader.product)).filter(models.BomHeader.id == bom_id).first()
    if not bom:
        raise HTTPException(status_code=404, detail="BOM not found")

    chain: list[models.BomHeader] = []
    current = bom
    visited = set()
    while current and current.id not in visited:
        visited.add(current.id)
        chain.append(current)
        if current.source_bom_id:
            current = db.query(models.BomHeader).filter(models.BomHeader.id == current.source_bom_id).first()
        else:
            break
    chain.reverse()

    history = []
    for item in chain:
        eca_action = (
            db.query(models.ChangeAction)
            .filter(
                models.ChangeAction.target_type == "BOM",
                models.ChangeAction.target_id == item.source_bom_id if item.source_bom_id else 0,
            )
            .first() if item.source_bom_id else None
        )
        change_no = ""
        change_status = ""
        action_no = ""
        effectivity_type = ""
        effective_batch = ""
        effective_date = ""
        release_gate = ""
        gate_message = ""
        if eca_action:
            change = db.query(models.Change).filter(models.Change.id == eca_action.change_id).first()
            change_no = change.change_no if change else ""
            change_status = change.status if change else ""
            action_no = eca_action.action_no
            effectivity_type = eca_action.effectivity_type
            effective_batch = eca_action.effective_batch
            effective_date = eca_action.effective_date
            if eca_action.generated_object_no:
                gate = get_eca_generated_object_gate(db, "BOM", eca_action.generated_object_no)
                if gate:
                    release_gate = "可提交" if gate["ready"] else "待变更闭环"
                    gate_message = gate["message"]

        generated_no = f"{item.bom_type}-{item.product.model}-{item.version}" if item.product else ""
        history.append({
            "id": item.id,
            "version": item.version,
            "status": item.status,
            "bom_type": item.bom_type,
            "object_no": generated_no,
            "owner": item.owner,
            "release_date": item.release_date,
            "effective_date": item.effective_date,
            "expiry_date": item.expiry_date,
            "effectivity_type": item.effectivity_type,
            "effective_batch": item.effective_batch,
            "source_bom_id": item.source_bom_id,
            "is_current": item.id == bom.id,
            "change_no": change_no,
            "change_status": change_status,
            "eca_action_no": action_no,
            "eca_effectivity_type": effectivity_type,
            "eca_effective_batch": effective_batch,
            "eca_effective_date": effective_date,
            "release_gate_status": release_gate,
            "release_gate_message": gate_message,
        })
    return history


@app.get("/api/documents/{document_id}/version-history")
def document_version_history(document_id: int, db: Session = Depends(get_db)) -> list[dict]:
    """追溯文档版本链路：同一产品同标题的版本序列及关联变更。"""
    doc = db.query(models.Document).filter(models.Document.id == document_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")

    siblings = (
        db.query(models.Document)
        .filter(
            models.Document.product_id == doc.product_id,
            models.Document.title == doc.title,
            models.Document.category == doc.category,
        )
        .order_by(models.Document.id)
        .all()
    )

    history = []
    for item in siblings:
        eca_action = (
            db.query(models.ChangeAction)
            .filter(
                models.ChangeAction.target_type == "文档",
                models.ChangeAction.generated_object_no == item.doc_no,
            )
            .first()
        )
        change_no = ""
        change_status = ""
        action_no = ""
        effectivity_type = ""
        effective_batch = ""
        effective_date = ""
        source_version = ""
        release_gate = ""
        gate_message = ""
        if eca_action:
            change = db.query(models.Change).filter(models.Change.id == eca_action.change_id).first()
            change_no = change.change_no if change else ""
            change_status = change.status if change else ""
            action_no = eca_action.action_no
            effectivity_type = eca_action.effectivity_type
            effective_batch = eca_action.effective_batch
            effective_date = eca_action.effective_date
            source_version = eca_action.target_version
            gate = get_eca_generated_object_gate(db, "文档", item.doc_no)
            if gate:
                release_gate = "可提交" if gate["ready"] else "待变更闭环"
                gate_message = gate["message"]

        history.append({
            "id": item.id,
            "doc_no": item.doc_no,
            "version": item.version,
            "status": item.status,
            "category": item.category,
            "owner": item.owner,
            "approval_status": item.approval_status,
            "updated_at": item.updated_at,
            "source_version": source_version,
            "is_current": item.id == doc.id,
            "change_no": change_no,
            "change_status": change_status,
            "eca_action_no": action_no,
            "eca_effectivity_type": effectivity_type,
            "eca_effective_batch": effective_batch,
            "eca_effective_date": effective_date,
            "release_gate_status": release_gate,
            "release_gate_message": gate_message,
        })
    return history


@app.get("/api/process-routes/{route_id}/version-history")
def process_route_version_history(route_id: int, db: Session = Depends(get_db)) -> list[dict]:
    """追溯工艺路线版本链路：来源路线、生成变更单和发布门状态。"""
    route = db.query(models.ProcessRoute).options(selectinload(models.ProcessRoute.product)).filter(models.ProcessRoute.id == route_id).first()
    if not route:
        raise HTTPException(status_code=404, detail="Process route not found")

    chain: list[models.ProcessRoute] = []
    current = route
    visited = set()
    while current and current.id not in visited:
        visited.add(current.id)
        chain.append(current)
        if current.source_route_id:
            current = db.query(models.ProcessRoute).filter(models.ProcessRoute.id == current.source_route_id).first()
        else:
            break
    chain.reverse()

    history = []
    for item in chain:
        eca_action = (
            db.query(models.ChangeAction)
            .filter(
                models.ChangeAction.target_type == "工艺路线",
                models.ChangeAction.target_id == item.source_route_id if item.source_route_id else 0,
            )
            .first() if item.source_route_id else None
        )
        change_no = ""
        change_status = ""
        action_no = ""
        effectivity_type = ""
        effective_batch = ""
        effective_date = ""
        release_gate = ""
        gate_message = ""
        if eca_action:
            change = db.query(models.Change).filter(models.Change.id == eca_action.change_id).first()
            change_no = change.change_no if change else ""
            change_status = change.status if change else ""
            action_no = eca_action.action_no
            effectivity_type = eca_action.effectivity_type
            effective_batch = eca_action.effective_batch
            effective_date = eca_action.effective_date
            if eca_action.generated_object_no:
                gate = get_eca_generated_object_gate(db, "工艺路线", eca_action.generated_object_no)
                if gate:
                    release_gate = "可提交" if gate["ready"] else "待变更闭环"
                    gate_message = gate["message"]

        history.append({
            "id": item.id,
            "route_no": item.route_no,
            "version": item.version,
            "status": item.status,
            "name": item.name,
            "owner": item.owner,
            "release_date": item.release_date,
            "effective_batch": item.effective_batch,
            "source_route_id": item.source_route_id,
            "is_current": item.id == route.id,
            "change_no": change_no,
            "change_status": change_status,
            "eca_action_no": action_no,
            "eca_effectivity_type": effectivity_type,
            "eca_effective_batch": effective_batch,
            "eca_effective_date": effective_date,
            "release_gate_status": release_gate,
            "release_gate_message": gate_message,
        })
    return history


@app.get("/api/integration-jobs")
def integration_jobs(
    status: str = "",
    target_system: str = "",
    object_type: str = "",
    keyword: str = "",
    page: int = 1,
    page_size: int = 20,
    db: Session = Depends(get_db),
) -> dict:
    query = db.query(models.IntegrationJob)
    if status:
        query = query.filter(models.IntegrationJob.status == status)
    if target_system:
        query = query.filter(models.IntegrationJob.target_system == target_system)
    if object_type:
        query = query.filter(models.IntegrationJob.object_type == object_type)
    if keyword:
        like_value = f"%{keyword}%"
        query = query.filter(
            (models.IntegrationJob.object_no.like(like_value))
            | (models.IntegrationJob.product_model.like(like_value))
            | (models.IntegrationJob.job_no.like(like_value))
        )
    total = query.count()
    rows = query.order_by(models.IntegrationJob.id.desc()).offset((page - 1) * page_size).limit(page_size).all()
    return {"items": [serialize_integration_job(row) for row in rows], "total": total, "page": page, "page_size": page_size}


@app.get("/api/integration-jobs/summary")
def integration_jobs_summary(db: Session = Depends(get_db)) -> dict:
    rows = db.query(models.IntegrationJob.status, func.count(models.IntegrationJob.id)).group_by(models.IntegrationJob.status).all()
    status_counts = {status: count for status, count in rows}
    target_rows = db.query(models.IntegrationJob.target_system, func.count(models.IntegrationJob.id)).group_by(models.IntegrationJob.target_system).all()
    return {
        "total": sum(status_counts.values()),
        "waiting": status_counts.get("等待", 0),
        "processing": status_counts.get("处理中", 0),
        "failed": status_counts.get("失败", 0),
        "success": status_counts.get("成功", 0),
        "target_systems": [{"target_system": target, "count": count} for target, count in target_rows],
    }


@app.post("/api/integration-jobs/{job_id}/start")
def start_integration_job(job_id: int, payload: IntegrationJobActionPayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("integration"))) -> dict:
    job = db.query(models.IntegrationJob).filter(models.IntegrationJob.id == job_id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Integration job not found")
    if job.status == "成功":
        raise HTTPException(status_code=409, detail="Successful integration job cannot be restarted")
    job.status = "处理中"
    job.attempt_count += 1
    job.last_sync_at = today_text()
    job.response_message = payload.response_message or f"{payload.acted_by} 已开始同步"
    db.commit()
    db.refresh(job)
    return serialize_integration_job(job)


@app.post("/api/integration-jobs/{job_id}/success")
def complete_integration_job(job_id: int, payload: IntegrationJobActionPayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("integration"))) -> dict:
    job = db.query(models.IntegrationJob).filter(models.IntegrationJob.id == job_id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Integration job not found")
    job.status = "成功"
    job.last_sync_at = today_text()
    job.external_id = payload.external_id or job.external_id
    job.response_message = payload.response_message or "下游系统已确认接收"
    db.commit()
    db.refresh(job)
    return serialize_integration_job(job)


@app.post("/api/integration-jobs/{job_id}/fail")
def fail_integration_job(job_id: int, payload: IntegrationJobActionPayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("integration"))) -> dict:
    job = db.query(models.IntegrationJob).filter(models.IntegrationJob.id == job_id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Integration job not found")
    job.status = "失败"
    job.last_sync_at = today_text()
    job.response_message = payload.response_message or "下游系统返回失败"
    db.commit()
    db.refresh(job)
    return serialize_integration_job(job)


@app.post("/api/integration-jobs/{job_id}/retry")
def retry_integration_job(job_id: int, payload: IntegrationJobActionPayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("integration"))) -> dict:
    job = db.query(models.IntegrationJob).filter(models.IntegrationJob.id == job_id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Integration job not found")
    if job.status == "成功":
        raise HTTPException(status_code=409, detail="Successful integration job cannot be retried")
    job.status = "等待"
    job.response_message = payload.response_message or f"{payload.acted_by} 已加入重试队列"
    db.commit()
    db.refresh(job)
    return serialize_integration_job(job)


@app.get("/api/admin/foundation/system-parameters")
def system_parameters(page: int = 1, page_size: int = 20, keyword: str = "", db: Session = Depends(get_db)) -> dict:
    q = db.query(models.SystemParameter)
    if keyword:
        kw = f"%{keyword}%"
        q = q.filter(
            models.SystemParameter.param_key.ilike(kw)
            | models.SystemParameter.param_value.ilike(kw)
            | models.SystemParameter.param_group.ilike(kw)
            | models.SystemParameter.description.ilike(kw)
        )
    total = q.count()
    rows = q.order_by(models.SystemParameter.param_group, models.SystemParameter.id).offset((page - 1) * page_size).limit(page_size).all()
    return {
        "items": [
            {"id": row.id, "param_key": row.param_key, "param_value": row.param_value, "param_group": row.param_group, "description": row.description}
            for row in rows
        ],
        "total": total,
        "page": page,
        "page_size": page_size,
    }


@app.post("/api/admin/foundation/system-parameters", status_code=201)
def create_system_parameter(payload: SystemParameterPayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("system"))) -> dict:
    row = models.SystemParameter(**payload.model_dump())
    db.add(row)
    commit_or_409(db, "System parameter key already exists")
    db.refresh(row)
    return {"id": row.id, "param_key": row.param_key, "param_value": row.param_value, "param_group": row.param_group, "description": row.description}


@app.put("/api/admin/foundation/system-parameters/{param_id}")
def update_system_parameter(param_id: int, payload: SystemParameterUpdatePayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("system"))) -> dict:
    row = db.query(models.SystemParameter).filter(models.SystemParameter.id == param_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="System parameter not found")
    update_model(row, payload)
    commit_or_409(db, "System parameter key already exists")
    db.refresh(row)
    return {"id": row.id, "param_key": row.param_key, "param_value": row.param_value, "param_group": row.param_group, "description": row.description}


@app.delete("/api/admin/foundation/system-parameters/{param_id}")
def delete_system_parameter(param_id: int, db: Session = Depends(get_db), _: dict = Depends(require_permission("system"))) -> dict:
    row = db.query(models.SystemParameter).filter(models.SystemParameter.id == param_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="System parameter not found")
    db.delete(row)
    db.commit()
    return {"ok": True}


@app.get("/api/audit-logs")
def audit_logs(page: int = 1, page_size: int = 20, keyword: str = "", object_type: str | None = None, action: str | None = None, db: Session = Depends(get_db)) -> dict:
    query = db.query(models.OperationLog)
    if object_type:
        query = query.filter(models.OperationLog.object_type == object_type)
    if action:
        query = query.filter(models.OperationLog.action == action)
    if keyword:
        kw = f"%{keyword}%"
        query = query.filter(models.OperationLog.object_no.ilike(kw) | models.OperationLog.summary.ilike(kw) | models.OperationLog.operated_by.ilike(kw))
    total = query.count()
    rows = query.order_by(models.OperationLog.id.desc()).offset((page - 1) * page_size).limit(page_size).all()
    return {
        "items": [
            {"id": row.id, "action": row.action, "object_type": row.object_type, "object_id": row.object_id, "object_no": row.object_no, "summary": row.summary, "operated_by": row.operated_by, "operated_at": row.operated_at}
            for row in rows
        ],
        "total": total,
        "page": page,
        "page_size": page_size,
    }


@app.get("/api/reports/completeness")
def report_completeness(db: Session = Depends(get_db)) -> dict:
    """数据完整度报表：产品资料/文档/BOM/工艺齐套率"""
    products = db.query(models.Product).all()
    total_products = len(products)
    product_rows = []
    doc_ready = 0
    bom_ready = 0
    route_ready = 0
    req_ready = 0
    for p in products:
        has_doc = db.query(models.Document).filter(models.Document.product_id == p.id).count() > 0
        has_bom = db.query(models.BomHeader).filter(models.BomHeader.product_id == p.id).count() > 0
        has_route = db.query(models.ProcessRoute).filter(models.ProcessRoute.product_id == p.id).count() > 0
        has_req = db.query(models.Requirement).filter(models.Requirement.product_id == p.id).count() > 0
        if has_doc:
            doc_ready += 1
        if has_bom:
            bom_ready += 1
        if has_route:
            route_ready += 1
        if has_req:
            req_ready += 1
        # 单产品完整度：四项全有视为齐套
        completeness = sum([has_doc, has_bom, has_route, has_req])
        product_rows.append({
            "id": p.id,
            "model": p.model,
            "name": p.name,
            "lifecycle": p.lifecycle,
            "owner": p.owner,
            "has_doc": has_doc,
            "has_bom": has_bom,
            "has_route": has_route,
            "has_req": has_req,
            "completeness": round(completeness * 25),  # 0/25/50/75/100
        })

    # 文档签核率
    doc_total = db.query(models.Document).count()
    doc_signed = db.query(models.Document).filter(models.Document.approval_status == "已签核").count()
    # BOM 已发布率
    bom_total = db.query(models.BomHeader).count()
    bom_released = db.query(models.BomHeader).filter(models.BomHeader.status == "已发布").count()
    # 工艺已发布率
    route_total = db.query(models.ProcessRoute).count()
    route_released = db.query(models.ProcessRoute).filter(models.ProcessRoute.status == "已发布").count()

    return {
        "summary": {
            "product_total": total_products,
            "doc_coverage": round((doc_ready / total_products) * 100) if total_products else 0,
            "bom_coverage": round((bom_ready / total_products) * 100) if total_products else 0,
            "route_coverage": round((route_ready / total_products) * 100) if total_products else 0,
            "req_coverage": round((req_ready / total_products) * 100) if total_products else 0,
            "doc_signed_rate": round((doc_signed / doc_total) * 100) if doc_total else 0,
            "bom_released_rate": round((bom_released / bom_total) * 100) if bom_total else 0,
            "route_released_rate": round((route_released / route_total) * 100) if route_total else 0,
            "full_ready_count": sum(1 for r in product_rows if r["completeness"] == 100),
        },
        "products": product_rows,
    }


@app.get("/api/reports/change-cycle")
def report_change_cycle(db: Session = Depends(get_db)) -> dict:
    """变更周期报表：ECR/ECO/ECN 状态分布、ECA 关闭率"""
    change_status_rows = db.query(models.Change.status, func.count(models.Change.id)).group_by(models.Change.status).all()
    change_type_rows = db.query(models.Change.change_type, func.count(models.Change.id)).group_by(models.Change.change_type).all()
    change_priority_rows = db.query(models.Change.priority, func.count(models.Change.id)).group_by(models.Change.priority).all()

    # ECA 关闭率
    eca_total = db.query(models.ChangeAction).count()
    eca_closed = db.query(models.ChangeAction).filter(models.ChangeAction.status == "已完成").count()
    eca_pending = db.query(models.ChangeAction).filter(models.ChangeAction.status != "已完成").count()

    # 按变更单聚合 ECA 完成情况
    changes = db.query(models.Change).order_by(models.Change.id.desc()).limit(20).all()
    change_rows = []
    for c in changes:
        actions = db.query(models.ChangeAction).filter(models.ChangeAction.change_id == c.id).all()
        total = len(actions)
        closed = sum(1 for a in actions if a.status == "已完成")
        change_rows.append({
            "id": c.id,
            "change_no": c.change_no,
            "title": c.title,
            "change_type": c.change_type,
            "status": c.status,
            "priority": c.priority,
            "owner": c.owner,
            "submitted_at": c.submitted_at,
            "eca_total": total,
            "eca_closed": closed,
            "eca_close_rate": round((closed / total) * 100) if total else 0,
        })

    return {
        "summary": {
            "change_total": db.query(models.Change).count(),
            "eca_total": eca_total,
            "eca_closed": eca_closed,
            "eca_pending": eca_pending,
            "eca_close_rate": round((eca_closed / eca_total) * 100) if eca_total else 0,
        },
        "by_status": [{"name": name, "value": value} for name, value in change_status_rows],
        "by_type": [{"name": name, "value": value} for name, value in change_type_rows],
        "by_priority": [{"name": name, "value": value} for name, value in change_priority_rows],
        "recent_changes": change_rows,
    }


@app.get("/api/reports/project-progress")
def report_project_progress(db: Session = Depends(get_db)) -> dict:
    """项目进度报表：阶段门分布、逾期任务、风险分布"""
    project_phase_rows = db.query(models.Project.phase, func.count(models.Project.id)).group_by(models.Project.phase).all()
    project_risk_rows = db.query(models.Project.risk_level, func.count(models.Project.id)).group_by(models.Project.risk_level).all()

    projects = db.query(models.Project).order_by(models.Project.id.desc()).all()
    # 逾期任务：due_date 非空且小于今天且未完成
    today_str = datetime.now().strftime("%Y-%m-%d")
    overdue_tasks = (
        db.query(models.ProjectTask)
        .filter(models.ProjectTask.due_date != "", models.ProjectTask.due_date < today_str, models.ProjectTask.status != "已完成")
        .order_by(models.ProjectTask.due_date)
        .all()
    )
    overdue_rows = []
    for t in overdue_tasks:
        proj = db.query(models.Project).filter(models.Project.id == t.project_id).first()
        overdue_rows.append({
            "id": t.id,
            "name": t.name,
            "phase": t.phase,
            "owner": t.owner,
            "due_date": t.due_date,
            "status": t.status,
            "project_no": proj.project_no if proj else "",
            "project_name": proj.name if proj else "",
        })

    # 项目维度汇总
    project_rows = []
    for p in projects:
        tasks = db.query(models.ProjectTask).filter(models.ProjectTask.project_id == p.id).all()
        task_total = len(tasks)
        task_done = sum(1 for t in tasks if t.status == "已完成")
        risks = db.query(models.ProjectRisk).filter(models.ProjectRisk.project_id == p.id).all()
        open_risks = sum(1 for r in risks if r.status != "已关闭")
        project_rows.append({
            "id": p.id,
            "project_no": p.project_no,
            "name": p.name,
            "phase": p.phase,
            "progress": p.progress,
            "owner": p.owner,
            "risk_level": p.risk_level,
            "task_total": task_total,
            "task_done": task_done,
            "task_done_rate": round((task_done / task_total) * 100) if task_total else 0,
            "open_risks": open_risks,
            "end_date": p.end_date,
        })

    # 风险类型分布
    risk_type_rows = db.query(models.ProjectRisk.risk_type, func.count(models.ProjectRisk.id)).group_by(models.ProjectRisk.risk_type).all()

    return {
        "summary": {
            "project_total": len(projects),
            "overdue_task_count": len(overdue_rows),
            "avg_progress": round(sum(p.progress for p in projects) / len(projects)) if projects else 0,
            "open_risk_count": db.query(models.ProjectRisk).filter(models.ProjectRisk.status != "已关闭").count(),
        },
        "by_phase": [{"name": name, "value": value} for name, value in project_phase_rows],
        "by_risk": [{"name": name, "value": value} for name, value in project_risk_rows],
        "by_risk_type": [{"name": name, "value": value} for name, value in risk_type_rows],
        "overdue_tasks": overdue_rows,
        "projects": project_rows,
    }


@app.get("/api/reports/quality-closure")
def report_quality_closure(db: Session = Depends(get_db)) -> dict:
    """质量闭环报表：CAPA 关闭率、问题严重度分布、状态趋势"""
    issue_total = db.query(models.QualityIssue).count()
    issue_closed = db.query(models.QualityIssue).filter(models.QualityIssue.status == "已关闭").count()
    issue_open = issue_total - issue_closed

    issue_severity_rows = db.query(models.QualityIssue.severity, func.count(models.QualityIssue.id)).group_by(models.QualityIssue.severity).all()
    issue_status_rows = db.query(models.QualityIssue.status, func.count(models.QualityIssue.id)).group_by(models.QualityIssue.status).all()

    capa_total = db.query(models.QualityCAPA).count()
    capa_closed = db.query(models.QualityCAPA).filter(models.QualityCAPA.status == "已关闭").count()
    capa_open = capa_total - capa_closed

    capa_source_rows = db.query(models.QualityCAPA.source, func.count(models.QualityCAPA.id)).group_by(models.QualityCAPA.source).all()

    # 质量问题清单
    issues = db.query(models.QualityIssue).order_by(models.QualityIssue.id.desc()).limit(20).all()
    issue_rows = []
    for i in issues:
        capas = db.query(models.QualityCAPA).filter(models.QualityCAPA.issue_id == i.id).all()
        issue_rows.append({
            "id": i.id,
            "issue_no": i.issue_no,
            "title": i.title,
            "product_model": i.product_model,
            "lot_no": i.lot_no,
            "severity": i.severity,
            "status": i.status,
            "owner": i.owner,
            "capa_count": len(capas),
            "capa_closed": sum(1 for c in capas if c.status == "已关闭"),
        })

    # 测试良率趋势（来自 QualityLot）
    quality_trend_rows = (
        db.query(models.QualityLot.tested_at, func.avg(models.QualityLot.cp_yield), func.avg(models.QualityLot.ft_yield))
        .group_by(models.QualityLot.tested_at)
        .order_by(models.QualityLot.tested_at)
        .limit(15)
        .all()
    )

    return {
        "summary": {
            "issue_total": issue_total,
            "issue_open": issue_open,
            "issue_close_rate": round((issue_closed / issue_total) * 100) if issue_total else 0,
            "capa_total": capa_total,
            "capa_open": capa_open,
            "capa_close_rate": round((capa_closed / capa_total) * 100) if capa_total else 0,
        },
        "issue_by_severity": [{"name": name, "value": value} for name, value in issue_severity_rows],
        "issue_by_status": [{"name": name, "value": value} for name, value in issue_status_rows],
        "capa_by_source": [{"name": name, "value": value} for name, value in capa_source_rows],
        "quality_trend": [{"date": d, "cp": round(cp, 1), "ft": round(ft, 1)} for d, cp, ft in quality_trend_rows],
        "issues": issue_rows,
    }


REPORT_SNAPSHOT_TYPES = {
    "completeness": ("数据完整度", report_completeness),
    "change": ("变更周期", report_change_cycle),
    "project": ("项目进度", report_project_progress),
    "quality": ("质量闭环", report_quality_closure),
}


def _report_snapshot_dict(row: models.ReportSnapshot) -> dict:
    import json

    try:
        summary = json.loads(row.summary_json or "{}")
    except Exception:
        summary = {}
    return {
        "id": row.id,
        "snapshot_no": row.snapshot_no,
        "report_type": row.report_type,
        "report_name": row.report_name,
        "summary": summary,
        "generated_by": row.generated_by,
        "generated_at": row.generated_at,
        "schedule_key": row.schedule_key,
    }


@app.get("/api/reports/snapshots")
def report_snapshots(page: int = 1, page_size: int = 20, report_type: str = "", db: Session = Depends(get_db)) -> dict:
    q = db.query(models.ReportSnapshot)
    if report_type:
        q = q.filter(models.ReportSnapshot.report_type == report_type)
    total = q.count()
    rows = q.order_by(models.ReportSnapshot.id.desc()).offset((page - 1) * page_size).limit(page_size).all()
    return {
        "items": [_report_snapshot_dict(row) for row in rows],
        "total": total,
        "page": page,
        "page_size": page_size,
    }


@app.post("/api/reports/snapshots", status_code=201)
def create_report_snapshot(payload: ReportSnapshotPayload, db: Session = Depends(get_db)) -> dict:
    import json
    from datetime import datetime

    config = REPORT_SNAPSHOT_TYPES.get(payload.report_type)
    if not config:
        raise HTTPException(status_code=400, detail="Unknown report type")
    report_name, report_fn = config
    data = report_fn(db)
    now = datetime.now()
    row = models.ReportSnapshot(
        snapshot_no=f"RPT-{payload.report_type.upper()}-{now.strftime('%Y%m%d%H%M%S%f')}",
        report_type=payload.report_type,
        report_name=report_name,
        summary_json=json.dumps(data.get("summary", {}), ensure_ascii=False),
        payload_json=json.dumps(data, ensure_ascii=False),
        generated_by=payload.generated_by,
        generated_at=now.strftime("%Y-%m-%d %H:%M"),
        schedule_key=payload.schedule_key or "manual",
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    return _report_snapshot_dict(row)


@app.get("/api/reports/completeness/export")
def export_report_completeness(db: Session = Depends(get_db)):
    from openpyxl import Workbook
    data = report_completeness(db)
    wb = Workbook()
    ws = wb.active
    ws.title = "数据完整度"
    s = data["summary"]
    ws.append(["指标", "值"])
    ws.append(["产品总数", s.get("product_total", 0)])
    ws.append(["文档覆盖率", f"{s.get('doc_coverage', 0)}%"])
    ws.append(["BOM覆盖率", f"{s.get('bom_coverage', 0)}%"])
    ws.append(["工艺覆盖率", f"{s.get('route_coverage', 0)}%"])
    ws.append(["需求覆盖率", f"{s.get('req_coverage', 0)}%"])
    ws.append(["文档签核率", f"{s.get('doc_signed_rate', 0)}%"])
    ws.append(["BOM发布率", f"{s.get('bom_released_rate', 0)}%"])
    ws.append(["工艺发布率", f"{s.get('route_released_rate', 0)}%"])
    ws.append(["资料齐套产品数", s.get("full_ready_count", 0)])
    ws.append([])
    ws.append(["型号", "名称", "生命周期", "负责人", "需求", "BOM", "工艺", "文档", "完整度%"])
    for p in data.get("products", []):
        ws.append([p.get("model"), p.get("name"), p.get("lifecycle"), p.get("owner"),
                    "有" if p.get("has_req") else "无", "有" if p.get("has_bom") else "无",
                    "有" if p.get("has_route") else "无", "有" if p.get("has_doc") else "无",
                    p.get("completeness", 0)])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(content=output.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": 'attachment; filename="report_completeness.xlsx"'})


@app.get("/api/reports/change-cycle/export")
def export_report_change_cycle(db: Session = Depends(get_db)):
    from openpyxl import Workbook
    data = report_change_cycle(db)
    wb = Workbook()
    ws = wb.active
    ws.title = "变更周期"
    s = data["summary"]
    ws.append(["指标", "值"])
    ws.append(["变更单总数", s.get("change_total", 0)])
    ws.append(["ECA总数", s.get("eca_total", 0)])
    ws.append(["ECA已完成", s.get("eca_closed", 0)])
    ws.append(["ECA待处理", s.get("eca_pending", 0)])
    ws.append(["ECA关闭率", f"{s.get('eca_close_rate', 0)}%"])
    ws.append([])
    ws.append(["状态分布"])
    ws.append(["状态", "数量"])
    for r in data.get("by_status", []):
        ws.append([r["name"], r["value"]])
    ws.append([])
    ws.append(["类型分布"])
    ws.append(["类型", "数量"])
    for r in data.get("by_type", []):
        ws.append([r["name"], r["value"]])
    ws.append([])
    ws.append(["优先级分布"])
    ws.append(["优先级", "数量"])
    for r in data.get("by_priority", []):
        ws.append([r["name"], r["value"]])
    ws.append([])
    ws.append(["近期变更ECA关闭率明细"])
    ws.append(["变更单", "标题", "类型", "优先级", "状态", "ECA已完成", "ECA总数", "关闭率%"])
    for c in data.get("recent_changes", []):
        ws.append([c.get("change_no"), c.get("title"), c.get("change_type"), c.get("priority"), c.get("status"), c.get("eca_closed"), c.get("eca_total"), c.get("eca_close_rate")])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(content=output.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": 'attachment; filename="report_change_cycle.xlsx"'})


@app.get("/api/reports/project-progress/export")
def export_report_project_progress(db: Session = Depends(get_db)):
    from openpyxl import Workbook
    data = report_project_progress(db)
    wb = Workbook()
    ws = wb.active
    ws.title = "项目进度"
    s = data["summary"]
    ws.append(["指标", "值"])
    ws.append(["项目总数", s.get("project_total", 0)])
    ws.append(["逾期任务数", s.get("overdue_task_count", 0)])
    ws.append(["平均进度", f"{s.get('avg_progress', 0)}%"])
    ws.append(["未关闭风险数", s.get("open_risk_count", 0)])
    ws.append([])
    ws.append(["阶段门分布"])
    ws.append(["阶段", "数量"])
    for r in data.get("by_phase", []):
        ws.append([r["name"], r["value"]])
    ws.append([])
    ws.append(["风险等级分布"])
    ws.append(["风险等级", "数量"])
    for r in data.get("by_risk", []):
        ws.append([r["name"], r["value"]])
    ws.append([])
    ws.append(["风险类型分布"])
    ws.append(["风险类型", "数量"])
    for r in data.get("by_risk_type", []):
        ws.append([r["name"], r["value"]])
    ws.append([])
    ws.append(["逾期任务明细"])
    ws.append(["项目编号", "项目名称", "任务", "阶段", "负责人", "截止日期", "状态"])
    for t in data.get("overdue_tasks", []):
        ws.append([t.get("project_no"), t.get("project_name"), t.get("name"), t.get("phase"), t.get("owner"), t.get("due_date"), t.get("status")])
    ws.append([])
    ws.append(["项目进度明细"])
    ws.append(["项目编号", "项目名称", "阶段", "进度%", "负责人", "任务完成率%", "未关闭风险"])
    for p in data.get("projects", []):
        ws.append([p.get("project_no"), p.get("name"), p.get("phase"), p.get("progress"), p.get("owner"), p.get("task_done_rate"), p.get("open_risks")])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(content=output.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": 'attachment; filename="report_project_progress.xlsx"'})


@app.get("/api/reports/quality-closure/export")
def export_report_quality_closure(db: Session = Depends(get_db)):
    from openpyxl import Workbook
    data = report_quality_closure(db)
    wb = Workbook()
    ws = wb.active
    ws.title = "质量闭环"
    s = data["summary"]
    ws.append(["指标", "值"])
    ws.append(["质量问题总数", s.get("issue_total", 0)])
    ws.append(["问题敞口数", s.get("issue_open", 0)])
    ws.append(["问题关闭率", f"{s.get('issue_close_rate', 0)}%"])
    ws.append(["CAPA总数", s.get("capa_total", 0)])
    ws.append(["CAPA敞口数", s.get("capa_open", 0)])
    ws.append(["CAPA关闭率", f"{s.get('capa_close_rate', 0)}%"])
    ws.append([])
    ws.append(["问题严重度分布"])
    ws.append(["严重度", "数量"])
    for r in data.get("issue_by_severity", []):
        ws.append([r["name"], r["value"]])
    ws.append([])
    ws.append(["问题状态分布"])
    ws.append(["状态", "数量"])
    for r in data.get("issue_by_status", []):
        ws.append([r["name"], r["value"]])
    ws.append([])
    ws.append(["CAPA来源分布"])
    ws.append(["来源", "数量"])
    for r in data.get("capa_by_source", []):
        ws.append([r["name"], r["value"]])
    ws.append([])
    ws.append(["良率趋势"])
    ws.append(["日期", "CP良率%", "FT良率%"])
    for t in data.get("quality_trend", []):
        ws.append([t.get("date"), t.get("cp"), t.get("ft")])
    ws.append([])
    ws.append(["质量问题清单"])
    ws.append(["问题编号", "标题", "型号", "Lot", "严重度", "状态", "负责人", "CAPA已关闭", "CAPA总数"])
    for i in data.get("issues", []):
        ws.append([i.get("issue_no"), i.get("title"), i.get("product_model"), i.get("lot_no"), i.get("severity"), i.get("status"), i.get("owner"), i.get("capa_closed"), i.get("capa_count")])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(content=output.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": 'attachment; filename="report_quality_closure.xlsx"'})


# ---- 通用附件管理 ----

@app.get("/api/attachments")
def list_attachments(object_type: str = "", object_id: int = 0, db: Session = Depends(get_db)) -> list[dict]:
    q = db.query(models.Attachment)
    if object_type:
        q = q.filter(models.Attachment.object_type == object_type)
    if object_id:
        q = q.filter(models.Attachment.object_id == object_id)
    rows = q.order_by(models.Attachment.id.desc()).all()
    return [_attachment_dict(r) for r in rows]


@app.post("/api/attachments/upload", status_code=201)
async def upload_attachment(object_type: str = Form(...), object_id: int = Form(...), description: str = Form(""), file: UploadFile = File(...), db: Session = Depends(get_db)):
    from datetime import datetime
    from uuid import uuid4

    os.makedirs(FILE_UPLOAD_DIR, exist_ok=True)
    original_name = os.path.basename(file.filename or "attachment")
    safe_name = f"{object_type}_{object_id}_{uuid4().hex}_{original_name}"
    file_path = os.path.join(FILE_UPLOAD_DIR, safe_name)
    content = await file.read()
    with open(file_path, "wb") as f:
        f.write(content)
    att = models.Attachment(
        object_type=object_type,
        object_id=object_id,
        file_name=original_name,
        file_path=file_path,
        file_size=len(content),
        file_type=file.content_type or "",
        description=description,
        uploaded_by="系统用户",
        uploaded_at=datetime.now().strftime("%Y-%m-%d %H:%M"),
    )
    db.add(att)
    db.commit()
    db.refresh(att)
    return _attachment_dict(att)


@app.get("/api/attachments/{attachment_id}/download")
def download_attachment(attachment_id: int, db: Session = Depends(get_db)):
    att = db.query(models.Attachment).filter(models.Attachment.id == attachment_id).first()
    if not att:
        raise HTTPException(status_code=404, detail="Attachment not found")
    if not os.path.exists(att.file_path):
        raise HTTPException(status_code=404, detail="File not found on disk")
    with open(att.file_path, "rb") as f:
        content = f.read()
    return Response(
        content=content,
        media_type="application/octet-stream",
        headers={"Content-Disposition": f'attachment; filename="{att.file_name}"'},
    )


@app.delete("/api/attachments/{attachment_id}")
def delete_attachment(attachment_id: int, db: Session = Depends(get_db)):
    att = db.query(models.Attachment).filter(models.Attachment.id == attachment_id).first()
    if not att:
        raise HTTPException(status_code=404, detail="Attachment not found")
    if os.path.exists(att.file_path):
        os.remove(att.file_path)
    db.delete(att)
    db.commit()
    return {"ok": True}


def _attachment_dict(r: models.Attachment) -> dict:
    return {
        "id": r.id,
        "object_type": r.object_type,
        "object_id": r.object_id,
        "file_name": r.file_name,
        "file_size": r.file_size,
        "file_type": r.file_type,
        "description": r.description,
        "uploaded_by": r.uploaded_by,
        "uploaded_at": r.uploaded_at,
    }


@app.get("/api/substitute-materials")
def substitute_materials(page: int = 1, page_size: int = 20, keyword: str = "", db: Session = Depends(get_db)) -> dict:
    q = db.query(models.SubstituteMaterial)
    if keyword:
        kw = f"%{keyword}%"
        q = q.filter(
            models.SubstituteMaterial.material_code.ilike(kw)
            | models.SubstituteMaterial.material_name.ilike(kw)
            | models.SubstituteMaterial.substitute_code.ilike(kw)
            | models.SubstituteMaterial.substitute_name.ilike(kw)
        )
    total = q.count()
    rows = q.order_by(models.SubstituteMaterial.material_code).offset((page - 1) * page_size).limit(page_size).all()
    return {"items": [{"id": r.id, "material_code": r.material_code, "material_name": r.material_name, "substitute_code": r.substitute_code, "substitute_name": r.substitute_name, "material_id": r.material_id, "substitute_material_id": r.substitute_material_id, "substitute_type": r.substitute_type, "strategy": r.strategy, "risk_level": r.risk_level, "status": r.status, "effective_date": r.effective_date, "expiry_date": r.expiry_date, "description": r.description} for r in rows], "total": total, "page": page, "page_size": page_size}


@app.post("/api/substitute-materials", status_code=201)
def create_substitute_material(payload: SubstituteMaterialPayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("material"))) -> dict:
    row = models.SubstituteMaterial(**payload.model_dump())
    db.add(row)
    db.commit()
    db.refresh(row)
    return {"id": row.id, "material_id": row.material_id, "substitute_material_id": row.substitute_material_id}
def update_substitute_material(row_id: int, payload: SubstituteMaterialUpdatePayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("material"))) -> dict:
    row = db.query(models.SubstituteMaterial).filter(models.SubstituteMaterial.id == row_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Substitute material not found")
    update_model(row, payload)
    db.commit()
    return {"ok": True}


@app.delete("/api/substitute-materials/{row_id}")
def delete_substitute_material(row_id: int, db: Session = Depends(get_db), _: dict = Depends(require_permission("material"))) -> dict:
    row = db.query(models.SubstituteMaterial).filter(models.SubstituteMaterial.id == row_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Substitute material not found")
    db.delete(row)
    db.commit()
    return {"ok": True}


@app.get("/api/suppliers")
def suppliers(page: int = 1, page_size: int = 20, keyword: str = "", db: Session = Depends(get_db)) -> dict:
    q = db.query(models.Supplier)
    if keyword:
        kw = f"%{keyword}%"
        q = q.filter(models.Supplier.code.ilike(kw) | models.Supplier.name.ilike(kw))
    total = q.count()
    rows = q.order_by(models.Supplier.code).offset((page - 1) * page_size).limit(page_size).all()
    return {"items": [{"id": r.id, "code": r.code, "name": r.name, "supplier_type": r.supplier_type, "contact": r.contact, "phone": r.phone, "email": r.email, "address": r.address, "certification": r.certification, "risk_level": r.risk_level, "status": r.status, "description": r.description} for r in rows], "total": total, "page": page, "page_size": page_size}


@app.post("/api/suppliers", status_code=201)
def create_supplier(payload: SupplierPayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("material"))) -> dict:
    row = models.Supplier(**payload.model_dump())
    db.add(row)
    commit_or_409(db, "Supplier code already exists")
    db.refresh(row)
    return {"id": row.id, "code": row.code, "name": row.name}


@app.put("/api/suppliers/{supplier_id}")
def update_supplier(supplier_id: int, payload: SupplierUpdatePayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("material"))) -> dict:
    row = db.query(models.Supplier).filter(models.Supplier.id == supplier_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Supplier not found")
    update_model(row, payload)
    commit_or_409(db, "Supplier code already exists")
    return {"ok": True}


@app.delete("/api/suppliers/{supplier_id}")
def delete_supplier(supplier_id: int, db: Session = Depends(get_db), _: dict = Depends(require_permission("material"))) -> dict:
    row = db.query(models.Supplier).filter(models.Supplier.id == supplier_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Supplier not found")
    db.delete(row)
    db.commit()
    return {"ok": True}


@app.get("/api/projects")
def projects(page: int = 1, page_size: int = 20, keyword: str = "", db: Session = Depends(get_db)) -> dict:
    q = db.query(models.Project).options(selectinload(models.Project.tasks), selectinload(models.Project.deliverables), selectinload(models.Project.risks))
    if keyword:
        kw = f"%{keyword}%"
        q = q.filter(models.Project.project_no.ilike(kw) | models.Project.name.ilike(kw) | models.Project.product_model.ilike(kw))
    total = q.count()
    rows = q.order_by(models.Project.id).offset((page - 1) * page_size).limit(page_size).all()
    return {
        "items": [
            {
                "id": row.id,
                "project_no": row.project_no,
                "name": row.name,
                "product_model": row.product_model,
                "phase": row.phase,
                "progress": row.progress,
                "owner": row.owner,
                "start_date": row.start_date,
                "end_date": row.end_date,
                "risk_level": row.risk_level,
                "tasks": [{"id": task.id, "name": task.name, "phase": task.phase, "owner": task.owner, "status": task.status, "due_date": task.due_date} for task in row.tasks],
                "deliverables": [{"id": d.id, "name": d.name, "deliverable_type": d.deliverable_type, "phase": d.phase, "owner": d.owner, "status": d.status, "due_date": d.due_date, "completed_at": d.completed_at, "description": d.description, "object_type": d.object_type, "object_id": d.object_id} for d in row.deliverables],
                "risks": [{"id": r.id, "risk_type": r.risk_type, "description": r.description, "impact": r.impact, "probability": r.probability, "severity": r.severity, "owner": r.owner, "status": r.status, "mitigation": r.mitigation} for r in row.risks],
            }
            for row in rows
        ],
        "total": total,
        "page": page,
        "page_size": page_size,
    }


@app.post("/api/projects", status_code=201)
def create_project(payload: ProjectPayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("project"))) -> dict:
    row = models.Project(**payload.model_dump())
    db.add(row)
    commit_or_409(db, "Project number already exists")
    db.refresh(row)
    return {"id": row.id, "project_no": row.project_no}


@app.put("/api/projects/{project_id}")
def update_project(project_id: int, payload: ProjectUpdatePayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("project"))) -> dict:
    row = db.query(models.Project).filter(models.Project.id == project_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Project not found")
    update_model(row, payload)
    db.commit()
    return {"ok": True}


@app.delete("/api/projects/{project_id}")
def delete_project(project_id: int, db: Session = Depends(get_db), _: dict = Depends(require_permission("project"))) -> dict:
    row = db.query(models.Project).filter(models.Project.id == project_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Project not found")
    db.delete(row)
    db.commit()
    return {"ok": True}


@app.get("/api/project-templates")
def project_templates(page: int = 1, page_size: int = 20, keyword: str = "", db: Session = Depends(get_db)) -> dict:
    q = db.query(models.ProjectTemplate)
    if keyword:
        kw = f"%{keyword}%"
        q = q.filter(models.ProjectTemplate.code.ilike(kw) | models.ProjectTemplate.name.ilike(kw))
    total = q.count()
    rows = q.order_by(models.ProjectTemplate.code).offset((page - 1) * page_size).limit(page_size).all()
    return {"items": [{"id": r.id, "code": r.code, "name": r.name, "description": r.description, "stages": r.stages, "status": r.status} for r in rows], "total": total, "page": page, "page_size": page_size}


@app.post("/api/project-templates", status_code=201)
def create_project_template(payload: ProjectTemplatePayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("project"))) -> dict:
    row = models.ProjectTemplate(**payload.model_dump())
    db.add(row)
    commit_or_409(db, "Template code already exists")
    db.refresh(row)
    return {"id": row.id, "code": row.code}


@app.put("/api/project-templates/{template_id}")
def update_project_template(template_id: int, payload: ProjectTemplateUpdatePayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("project"))) -> dict:
    row = db.query(models.ProjectTemplate).filter(models.ProjectTemplate.id == template_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Template not found")
    update_model(row, payload)
    db.commit()
    return {"ok": True}


@app.delete("/api/project-templates/{template_id}")
def delete_project_template(template_id: int, db: Session = Depends(get_db), _: dict = Depends(require_permission("project"))) -> dict:
    row = db.query(models.ProjectTemplate).filter(models.ProjectTemplate.id == template_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Template not found")
    db.delete(row)
    db.commit()
    return {"ok": True}


@app.post("/api/projects/from-template", status_code=201)
def create_project_from_template(payload: ProjectFromTemplatePayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("project"))) -> dict:
    import json
    template = db.query(models.ProjectTemplate).filter(models.ProjectTemplate.id == payload.template_id).first()
    if not template:
        raise HTTPException(status_code=404, detail="Template not found")
    try:
        stages = json.loads(template.stages) if template.stages else ["概念", "设计", "流片", "验证", "试产"]
    except (json.JSONDecodeError, TypeError):
        stages = ["概念", "设计", "流片", "验证", "试产"]
    first_phase = stages[0] if stages else "概念"
    project = models.Project(
        project_no=payload.project_no,
        name=payload.name,
        product_model=payload.product_model,
        phase=first_phase,
        progress=0,
        owner=payload.owner,
        start_date=payload.start_date,
        end_date=payload.end_date,
        risk_level="低",
    )
    db.add(project)
    commit_or_409(db, "Project number already exists")
    db.refresh(project)
    for stage in stages:
        task = models.ProjectTask(
            project_id=project.id,
            name=f"{stage}阶段任务",
            phase=stage,
            owner=payload.owner,
            status="待处理" if stage != first_phase else "进行中",
            due_date="",
            start_date=payload.start_date if stage == first_phase else "",
        )
        db.add(task)
    db.commit()
    return {"id": project.id, "project_no": project.project_no, "phase": project.phase, "tasks_created": len(stages)}


@app.get("/api/projects/{project_id}/deliverables")
def project_deliverables(project_id: int, db: Session = Depends(get_db)) -> list[dict]:
    ensure_project_exists(db, project_id)
    rows = db.query(models.ProjectDeliverable).filter(models.ProjectDeliverable.project_id == project_id).order_by(models.ProjectDeliverable.id).all()
    return [{"id": r.id, "name": r.name, "deliverable_type": r.deliverable_type, "phase": r.phase, "owner": r.owner, "status": r.status, "due_date": r.due_date, "completed_at": r.completed_at, "description": r.description, "object_type": r.object_type, "object_id": r.object_id} for r in rows]


@app.post("/api/projects/{project_id}/deliverables", status_code=201)
def create_project_deliverable(project_id: int, payload: ProjectDeliverablePayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("project"))) -> dict:
    ensure_project_exists(db, project_id)
    row = models.ProjectDeliverable(project_id=project_id, **payload.model_dump())
    db.add(row)
    db.commit()
    db.refresh(row)
    return {"id": row.id}


@app.put("/api/project-deliverables/{deliverable_id}")
def update_project_deliverable(deliverable_id: int, payload: ProjectDeliverableUpdatePayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("project"))) -> dict:
    row = db.query(models.ProjectDeliverable).filter(models.ProjectDeliverable.id == deliverable_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Deliverable not found")
    update_model(row, payload)
    db.commit()
    return {"ok": True}


@app.delete("/api/project-deliverables/{deliverable_id}")
def delete_project_deliverable(deliverable_id: int, db: Session = Depends(get_db), _: dict = Depends(require_permission("project"))) -> dict:
    row = db.query(models.ProjectDeliverable).filter(models.ProjectDeliverable.id == deliverable_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Deliverable not found")
    db.delete(row)
    db.commit()
    return {"ok": True}


@app.get("/api/projects/{project_id}/risks")
def project_risks(project_id: int, db: Session = Depends(get_db)) -> list[dict]:
    ensure_project_exists(db, project_id)
    rows = db.query(models.ProjectRisk).filter(models.ProjectRisk.project_id == project_id).order_by(models.ProjectRisk.id).all()
    return [{"id": r.id, "risk_type": r.risk_type, "description": r.description, "impact": r.impact, "probability": r.probability, "severity": r.severity, "owner": r.owner, "status": r.status, "mitigation": r.mitigation} for r in rows]


@app.post("/api/projects/{project_id}/risks", status_code=201)
def create_project_risk(project_id: int, payload: ProjectRiskPayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("project"))) -> dict:
    ensure_project_exists(db, project_id)
    row = models.ProjectRisk(project_id=project_id, **payload.model_dump())
    db.add(row)
    db.commit()
    db.refresh(row)
    return {"id": row.id}


@app.put("/api/project-risks/{risk_id}")
def update_project_risk(risk_id: int, payload: ProjectRiskUpdatePayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("project"))) -> dict:
    row = db.query(models.ProjectRisk).filter(models.ProjectRisk.id == risk_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Risk not found")
    update_model(row, payload)
    db.commit()
    return {"ok": True}


@app.delete("/api/project-risks/{risk_id}")
def delete_project_risk(risk_id: int, db: Session = Depends(get_db), _: dict = Depends(require_permission("project"))) -> dict:
    row = db.query(models.ProjectRisk).filter(models.ProjectRisk.id == risk_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Risk not found")
    db.delete(row)
    db.commit()
    return {"ok": True}


PROJECT_PHASES = ["概念", "设计", "流片", "验证", "试产", "量产导入"]


@app.get("/api/projects/{project_id}/tasks")
def project_tasks(project_id: int, db: Session = Depends(get_db)) -> list[dict]:
    ensure_project_exists(db, project_id)
    rows = db.query(models.ProjectTask).filter(models.ProjectTask.project_id == project_id).order_by(models.ProjectTask.id).all()
    return [{"id": r.id, "name": r.name, "phase": r.phase, "owner": r.owner, "status": r.status, "due_date": r.due_date, "start_date": r.start_date, "parent_id": r.parent_id, "depends_on": r.depends_on} for r in rows]


@app.post("/api/projects/{project_id}/tasks", status_code=201)
def create_project_task(project_id: int, payload: ProjectTaskPayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("project"))) -> dict:
    ensure_project_exists(db, project_id)
    row = models.ProjectTask(project_id=project_id, **payload.model_dump())
    db.add(row)
    db.commit()
    db.refresh(row)
    return {"id": row.id}


@app.put("/api/project-tasks/{task_id}")
def update_project_task(task_id: int, payload: ProjectTaskUpdatePayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("project"))) -> dict:
    row = db.query(models.ProjectTask).filter(models.ProjectTask.id == task_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Task not found")
    update_model(row, payload)
    db.commit()
    return {"ok": True}


@app.delete("/api/project-tasks/{task_id}")
def delete_project_task(task_id: int, db: Session = Depends(get_db), _: dict = Depends(require_permission("project"))) -> dict:
    row = db.query(models.ProjectTask).filter(models.ProjectTask.id == task_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Task not found")
    db.delete(row)
    db.commit()
    return {"ok": True}


@app.post("/api/projects/{project_id}/advance-phase")
def advance_project_phase(project_id: int, payload: ProjectPhaseGatePayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("project"))) -> dict:
    """推进项目阶段门：校验当前阶段交付物全部完成，推进到下一阶段并更新进度。"""
    project = db.query(models.Project).filter(models.Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    if project.phase == "量产导入":
        raise HTTPException(status_code=409, detail="Project already at final phase")

    current_phase = project.phase
    deliverables = (
        db.query(models.ProjectDeliverable)
        .filter(models.ProjectDeliverable.project_id == project_id, models.ProjectDeliverable.phase == current_phase)
        .all()
    )
    pending = [d for d in deliverables if d.status not in {"已完成", "已关闭"}]
    if pending:
        names = "、".join(d.name for d in pending[:3])
        raise HTTPException(status_code=409, detail=f"当前阶段有 {len(pending)} 个交付物未完成：{names}")

    try:
        idx = PROJECT_PHASES.index(current_phase)
    except ValueError:
        idx = 0
    next_phase = PROJECT_PHASES[min(idx + 1, len(PROJECT_PHASES) - 1)]
    old_phase = project.phase
    project.phase = next_phase
    project.progress = min(100, int((idx + 1) / len(PROJECT_PHASES) * 100))
    db.commit()
    db.refresh(project)
    return {
        "ok": True,
        "old_phase": old_phase,
        "new_phase": next_phase,
        "progress": project.progress,
        "message": f"阶段门从「{old_phase}」推进到「{next_phase}」",
    }


@app.get("/api/quality")
def quality(page: int = 1, page_size: int = 20, keyword: str = "", db: Session = Depends(get_db)) -> dict:
    lots_q = db.query(models.QualityLot).options(selectinload(models.QualityLot.product))
    issues_q = db.query(models.QualityIssue)
    if keyword:
        kw = f"%{keyword}%"
        lots_q = lots_q.filter(models.QualityLot.lot_no.ilike(kw) | models.QualityLot.wafer_id.ilike(kw))
        issues_q = issues_q.filter(models.QualityIssue.issue_no.ilike(kw) | models.QualityIssue.title.ilike(kw))
    lots_total = lots_q.count()
    issues_total = issues_q.count()
    lots = lots_q.order_by(models.QualityLot.id).offset((page - 1) * page_size).limit(page_size).all()
    issues = issues_q.order_by(models.QualityIssue.id).offset((page - 1) * page_size).limit(page_size).all()
    return {
        "lots": [
            {
                "id": lot.id,
                "product_model": lot.product.model,
                "lot_no": lot.lot_no,
                "wafer_id": lot.wafer_id,
                "stage": lot.stage,
                "cp_yield": lot.cp_yield,
                "ft_yield": lot.ft_yield,
                "bin1_rate": lot.bin1_rate,
                "issue_count": lot.issue_count,
                "status": lot.status,
                "tested_at": lot.tested_at,
            }
            for lot in lots
        ],
        "lots_total": lots_total,
        "issues": [
            {
                "id": issue.id,
                "issue_no": issue.issue_no,
                "product_model": issue.product_model,
                "lot_no": issue.lot_no,
                "title": issue.title,
                "severity": issue.severity,
                "status": issue.status,
                "owner": issue.owner,
                "root_cause": issue.root_cause,
                "corrective_action": issue.corrective_action,
            }
            for issue in issues
        ],
        "issues_total": issues_total,
        "page": page,
        "page_size": page_size,
    }


@app.get("/api/quality/capas")
def quality_capas(page: int = 1, page_size: int = 20, keyword: str = "", db: Session = Depends(get_db)) -> dict:
    q = db.query(models.QualityCAPA)
    if keyword:
        kw = f"%{keyword}%"
        q = q.filter(models.QualityCAPA.capa_no.ilike(kw) | models.QualityCAPA.title.ilike(kw))
    total = q.count()
    rows = q.order_by(models.QualityCAPA.id.desc()).offset((page - 1) * page_size).limit(page_size).all()
    return {
        "items": [
            {"id": r.id, "capa_no": r.capa_no, "issue_id": r.issue_id, "title": r.title, "source": r.source, "root_cause": r.root_cause, "corrective_action": r.corrective_action, "preventive_action": r.preventive_action, "owner": r.owner, "status": r.status, "due_date": r.due_date, "closed_at": r.closed_at, "result": r.result}
            for r in rows
        ],
        "total": total,
        "page": page,
        "page_size": page_size,
    }


@app.post("/api/quality/capas", status_code=201)
def create_quality_capa(payload: QualityCAPAPayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("quality"))) -> dict:
    if not payload.capa_no:
        count = db.query(models.QualityCAPA).count() + 1
        payload.capa_no = f"CAPA-{count:04d}"
    row = models.QualityCAPA(**payload.model_dump())
    db.add(row)
    db.commit()
    db.refresh(row)
    return {"id": row.id, "capa_no": row.capa_no}


@app.put("/api/quality/capas/{capa_id}")
def update_quality_capa(capa_id: int, payload: QualityCAPAUpdatePayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("quality"))) -> dict:
    row = db.query(models.QualityCAPA).filter(models.QualityCAPA.id == capa_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="CAPA not found")
    update_model(row, payload)
    db.commit()
    return {"ok": True}


@app.delete("/api/quality/capas/{capa_id}")
def delete_quality_capa(capa_id: int, db: Session = Depends(get_db), _: dict = Depends(require_permission("quality"))) -> dict:
    row = db.query(models.QualityCAPA).filter(models.QualityCAPA.id == capa_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="CAPA not found")
    db.delete(row)
    db.commit()
    return {"ok": True}


@app.post("/api/quality/issues/{issue_id}/create-capa", status_code=201)
def create_capa_from_issue(issue_id: int, db: Session = Depends(get_db), _: dict = Depends(require_permission("quality"))) -> dict:
    issue = db.query(models.QualityIssue).filter(models.QualityIssue.id == issue_id).first()
    if not issue:
        raise HTTPException(status_code=404, detail="Issue not found")
    count = db.query(models.QualityCAPA).count() + 1
    capa = models.QualityCAPA(
        capa_no=f"CAPA-{count:04d}",
        issue_id=issue.id,
        title=f"CAPA: {issue.title}",
        source="质量问题",
        root_cause=issue.root_cause or "",
        corrective_action=issue.corrective_action or "",
        owner=issue.owner,
        status="待处理",
    )
    db.add(capa)
    issue.status = "CAPA 执行中"
    db.commit()
    db.refresh(capa)
    return {"id": capa.id, "capa_no": capa.capa_no}


@app.post("/api/quality/issues", status_code=201)
def create_quality_issue(payload: QualityIssuePayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("quality"))) -> dict:
    if not payload.issue_no:
        count = db.query(models.QualityIssue).count() + 1
        payload.issue_no = f"QI-{count:04d}"
    row = models.QualityIssue(**payload.model_dump())
    db.add(row)
    commit_or_409(db, "Issue number already exists")
    db.refresh(row)
    return {"id": row.id, "issue_no": row.issue_no}


@app.put("/api/quality/issues/{issue_id}")
def update_quality_issue(issue_id: int, payload: QualityIssueUpdatePayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("quality"))) -> dict:
    row = db.query(models.QualityIssue).filter(models.QualityIssue.id == issue_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Quality issue not found")
    update_model(row, payload)
    db.commit()
    return {"ok": True}


@app.delete("/api/quality/issues/{issue_id}")
def delete_quality_issue(issue_id: int, db: Session = Depends(get_db), _: dict = Depends(require_permission("quality"))) -> dict:
    row = db.query(models.QualityIssue).filter(models.QualityIssue.id == issue_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Quality issue not found")
    if row.status in {"CAPA 执行中", "已关闭"}:
        raise HTTPException(status_code=409, detail="Cannot delete issue with CAPA in progress or closed")
    db.delete(row)
    db.commit()
    return {"ok": True}


@app.post("/api/quality/issues/{issue_id}/close")
def close_quality_issue(issue_id: int, payload: WorkflowTaskActionPayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("quality"))) -> dict:
    issue = db.query(models.QualityIssue).filter(models.QualityIssue.id == issue_id).first()
    if not issue:
        raise HTTPException(status_code=404, detail="Quality issue not found")
    if issue.status == "已关闭":
        raise HTTPException(status_code=409, detail="Issue already closed")
    issue.status = "已关闭"
    issue.corrective_action = f"{issue.corrective_action}\n[关闭人：{payload.acted_by}，日期：{today_text()}，备注：{payload.comment}]".strip()
    db.commit()
    return {"ok": True}


@app.post("/api/quality/issues/{issue_id}/trigger-ecr", status_code=201)
def trigger_ecr_from_issue(issue_id: int, db: Session = Depends(get_db), _: dict = Depends(require_permission("quality"))) -> dict:
    """质量问题触发 ECR：自动创建关联产品的变更单草稿。"""
    issue = db.query(models.QualityIssue).filter(models.QualityIssue.id == issue_id).first()
    if not issue:
        raise HTTPException(status_code=404, detail="Quality issue not found")
    if issue.status == "已关闭":
        raise HTTPException(status_code=409, detail="Cannot trigger ECR from closed issue")

    product = (
        db.query(models.Product)
        .filter(models.Product.model == issue.product_model)
        .first()
        if issue.product_model else None
    )
    if not product:
        raise HTTPException(status_code=409, detail=f"Product model '{issue.product_model}' not found, cannot trigger ECR")

    existing = (
        db.query(models.Change)
        .filter(models.Change.title.like(f"%{issue.issue_no}%"))
        .first()
    )
    if existing:
        raise HTTPException(status_code=409, detail=f"ECR already exists for {issue.issue_no}: {existing.change_no}")

    count = db.query(models.Change).count() + 1
    change_no = f"ECR-{product.model}-QI{count:03d}"
    change = models.Change(
        product_id=product.id,
        change_no=change_no,
        title=f"质量问题 {issue.issue_no} 触发变更：{issue.title}",
        change_type="ECR",
        reason=f"质量问题 {issue.issue_no} 触发。\n问题描述：{issue.title}\n根因：{issue.root_cause}\n纠正措施：{issue.corrective_action}",
        status="草稿",
        priority="高" if issue.severity == "高" else "中",
        owner=issue.owner or product.owner,
        submitted_at="",
        before_desc=f"Lot {issue.lot_no} 出现质量问题：{issue.title}",
        after_desc="待影响分析后确定",
    )
    db.add(change)
    issue.status = "已触发 ECR"
    db.commit()
    db.refresh(change)
    return {"id": change.id, "change_no": change.change_no, "product_id": product.id}


@app.post("/api/quality/capas/{capa_id}/close")
def close_quality_capa(capa_id: int, payload: WorkflowTaskActionPayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("quality"))) -> dict:
    """关闭 CAPA 并联动关闭关联的质量问题。"""
    capa = db.query(models.QualityCAPA).filter(models.QualityCAPA.id == capa_id).first()
    if not capa:
        raise HTTPException(status_code=404, detail="CAPA not found")
    if capa.status == "已关闭":
        raise HTTPException(status_code=409, detail="CAPA already closed")
    capa.status = "已关闭"
    capa.closed_at = today_text()
    capa.result = f"{capa.result}\n[关闭人：{payload.acted_by}，日期：{capa.closed_at}，备注：{payload.comment}]".strip()

    issue_closed = False
    if capa.issue_id:
        issue = db.query(models.QualityIssue).filter(models.QualityIssue.id == capa.issue_id).first()
        if issue and issue.status != "已关闭":
            issue.status = "已关闭"
            issue_closed = True
    db.commit()
    return {"ok": True, "issue_closed": issue_closed}


@app.get("/api/quality/reports")
def quality_reports(page: int = 1, page_size: int = 20, keyword: str = "", db: Session = Depends(get_db)) -> dict:
    q = db.query(models.QualityReport)
    if keyword:
        kw = f"%{keyword}%"
        q = q.filter(models.QualityReport.report_no.ilike(kw) | models.QualityReport.title.ilike(kw) | models.QualityReport.product_model.ilike(kw))
    total = q.count()
    rows = q.order_by(models.QualityReport.id.desc()).offset((page - 1) * page_size).limit(page_size).all()
    return {
        "items": [
            {"id": r.id, "report_no": r.report_no, "title": r.title, "report_type": r.report_type, "product_model": r.product_model, "issue_nos": r.issue_nos, "capa_nos": r.capa_nos, "summary": r.summary, "root_cause": r.root_cause, "corrective_action": r.corrective_action, "preventive_action": r.preventive_action, "owner": r.owner, "status": r.status, "archived_at": r.archived_at, "archived_by": r.archived_by}
            for r in rows
        ],
        "total": total,
        "page": page,
        "page_size": page_size,
    }


@app.post("/api/quality/reports", status_code=201)
def create_quality_report(payload: QualityReportPayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("quality"))) -> dict:
    if not payload.report_no:
        count = db.query(models.QualityReport).count() + 1
        payload.report_no = f"QR-{count:04d}"
    row = models.QualityReport(**payload.model_dump(), archived_at=today_text())
    db.add(row)
    commit_or_409(db, "Report number already exists")
    db.refresh(row)
    return {"id": row.id, "report_no": row.report_no}


@app.post("/api/quality/reports/archive-from-issues", status_code=201)
def archive_quality_report_from_issues(payload: QualityReportArchiveFromIssuePayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("quality"))) -> dict:
    """从已关闭的质量问题+CAPA 批量生成质量归档报告。"""
    if not payload.issue_ids:
        raise HTTPException(status_code=409, detail="No issues selected")
    issues = db.query(models.QualityIssue).filter(models.QualityIssue.id.in_(payload.issue_ids)).all()
    if not issues:
        raise HTTPException(status_code=404, detail="Issues not found")
    not_closed = [i for i in issues if i.status != "已关闭"]
    if not_closed:
        raise HTTPException(status_code=409, detail=f"Cannot archive: issues {[i.issue_no for i in not_closed]} are not closed")

    issue_ids = [i.id for i in issues]
    capas = db.query(models.QualityCAPA).filter(models.QualityCAPA.issue_id.in_(issue_ids)).all()
    issue_nos = "、".join(i.issue_no for i in issues)
    capa_nos = "、".join(c.capa_no for c in capas)
    product_models = list({i.product_model for i in issues if i.product_model})
    product_model = "、".join(product_models)
    root_causes = "\n".join(f"[{i.issue_no}] {i.root_cause}" for i in issues if i.root_cause)
    corrective = "\n".join(f"[{i.issue_no}] {i.corrective_action}" for i in issues if i.corrective_action)
    preventive = "\n".join(f"[{c.capa_no}] {c.preventive_action}" for c in capas if c.preventive_action)
    summary = f"本报告归档 {len(issues)} 个质量问题、{len(capas)} 个 CAPA。"

    count = db.query(models.QualityReport).count() + 1
    report_no = f"QR-{count:04d}"
    title = payload.title or f"质量归档报告 {report_no}"
    report = models.QualityReport(
        report_no=report_no,
        title=title,
        report_type="质量问题归档",
        product_model=product_model,
        issue_nos=issue_nos,
        capa_nos=capa_nos,
        summary=summary,
        root_cause=root_causes,
        corrective_action=corrective,
        preventive_action=preventive,
        owner=payload.owner or (issues[0].owner if issues else ""),
        status="已归档",
        archived_at=today_text(),
        archived_by=payload.owner or "系统",
    )
    db.add(report)
    db.commit()
    db.refresh(report)
    return {"id": report.id, "report_no": report.report_no}


@app.put("/api/quality/reports/{report_id}")
def update_quality_report(report_id: int, payload: QualityReportUpdatePayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("quality"))) -> dict:
    row = db.query(models.QualityReport).filter(models.QualityReport.id == report_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Report not found")
    update_model(row, payload)
    db.commit()
    return {"ok": True}


@app.delete("/api/quality/reports/{report_id}")
def delete_quality_report(report_id: int, db: Session = Depends(get_db), _: dict = Depends(require_permission("quality"))) -> dict:
    row = db.query(models.QualityReport).filter(models.QualityReport.id == report_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Report not found")
    db.delete(row)
    db.commit()
    return {"ok": True}


@app.get("/api/problem-reports")
def problem_reports(page: int = 1, page_size: int = 20, keyword: str = "", db: Session = Depends(get_db)) -> dict:
    q = db.query(models.ProblemReport)
    if keyword:
        kw = f"%{keyword}%"
        q = q.filter(models.ProblemReport.pr_no.ilike(kw) | models.ProblemReport.title.ilike(kw))
    total = q.count()
    rows = q.order_by(models.ProblemReport.id.desc()).offset((page - 1) * page_size).limit(page_size).all()
    return {
        "items": [
            {
                "id": r.id,
                "pr_no": r.pr_no,
                "title": r.title,
                "problem_type": r.problem_type,
                "severity": r.severity,
                "source": r.source,
                "product_id": r.product_id,
                "product_model": r.product_model,
                "description": r.description,
                "suggested_action": r.suggested_action,
                "status": r.status,
                "reporter": r.reporter,
                "reported_at": r.reported_at,
                "related_change_no": r.related_change_no,
                "remark": r.remark,
            }
            for r in rows
        ],
        "total": total,
        "page": page,
        "page_size": page_size,
    }


@app.post("/api/problem-reports", status_code=201)
def create_problem_report(payload: ProblemReportPayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("change"))) -> dict:
    count = db.query(models.ProblemReport).count() + 1
    pr_no = payload.pr_no or f"PR-{count:04d}"
    row = models.ProblemReport(**{**payload.model_dump(), "pr_no": pr_no})
    db.add(row)
    db.commit()
    db.refresh(row)
    return {"id": row.id, "pr_no": row.pr_no}


@app.put("/api/problem-reports/{report_id}")
def update_problem_report(report_id: int, payload: ProblemReportUpdatePayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("change"))) -> dict:
    row = db.query(models.ProblemReport).filter(models.ProblemReport.id == report_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Problem report not found")
    update_model(row, payload)
    db.commit()
    return {"ok": True}


@app.delete("/api/problem-reports/{report_id}")
def delete_problem_report(report_id: int, db: Session = Depends(get_db), _: dict = Depends(require_permission("change"))) -> dict:
    row = db.query(models.ProblemReport).filter(models.ProblemReport.id == report_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Problem report not found")
    db.delete(row)
    db.commit()
    return {"ok": True}


@app.get("/api/process-parameters")
def process_parameters(page: int = 1, page_size: int = 20, keyword: str = "", db: Session = Depends(get_db)) -> dict:
    q = db.query(models.ProcessParameter)
    if keyword:
        kw = f"%{keyword}%"
        q = q.filter(models.ProcessParameter.param_code.ilike(kw) | models.ProcessParameter.param_name.ilike(kw))
    total = q.count()
    rows = q.order_by(models.ProcessParameter.id.asc()).offset((page - 1) * page_size).limit(page_size).all()
    return {
        "items": [
            {
                "id": r.id,
                "param_code": r.param_code,
                "param_name": r.param_name,
                "param_type": r.param_type,
                "unit": r.unit,
                "category": r.category,
                "default_value": r.default_value,
                "min_value": r.min_value,
                "max_value": r.max_value,
                "description": r.description,
                "status": r.status,
            }
            for r in rows
        ],
        "total": total,
        "page": page,
        "page_size": page_size,
    }


@app.post("/api/process-parameters", status_code=201)
def create_process_parameter(payload: ProcessParameterPayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("process"))) -> dict:
    count = db.query(models.ProcessParameter).count() + 1
    param_code = payload.param_code or f"PARAM-{count:04d}"
    row = models.ProcessParameter(**{**payload.model_dump(), "param_code": param_code})
    db.add(row)
    db.commit()
    db.refresh(row)
    return {"id": row.id, "param_code": row.param_code}


@app.put("/api/process-parameters/{param_id}")
def update_process_parameter(param_id: int, payload: ProcessParameterUpdatePayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("process"))) -> dict:
    row = db.query(models.ProcessParameter).filter(models.ProcessParameter.id == param_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Process parameter not found")
    update_model(row, payload)
    db.commit()
    return {"ok": True}


@app.delete("/api/process-parameters/{param_id}")
def delete_process_parameter(param_id: int, db: Session = Depends(get_db), _: dict = Depends(require_permission("process"))) -> dict:
    row = db.query(models.ProcessParameter).filter(models.ProcessParameter.id == param_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Process parameter not found")
    db.delete(row)
    db.commit()
    return {"ok": True}

