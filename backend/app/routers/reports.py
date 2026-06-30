import io
import os
from datetime import date, datetime

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from fastapi.responses import Response
from sqlalchemy import func
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session, selectinload

from .. import models
from ..database import get_db
from ..deps import current_user_context, has_permission, require_permission
from ..schemas import *  # noqa: F401,F403
from ..serializers import *  # noqa: F401,F403
from ..services.change import (
    analyze_change_impacts,
    apply_change_action_revision,
    close_change_when_actions_done,
    create_change_release_jobs,
    get_eca_generated_object_gate,
    validate_action_effectivity,
    validate_change_action_target,
    validate_eca_generated_object_ready,
)
from ..services.helpers import (
    audit_log,
    commit_or_409,
    day_before,
    ensure_product_exists,
    ensure_project_exists,
    today_text,
    update_model,
)
from ..services.integration import create_integration_job
from ..services.process import (
    apply_bom_item_process_binding,
    ensure_route_editable,
    validate_process_route_ready,
)
from ..services.versioning import (
    close_previous_effective_boms,
    is_current_effective_bom,
    next_revision,
    next_unique_bom_version,
    next_unique_document_no,
    next_unique_process_version,
    next_unique_route_no,
)
from ..services.workflow import start_workflow


router = APIRouter()


@router.get("/api/reports/completeness")
def report_completeness(db: Session = Depends(get_db)) -> dict:
    """数据完整度报表：产品资料/文档/BOM/工艺齐套率"""
    products = db.query(models.Product).all()
    total_products = len(products)
    product_rows = []
    doc_ready = 0
    bom_ready = 0
    route_ready = 0
    req_ready = 0
    for p in products:
        has_doc = db.query(models.Document).filter(models.Document.product_id == p.id).count() > 0
        has_bom = db.query(models.BomHeader).filter(models.BomHeader.product_id == p.id).count() > 0
        has_route = db.query(models.ProcessRoute).filter(models.ProcessRoute.product_id == p.id).count() > 0
        has_req = db.query(models.Requirement).filter(models.Requirement.product_id == p.id).count() > 0
        if has_doc:
            doc_ready += 1
        if has_bom:
            bom_ready += 1
        if has_route:
            route_ready += 1
        if has_req:
            req_ready += 1
        # 单产品完整度：四项全有视为齐套
        completeness = sum([has_doc, has_bom, has_route, has_req])
        product_rows.append({
            "id": p.id,
            "model": p.model,
            "name": p.name,
            "lifecycle": p.lifecycle,
            "owner": p.owner,
            "has_doc": has_doc,
            "has_bom": has_bom,
            "has_route": has_route,
            "has_req": has_req,
            "completeness": round(completeness * 25),  # 0/25/50/75/100
        })

    # 文档签核率
    doc_total = db.query(models.Document).count()
    doc_signed = db.query(models.Document).filter(models.Document.approval_status == "已签核").count()
    # BOM 已发布率
    bom_total = db.query(models.BomHeader).count()
    bom_released = db.query(models.BomHeader).filter(models.BomHeader.status == "已发布").count()
    # 工艺已发布率
    route_total = db.query(models.ProcessRoute).count()
    route_released = db.query(models.ProcessRoute).filter(models.ProcessRoute.status == "已发布").count()

    return {
        "summary": {
            "product_total": total_products,
            "doc_coverage": round((doc_ready / total_products) * 100) if total_products else 0,
            "bom_coverage": round((bom_ready / total_products) * 100) if total_products else 0,
            "route_coverage": round((route_ready / total_products) * 100) if total_products else 0,
            "req_coverage": round((req_ready / total_products) * 100) if total_products else 0,
            "doc_signed_rate": round((doc_signed / doc_total) * 100) if doc_total else 0,
            "bom_released_rate": round((bom_released / bom_total) * 100) if bom_total else 0,
            "route_released_rate": round((route_released / route_total) * 100) if route_total else 0,
            "full_ready_count": sum(1 for r in product_rows if r["completeness"] == 100),
        },
        "products": product_rows,
    }


@router.get("/api/reports/change-cycle")
def report_change_cycle(db: Session = Depends(get_db)) -> dict:
    """变更周期报表：ECR/ECO/ECN 状态分布、ECA 关闭率"""
    change_status_rows = db.query(models.Change.status, func.count(models.Change.id)).group_by(models.Change.status).all()
    change_type_rows = db.query(models.Change.change_type, func.count(models.Change.id)).group_by(models.Change.change_type).all()
    change_priority_rows = db.query(models.Change.priority, func.count(models.Change.id)).group_by(models.Change.priority).all()

    # ECA 关闭率
    eca_total = db.query(models.ChangeAction).count()
    eca_closed = db.query(models.ChangeAction).filter(models.ChangeAction.status == "已完成").count()
    eca_pending = db.query(models.ChangeAction).filter(models.ChangeAction.status != "已完成").count()

    # 按变更单聚合 ECA 完成情况
    changes = db.query(models.Change).order_by(models.Change.id.desc()).limit(20).all()
    change_rows = []
    for c in changes:
        actions = db.query(models.ChangeAction).filter(models.ChangeAction.change_id == c.id).all()
        total = len(actions)
        closed = sum(1 for a in actions if a.status == "已完成")
        change_rows.append({
            "id": c.id,
            "change_no": c.change_no,
            "title": c.title,
            "change_type": c.change_type,
            "status": c.status,
            "priority": c.priority,
            "owner": c.owner,
            "submitted_at": c.submitted_at,
            "eca_total": total,
            "eca_closed": closed,
            "eca_close_rate": round((closed / total) * 100) if total else 0,
        })

    return {
        "summary": {
            "change_total": db.query(models.Change).count(),
            "eca_total": eca_total,
            "eca_closed": eca_closed,
            "eca_pending": eca_pending,
            "eca_close_rate": round((eca_closed / eca_total) * 100) if eca_total else 0,
        },
        "by_status": [{"name": name, "value": value} for name, value in change_status_rows],
        "by_type": [{"name": name, "value": value} for name, value in change_type_rows],
        "by_priority": [{"name": name, "value": value} for name, value in change_priority_rows],
        "recent_changes": change_rows,
    }


@router.get("/api/reports/project-progress")
def report_project_progress(db: Session = Depends(get_db)) -> dict:
    """项目进度报表：阶段门分布、逾期任务、风险分布"""
    project_phase_rows = db.query(models.Project.phase, func.count(models.Project.id)).group_by(models.Project.phase).all()
    project_risk_rows = db.query(models.Project.risk_level, func.count(models.Project.id)).group_by(models.Project.risk_level).all()

    projects = db.query(models.Project).order_by(models.Project.id.desc()).all()
    # 逾期任务：due_date 非空且小于今天且未完成
    today_str = datetime.now().strftime("%Y-%m-%d")
    overdue_tasks = (
        db.query(models.ProjectTask)
        .filter(models.ProjectTask.due_date != "", models.ProjectTask.due_date < today_str, models.ProjectTask.status != "已完成")
        .order_by(models.ProjectTask.due_date)
        .all()
    )
    overdue_rows = []
    for t in overdue_tasks:
        proj = db.query(models.Project).filter(models.Project.id == t.project_id).first()
        overdue_rows.append({
            "id": t.id,
            "name": t.name,
            "phase": t.phase,
            "owner": t.owner,
            "due_date": t.due_date,
            "status": t.status,
            "project_no": proj.project_no if proj else "",
            "project_name": proj.name if proj else "",
        })

    # 项目维度汇总
    project_rows = []
    for p in projects:
        tasks = db.query(models.ProjectTask).filter(models.ProjectTask.project_id == p.id).all()
        task_total = len(tasks)
        task_done = sum(1 for t in tasks if t.status == "已完成")
        risks = db.query(models.ProjectRisk).filter(models.ProjectRisk.project_id == p.id).all()
        open_risks = sum(1 for r in risks if r.status != "已关闭")
        project_rows.append({
            "id": p.id,
            "project_no": p.project_no,
            "name": p.name,
            "phase": p.phase,
            "progress": p.progress,
            "owner": p.owner,
            "risk_level": p.risk_level,
            "task_total": task_total,
            "task_done": task_done,
            "task_done_rate": round((task_done / task_total) * 100) if task_total else 0,
            "open_risks": open_risks,
            "end_date": p.end_date,
        })

    # 风险类型分布
    risk_type_rows = db.query(models.ProjectRisk.risk_type, func.count(models.ProjectRisk.id)).group_by(models.ProjectRisk.risk_type).all()

    return {
        "summary": {
            "project_total": len(projects),
            "overdue_task_count": len(overdue_rows),
            "avg_progress": round(sum(p.progress for p in projects) / len(projects)) if projects else 0,
            "open_risk_count": db.query(models.ProjectRisk).filter(models.ProjectRisk.status != "已关闭").count(),
        },
        "by_phase": [{"name": name, "value": value} for name, value in project_phase_rows],
        "by_risk": [{"name": name, "value": value} for name, value in project_risk_rows],
        "by_risk_type": [{"name": name, "value": value} for name, value in risk_type_rows],
        "overdue_tasks": overdue_rows,
        "projects": project_rows,
    }


@router.get("/api/reports/quality-closure")
def report_quality_closure(db: Session = Depends(get_db)) -> dict:
    """质量闭环报表：CAPA 关闭率、问题严重度分布、状态趋势"""
    issue_total = db.query(models.QualityIssue).count()
    issue_closed = db.query(models.QualityIssue).filter(models.QualityIssue.status == "已关闭").count()
    issue_open = issue_total - issue_closed

    issue_severity_rows = db.query(models.QualityIssue.severity, func.count(models.QualityIssue.id)).group_by(models.QualityIssue.severity).all()
    issue_status_rows = db.query(models.QualityIssue.status, func.count(models.QualityIssue.id)).group_by(models.QualityIssue.status).all()

    capa_total = db.query(models.QualityCAPA).count()
    capa_closed = db.query(models.QualityCAPA).filter(models.QualityCAPA.status == "已关闭").count()
    capa_open = capa_total - capa_closed

    capa_source_rows = db.query(models.QualityCAPA.source, func.count(models.QualityCAPA.id)).group_by(models.QualityCAPA.source).all()

    # 质量问题清单
    issues = db.query(models.QualityIssue).order_by(models.QualityIssue.id.desc()).limit(20).all()
    issue_rows = []
    for i in issues:
        capas = db.query(models.QualityCAPA).filter(models.QualityCAPA.issue_id == i.id).all()
        issue_rows.append({
            "id": i.id,
            "issue_no": i.issue_no,
            "title": i.title,
            "product_model": i.product_model,
            "lot_no": i.lot_no,
            "severity": i.severity,
            "status": i.status,
            "owner": i.owner,
            "capa_count": len(capas),
            "capa_closed": sum(1 for c in capas if c.status == "已关闭"),
        })

    # 测试良率趋势（来自 QualityLot）
    quality_trend_rows = (
        db.query(models.QualityLot.tested_at, func.avg(models.QualityLot.cp_yield), func.avg(models.QualityLot.ft_yield))
        .group_by(models.QualityLot.tested_at)
        .order_by(models.QualityLot.tested_at)
        .limit(15)
        .all()
    )

    return {
        "summary": {
            "issue_total": issue_total,
            "issue_open": issue_open,
            "issue_close_rate": round((issue_closed / issue_total) * 100) if issue_total else 0,
            "capa_total": capa_total,
            "capa_open": capa_open,
            "capa_close_rate": round((capa_closed / capa_total) * 100) if capa_total else 0,
        },
        "issue_by_severity": [{"name": name, "value": value} for name, value in issue_severity_rows],
        "issue_by_status": [{"name": name, "value": value} for name, value in issue_status_rows],
        "capa_by_source": [{"name": name, "value": value} for name, value in capa_source_rows],
        "quality_trend": [{"date": d, "cp": round(cp, 1), "ft": round(ft, 1)} for d, cp, ft in quality_trend_rows],
        "issues": issue_rows,
    }


REPORT_SNAPSHOT_TYPES = {
    "completeness": ("数据完整度", report_completeness),
    "change": ("变更周期", report_change_cycle),
    "project": ("项目进度", report_project_progress),
    "quality": ("质量闭环", report_quality_closure),
}


def _report_snapshot_dict(row: models.ReportSnapshot) -> dict:
    import json

    try:
        summary = json.loads(row.summary_json or "{}")
    except Exception:
        summary = {}
    return {
        "id": row.id,
        "snapshot_no": row.snapshot_no,
        "report_type": row.report_type,
        "report_name": row.report_name,
        "summary": summary,
        "generated_by": row.generated_by,
        "generated_at": row.generated_at,
        "schedule_key": row.schedule_key,
    }


@router.get("/api/reports/snapshots")
def report_snapshots(page: int = 1, page_size: int = 20, report_type: str = "", db: Session = Depends(get_db)) -> dict:
    q = db.query(models.ReportSnapshot)
    if report_type:
        q = q.filter(models.ReportSnapshot.report_type == report_type)
    total = q.count()
    rows = q.order_by(models.ReportSnapshot.id.desc()).offset((page - 1) * page_size).limit(page_size).all()
    return {
        "items": [_report_snapshot_dict(row) for row in rows],
        "total": total,
        "page": page,
        "page_size": page_size,
    }


@router.post("/api/reports/snapshots", status_code=201)
def create_report_snapshot(payload: ReportSnapshotPayload, db: Session = Depends(get_db)) -> dict:
    import json
    from datetime import datetime

    config = REPORT_SNAPSHOT_TYPES.get(payload.report_type)
    if not config:
        raise HTTPException(status_code=400, detail="Unknown report type")
    report_name, report_fn = config
    data = report_fn(db)
    now = datetime.now()
    row = models.ReportSnapshot(
        snapshot_no=f"RPT-{payload.report_type.upper()}-{now.strftime('%Y%m%d%H%M%S%f')}",
        report_type=payload.report_type,
        report_name=report_name,
        summary_json=json.dumps(data.get("summary", {}), ensure_ascii=False),
        payload_json=json.dumps(data, ensure_ascii=False),
        generated_by=payload.generated_by,
        generated_at=now.strftime("%Y-%m-%d %H:%M"),
        schedule_key=payload.schedule_key or "manual",
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    return _report_snapshot_dict(row)


@router.get("/api/reports/completeness/export")
def export_report_completeness(db: Session = Depends(get_db)):
    from openpyxl import Workbook
    data = report_completeness(db)
    wb = Workbook()
    ws = wb.active
    ws.title = "数据完整度"
    s = data["summary"]
    ws.append(["指标", "值"])
    ws.append(["产品总数", s.get("product_total", 0)])
    ws.append(["文档覆盖率", f"{s.get('doc_coverage', 0)}%"])
    ws.append(["BOM覆盖率", f"{s.get('bom_coverage', 0)}%"])
    ws.append(["工艺覆盖率", f"{s.get('route_coverage', 0)}%"])
    ws.append(["需求覆盖率", f"{s.get('req_coverage', 0)}%"])
    ws.append(["文档签核率", f"{s.get('doc_signed_rate', 0)}%"])
    ws.append(["BOM发布率", f"{s.get('bom_released_rate', 0)}%"])
    ws.append(["工艺发布率", f"{s.get('route_released_rate', 0)}%"])
    ws.append(["资料齐套产品数", s.get("full_ready_count", 0)])
    ws.append([])
    ws.append(["型号", "名称", "生命周期", "负责人", "需求", "BOM", "工艺", "文档", "完整度%"])
    for p in data.get("products", []):
        ws.append([p.get("model"), p.get("name"), p.get("lifecycle"), p.get("owner"),
                    "有" if p.get("has_req") else "无", "有" if p.get("has_bom") else "无",
                    "有" if p.get("has_route") else "无", "有" if p.get("has_doc") else "无",
                    p.get("completeness", 0)])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(content=output.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": 'attachment; filename="report_completeness.xlsx"'})


@router.get("/api/reports/change-cycle/export")
def export_report_change_cycle(db: Session = Depends(get_db)):
    from openpyxl import Workbook
    data = report_change_cycle(db)
    wb = Workbook()
    ws = wb.active
    ws.title = "变更周期"
    s = data["summary"]
    ws.append(["指标", "值"])
    ws.append(["变更单总数", s.get("change_total", 0)])
    ws.append(["ECA总数", s.get("eca_total", 0)])
    ws.append(["ECA已完成", s.get("eca_closed", 0)])
    ws.append(["ECA待处理", s.get("eca_pending", 0)])
    ws.append(["ECA关闭率", f"{s.get('eca_close_rate', 0)}%"])
    ws.append([])
    ws.append(["状态分布"])
    ws.append(["状态", "数量"])
    for r in data.get("by_status", []):
        ws.append([r["name"], r["value"]])
    ws.append([])
    ws.append(["类型分布"])
    ws.append(["类型", "数量"])
    for r in data.get("by_type", []):
        ws.append([r["name"], r["value"]])
    ws.append([])
    ws.append(["优先级分布"])
    ws.append(["优先级", "数量"])
    for r in data.get("by_priority", []):
        ws.append([r["name"], r["value"]])
    ws.append([])
    ws.append(["近期变更ECA关闭率明细"])
    ws.append(["变更单", "标题", "类型", "优先级", "状态", "ECA已完成", "ECA总数", "关闭率%"])
    for c in data.get("recent_changes", []):
        ws.append([c.get("change_no"), c.get("title"), c.get("change_type"), c.get("priority"), c.get("status"), c.get("eca_closed"), c.get("eca_total"), c.get("eca_close_rate")])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(content=output.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": 'attachment; filename="report_change_cycle.xlsx"'})


@router.get("/api/reports/project-progress/export")
def export_report_project_progress(db: Session = Depends(get_db)):
    from openpyxl import Workbook
    data = report_project_progress(db)
    wb = Workbook()
    ws = wb.active
    ws.title = "项目进度"
    s = data["summary"]
    ws.append(["指标", "值"])
    ws.append(["项目总数", s.get("project_total", 0)])
    ws.append(["逾期任务数", s.get("overdue_task_count", 0)])
    ws.append(["平均进度", f"{s.get('avg_progress', 0)}%"])
    ws.append(["未关闭风险数", s.get("open_risk_count", 0)])
    ws.append([])
    ws.append(["阶段门分布"])
    ws.append(["阶段", "数量"])
    for r in data.get("by_phase", []):
        ws.append([r["name"], r["value"]])
    ws.append([])
    ws.append(["风险等级分布"])
    ws.append(["风险等级", "数量"])
    for r in data.get("by_risk", []):
        ws.append([r["name"], r["value"]])
    ws.append([])
    ws.append(["风险类型分布"])
    ws.append(["风险类型", "数量"])
    for r in data.get("by_risk_type", []):
        ws.append([r["name"], r["value"]])
    ws.append([])
    ws.append(["逾期任务明细"])
    ws.append(["项目编号", "项目名称", "任务", "阶段", "负责人", "截止日期", "状态"])
    for t in data.get("overdue_tasks", []):
        ws.append([t.get("project_no"), t.get("project_name"), t.get("name"), t.get("phase"), t.get("owner"), t.get("due_date"), t.get("status")])
    ws.append([])
    ws.append(["项目进度明细"])
    ws.append(["项目编号", "项目名称", "阶段", "进度%", "负责人", "任务完成率%", "未关闭风险"])
    for p in data.get("projects", []):
        ws.append([p.get("project_no"), p.get("name"), p.get("phase"), p.get("progress"), p.get("owner"), p.get("task_done_rate"), p.get("open_risks")])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(content=output.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": 'attachment; filename="report_project_progress.xlsx"'})


@router.get("/api/reports/quality-closure/export")
def export_report_quality_closure(db: Session = Depends(get_db)):
    from openpyxl import Workbook
    data = report_quality_closure(db)
    wb = Workbook()
    ws = wb.active
    ws.title = "质量闭环"
    s = data["summary"]
    ws.append(["指标", "值"])
    ws.append(["质量问题总数", s.get("issue_total", 0)])
    ws.append(["问题敞口数", s.get("issue_open", 0)])
    ws.append(["问题关闭率", f"{s.get('issue_close_rate', 0)}%"])
    ws.append(["CAPA总数", s.get("capa_total", 0)])
    ws.append(["CAPA敞口数", s.get("capa_open", 0)])
    ws.append(["CAPA关闭率", f"{s.get('capa_close_rate', 0)}%"])
    ws.append([])
    ws.append(["问题严重度分布"])
    ws.append(["严重度", "数量"])
    for r in data.get("issue_by_severity", []):
        ws.append([r["name"], r["value"]])
    ws.append([])
    ws.append(["问题状态分布"])
    ws.append(["状态", "数量"])
    for r in data.get("issue_by_status", []):
        ws.append([r["name"], r["value"]])
    ws.append([])
    ws.append(["CAPA来源分布"])
    ws.append(["来源", "数量"])
    for r in data.get("capa_by_source", []):
        ws.append([r["name"], r["value"]])
    ws.append([])
    ws.append(["良率趋势"])
    ws.append(["日期", "CP良率%", "FT良率%"])
    for t in data.get("quality_trend", []):
        ws.append([t.get("date"), t.get("cp"), t.get("ft")])
    ws.append([])
    ws.append(["质量问题清单"])
    ws.append(["问题编号", "标题", "型号", "Lot", "严重度", "状态", "负责人", "CAPA已关闭", "CAPA总数"])
    for i in data.get("issues", []):
        ws.append([i.get("issue_no"), i.get("title"), i.get("product_model"), i.get("lot_no"), i.get("severity"), i.get("status"), i.get("owner"), i.get("capa_closed"), i.get("capa_count")])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(content=output.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": 'attachment; filename="report_quality_closure.xlsx"'})


# ---- 通用附件管理 ----

ATTACHMENT_PERMISSION_BY_OBJECT = {
    "BOM": "bom",
    "Change": "change",
    "Project": "project",
    "Document": "document",
    "Requirement": "requirement",
    "QualityIssue": "quality",
}


def _require_attachment_permission(object_type: str, context: dict) -> None:
    permission = ATTACHMENT_PERMISSION_BY_OBJECT.get(object_type)
    if not permission:
        raise HTTPException(status_code=400, detail="Unsupported attachment object type")
    if not has_permission(context, permission):
        raise HTTPException(status_code=403, detail="Permission denied")
