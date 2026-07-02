import io

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import Response
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session, selectinload

from .. import models
from ..database import get_db
from ..deps import require_permission
from ..schemas import (
    BomBatchDeletePayload,
    BomBatchQuantityPayload,
    BomBatchReplacePayload,
    BomHeaderPayload,
    BomHeaderUpdatePayload,
    BomItemPayload,
    BomItemUpdatePayload,
    BomTransformPayload,
)
from ..serializers import (
    bom_item_compare_key,
    serialize_bom,
    serialize_bom_compare_item,
    serialize_bom_item,
)
from ..services.helpers import update_model
from ..services.process import apply_bom_item_process_binding

router = APIRouter()

LOCKED_BOM_STATES = {"Frozen", "Active", "Invalid"}
BOM_EXPORT_COLUMNS = ["序号", "物料类型", "物料名称", "物料版本", "需求量", "单位", "工序名称", "工序版本"]


def _get_bom(db: Session, bom_id: int) -> models.BomHeader:
    bom = (
        db.query(models.BomHeader)
        .options(selectinload(models.BomHeader.items))
        .filter(models.BomHeader.id == bom_id)
        .first()
    )
    if not bom:
        raise HTTPException(status_code=404, detail="BOM not found")
    return bom


def _ensure_bom_editable(bom: models.BomHeader) -> None:
    if bom.bom_state in LOCKED_BOM_STATES:
        raise HTTPException(status_code=409, detail=f"BOM in {bom.bom_state} state cannot be edited")


def _ensure_unique_bom(db: Session, bom_name: str, bom_version: str, exclude_id: int | None = None) -> None:
    q = db.query(models.BomHeader.id).filter(
        models.BomHeader.bom_name == bom_name,
        models.BomHeader.bom_version == bom_version,
    )
    if exclude_id is not None:
        q = q.filter(models.BomHeader.id != exclude_id)
    if q.first():
        raise HTTPException(status_code=409, detail="BOM name and version already exists")


def _refresh_bom(db: Session, bom_id: int) -> models.BomHeader:
    return (
        db.query(models.BomHeader)
        .options(selectinload(models.BomHeader.items))
        .filter(models.BomHeader.id == bom_id)
        .first()
    )


@router.get("/api/boms")
def boms(page: int = 1, page_size: int = 20, keyword: str = "", db: Session = Depends(get_db)) -> dict:
    q = db.query(models.BomHeader).options(selectinload(models.BomHeader.items))
    if keyword:
        kw = f"%{keyword}%"
        q = q.filter(
            models.BomHeader.bom_name.ilike(kw)
            | models.BomHeader.bom_version.ilike(kw)
            | models.BomHeader.description.ilike(kw)
            | models.BomHeader.owner.ilike(kw)
        )
    total = q.count()
    rows = q.order_by(models.BomHeader.id.desc()).offset((page - 1) * page_size).limit(page_size).all()
    return {"items": [serialize_bom(row) for row in rows], "total": total, "page": page, "page_size": page_size}


@router.post("/api/boms", status_code=201)
def create_bom(payload: BomHeaderPayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("bom"))) -> dict:
    _ensure_unique_bom(db, payload.bom_name, payload.bom_version)
    bom = models.BomHeader(**payload.model_dump())
    db.add(bom)
    db.commit()
    db.refresh(bom)
    return serialize_bom(_refresh_bom(db, bom.id))


@router.put("/api/boms/{bom_id}")
def update_bom(bom_id: int, payload: BomHeaderUpdatePayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("bom"))) -> dict:
    bom = _get_bom(db, bom_id)
    _ensure_bom_editable(bom)
    next_name = payload.bom_name if payload.bom_name is not None else bom.bom_name
    next_version = payload.bom_version if payload.bom_version is not None else bom.bom_version
    _ensure_unique_bom(db, next_name, next_version, exclude_id=bom.id)
    update_model(bom, payload)
    for item in bom.items:
        item.bom_name = bom.bom_name
        item.bom_version = bom.bom_version
    db.commit()
    return serialize_bom(_refresh_bom(db, bom.id))


@router.delete("/api/boms/{bom_id}")
def delete_bom(bom_id: int, db: Session = Depends(get_db), _: dict = Depends(require_permission("bom"))) -> dict:
    bom = _get_bom(db, bom_id)
    _ensure_bom_editable(bom)
    db.delete(bom)
    db.commit()
    return {"deleted": True}


@router.post("/api/boms/{bom_id}/items", status_code=201)
def create_bom_item(bom_id: int, payload: BomItemPayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("bom"))) -> dict:
    bom = _get_bom(db, bom_id)
    _ensure_bom_editable(bom)
    data = apply_bom_item_process_binding(db, payload, product_id=0)
    data["bom_name"] = bom.bom_name
    data["bom_version"] = bom.bom_version
    item = models.BomItem(bom_id=bom.id, **data)
    db.add(item)
    db.commit()
    db.refresh(item)
    return serialize_bom_item(item)


@router.put("/api/bom-items/{item_id}")
def update_bom_item(item_id: int, payload: BomItemUpdatePayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("bom"))) -> dict:
    item = db.query(models.BomItem).options(selectinload(models.BomItem.bom)).filter(models.BomItem.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="BOM item not found")
    _ensure_bom_editable(item.bom)
    data = apply_bom_item_process_binding(db, payload, product_id=0)
    data.pop("bom_name", None)
    data.pop("bom_version", None)
    for key, value in data.items():
        setattr(item, key, value)
    db.commit()
    db.refresh(item)
    return serialize_bom_item(item)


@router.delete("/api/bom-items/{item_id}")
def delete_bom_item(item_id: int, db: Session = Depends(get_db), _: dict = Depends(require_permission("bom"))) -> dict:
    item = db.query(models.BomItem).options(selectinload(models.BomItem.bom)).filter(models.BomItem.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="BOM item not found")
    _ensure_bom_editable(item.bom)
    db.delete(item)
    db.commit()
    return {"deleted": True}


@router.post("/api/boms/{bom_id}/submit")
def submit_bom(bom_id: int, db: Session = Depends(get_db), _: dict = Depends(require_permission("bom"))) -> dict:
    bom = _get_bom(db, bom_id)
    if not bom.items:
        raise HTTPException(status_code=409, detail="BOM has no items")
    if bom.bom_state != "Unfrozen":
        raise HTTPException(status_code=409, detail="Only Unfrozen BOM can be submitted")
    bom.bom_state = "Frozen"
    db.commit()
    return {"id": bom.id, "bom_state": bom.bom_state}


@router.post("/api/boms/{bom_id}/approve")
def approve_bom(bom_id: int, db: Session = Depends(get_db), _: dict = Depends(require_permission(["approval", "bom"]))) -> dict:
    bom = _get_bom(db, bom_id)
    if not bom.items:
        raise HTTPException(status_code=409, detail="BOM has no items")
    if bom.bom_state not in {"Frozen", "Active"}:
        raise HTTPException(status_code=409, detail="BOM must be Frozen before approval")
    bom.bom_state = "Active"
    db.commit()
    return {"id": bom.id, "bom_state": bom.bom_state, "closed_versions": []}


@router.post("/api/boms/{bom_id}/transform", status_code=201)
def transform_bom(bom_id: int, payload: BomTransformPayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("bom"))) -> dict:
    source = _get_bom(db, bom_id)
    target_name = f"{source.bom_name}-{payload.target_type}"
    target_version = payload.version or source.bom_version
    _ensure_unique_bom(db, target_name, target_version)
    target = models.BomHeader(
        bom_name=target_name,
        bom_version=target_version,
        bom_state="Unfrozen",
        description=f"由 {source.bom_name}@{source.bom_version} 复制生成",
        owner=payload.owner or source.owner,
    )
    db.add(target)
    db.flush()
    for item in source.items:
        db.add(
            models.BomItem(
                bom_id=target.id,
                idx=item.idx,
                bom_name=target.bom_name,
                bom_version=target.bom_version,
                material_type=item.material_type,
                material_def_name=item.material_def_name,
                material_def_version=item.material_def_version,
                require_quantity=item.require_quantity,
                unit=item.unit,
                process_step_name=item.process_step_name,
                process_step_version=item.process_step_version,
            )
        )
    db.commit()
    created = _refresh_bom(db, target.id)
    result = serialize_bom(created)
    result["transform_diff"] = {
        "source": {"id": source.id, "bom_name": source.bom_name, "bom_version": source.bom_version},
        "target": {"id": created.id, "bom_name": created.bom_name, "bom_version": created.bom_version},
        "items_total": len(created.items),
        "items_added": len(created.items),
        "items_removed": 0,
        "items_changed": 0,
        "process_unassigned": sum(1 for item in created.items if not item.process_step_name),
        "process_is_complete": all(item.process_step_name for item in created.items),
    }
    return result


@router.get("/api/boms/{bom_id}/compare/{target_bom_id}")
def compare_bom(bom_id: int, target_bom_id: int, db: Session = Depends(get_db)) -> dict:
    base = _get_bom(db, bom_id)
    target = _get_bom(db, target_bom_id)
    base_map = {bom_item_compare_key(item): item for item in base.items}
    target_map = {bom_item_compare_key(item): item for item in target.items}
    changes = []
    for key in sorted(set(base_map) | set(target_map)):
        base_item = base_map.get(key)
        target_item = target_map.get(key)
        if base_item and not target_item:
            changes.append(serialize_bom_compare_item("删除", None, base_item))
        elif target_item and not base_item:
            changes.append(serialize_bom_compare_item("新增", target_item, None))
        elif base_item and target_item and (
            base_item.require_quantity != target_item.require_quantity
            or base_item.unit != target_item.unit
            or base_item.material_type != target_item.material_type
            or base_item.material_def_version != target_item.material_def_version
        ):
            changes.append(serialize_bom_compare_item("变更", target_item, base_item))
    return {
        "base": {"id": base.id, "bom_name": base.bom_name, "bom_version": base.bom_version},
        "target": {"id": target.id, "bom_name": target.bom_name, "bom_version": target.bom_version},
        "summary": {
            "added": len([item for item in changes if item["change_type"] == "新增"]),
            "removed": len([item for item in changes if item["change_type"] == "删除"]),
            "changed": len([item for item in changes if item["change_type"] == "变更"]),
        },
        "changes": changes,
    }


@router.get("/api/boms/{bom_id}/process-coverage")
def bom_process_coverage(bom_id: int, db: Session = Depends(get_db)) -> dict:
    bom = _get_bom(db, bom_id)
    unassigned_items = [
        {
            "item_id": item.id,
            "material_def_name": item.material_def_name,
            "material_type": item.material_type,
            "require_quantity": item.require_quantity,
            "unit": item.unit,
            "process_step_name": item.process_step_name,
        }
        for item in sorted(bom.items, key=lambda row: row.id)
        if not item.process_step_name
    ]
    total = len(bom.items)
    unassigned = len(unassigned_items)
    assigned = total - unassigned
    return {
        "bom_id": bom.id,
        "bom_name": bom.bom_name,
        "bom_version": bom.bom_version,
        "total_items": total,
        "assigned": assigned,
        "unassigned": unassigned,
        "coverage_rate": round(assigned / total, 4) if total else 0.0,
        "is_complete": unassigned == 0,
        "unassigned_items": unassigned_items,
    }


@router.get("/api/boms/where-used/{material_name}")
def bom_where_used(material_name: str, db: Session = Depends(get_db)) -> list[dict]:
    rows = (
        db.query(models.BomItem)
        .join(models.BomHeader)
        .options(selectinload(models.BomItem.bom))
        .filter(models.BomItem.material_def_name == material_name)
        .order_by(models.BomHeader.bom_name, models.BomHeader.bom_version)
        .all()
    )
    return [
        {
            "bom_id": item.bom_id,
            "bom_name": item.bom.bom_name,
            "bom_version": item.bom.bom_version,
            "bom_state": item.bom.bom_state,
            "material_type": item.material_type,
            "material_def_name": item.material_def_name,
            "material_def_version": item.material_def_version,
            "require_quantity": item.require_quantity,
            "unit": item.unit,
            "process_step_name": item.process_step_name,
            "process_step_version": item.process_step_version,
        }
        for item in rows
    ]


@router.get("/api/boms/{bom_id}/lineage")
def bom_lineage(bom_id: int, db: Session = Depends(get_db)) -> dict:
    bom = _get_bom(db, bom_id)
    current = {"id": bom.id, "bom_name": bom.bom_name, "bom_version": bom.bom_version, "bom_state": bom.bom_state}
    related = (
        db.query(models.BomHeader)
        .filter(models.BomHeader.bom_name == bom.bom_name, models.BomHeader.id != bom.id)
        .order_by(models.BomHeader.bom_version)
        .all()
    )
    descendants = [{"id": row.id, "bom_name": row.bom_name, "bom_version": row.bom_version, "bom_state": row.bom_state} for row in related]
    return {"current": current, "ancestors": [], "descendants": descendants, "has_lineage": bool(descendants)}


def _style_bom_export_ws(ws, bom: models.BomHeader | object, include_data: bool = True) -> None:
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    center = Alignment(horizontal="center", vertical="center")
    ws.append(["BOM名称", getattr(bom, "bom_name", "")])
    ws.append(["BOM版本", getattr(bom, "bom_version", "")])
    ws.append(["状态", getattr(bom, "bom_state", "")])
    ws.append(["负责人", getattr(bom, "owner", "")])
    ws.append([])
    ws.append(BOM_EXPORT_COLUMNS)
    for col_idx in range(1, len(BOM_EXPORT_COLUMNS) + 1):
        cell = ws.cell(row=6, column=col_idx)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = [8, 16, 24, 12, 12, 10, 18, 12][col_idx - 1]
    if include_data:
        for item in getattr(bom, "items", []):
            ws.append([
                item.idx,
                item.material_type,
                item.material_def_name,
                item.material_def_version,
                item.require_quantity,
                item.unit,
                item.process_step_name,
                item.process_step_version,
            ])
            for col_idx in range(1, len(BOM_EXPORT_COLUMNS) + 1):
                ws.cell(row=ws.max_row, column=col_idx).border = thin_border
    ws.freeze_panes = "A7"


@router.get("/api/boms/{bom_id}/export")
def export_bom_excel(bom_id: int, db: Session = Depends(get_db)):
    bom = _get_bom(db, bom_id)
    wb = Workbook()
    ws = wb.active
    ws.title = "BOM"
    _style_bom_export_ws(ws, bom, include_data=True)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"{bom.bom_name}-{bom.bom_version}.xlsx"
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/api/boms/template")
def download_bom_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "BOM模板"

    class DummyBom:
        bom_name = "BOM-DEMO"
        bom_version = "001"
        bom_state = "Unfrozen"
        owner = ""

    _style_bom_export_ws(ws, DummyBom(), include_data=False)
    ws.append([1, "Consumable", "MAT-DEMO", "001", 1, "EA", "STEP-DEMO", "001"])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="BOM_template.xlsx"'},
    )


@router.post("/api/boms/{bom_id}/import")
async def import_bom_excel(bom_id: int, file: UploadFile = File(...), db: Session = Depends(get_db), _: dict = Depends(require_permission("bom"))) -> dict:
    bom = _get_bom(db, bom_id)
    _ensure_bom_editable(bom)
    content = await file.read()
    wb = load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active
    imported = 0
    warnings: list[dict] = []
    errors: list[dict] = []
    existing = {(item.material_type, item.material_def_name, item.process_step_name) for item in bom.items}
    for row_idx, row in enumerate(ws.iter_rows(min_row=7, values_only=True), 7):
        if not row or not row[2]:
            continue
        try:
            idx = int(row[0]) if row[0] not in (None, "") else None
            quantity = float(row[4]) if row[4] not in (None, "") else None
        except (TypeError, ValueError):
            warnings.append({"row": row_idx, "message": "序号或需求量格式错误，已跳过"})
            continue
        material_type = str(row[1] or "Consumable").strip()
        material_name = str(row[2] or "").strip()
        step_name = str(row[6] or "").strip()
        key = (material_type, material_name, step_name)
        if key in existing:
            errors.append({"row": row_idx, "message": f"{material_name} 在当前 BOM 中已存在"})
            continue
        existing.add(key)
        db.add(
            models.BomItem(
                bom_id=bom.id,
                idx=idx,
                bom_name=bom.bom_name,
                bom_version=bom.bom_version,
                material_type=material_type,
                material_def_name=material_name,
                material_def_version=str(row[3] or "").strip(),
                require_quantity=quantity,
                unit=str(row[5] or "").strip(),
                process_step_name=step_name,
                process_step_version=str(row[7] or "").strip(),
            )
        )
        imported += 1
    db.commit()
    return {"imported": imported, "bom_id": bom_id, "warnings": warnings, "errors": errors, "skipped": len(warnings)}


@router.post("/api/boms/{bom_id}/batch-replace")
def bom_batch_replace(bom_id: int, payload: BomBatchReplacePayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("bom"))) -> dict:
    bom = _get_bom(db, bom_id)
    _ensure_bom_editable(bom)
    items = db.query(models.BomItem).filter(models.BomItem.bom_id == bom.id, models.BomItem.material_def_name == payload.from_code).all()
    for item in items:
        item.material_def_name = payload.to_code
        if payload.to_name:
            item.material_def_version = payload.to_name
    db.commit()
    return {"updated": len(items), "bom_id": bom.id, "from_code": payload.from_code, "to_code": payload.to_code}


@router.post("/api/boms/{bom_id}/batch-quantity")
def bom_batch_quantity(bom_id: int, payload: BomBatchQuantityPayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("bom"))) -> dict:
    bom = _get_bom(db, bom_id)
    _ensure_bom_editable(bom)
    items = db.query(models.BomItem).filter(models.BomItem.bom_id == bom.id, models.BomItem.id.in_(payload.item_ids)).all()
    for item in items:
        item.require_quantity = payload.quantity
    db.commit()
    return {"updated": len(items), "bom_id": bom.id, "quantity": payload.quantity}


@router.post("/api/boms/{bom_id}/batch-delete")
def bom_batch_delete(bom_id: int, payload: BomBatchDeletePayload, db: Session = Depends(get_db), _: dict = Depends(require_permission("bom"))) -> dict:
    bom = _get_bom(db, bom_id)
    _ensure_bom_editable(bom)
    deleted = (
        db.query(models.BomItem)
        .filter(models.BomItem.bom_id == bom.id, models.BomItem.id.in_(payload.item_ids))
        .delete(synchronize_session=False)
    )
    db.commit()
    return {"deleted": deleted, "bom_id": bom.id}


@router.get("/api/boms/{bom_id}/version-history")
def bom_version_history(bom_id: int, db: Session = Depends(get_db)) -> list[dict]:
    bom = _get_bom(db, bom_id)
    rows = db.query(models.BomHeader).filter(models.BomHeader.bom_name == bom.bom_name).order_by(models.BomHeader.bom_version).all()
    return [
        {
            "id": row.id,
            "bom_name": row.bom_name,
            "bom_version": row.bom_version,
            "bom_state": row.bom_state,
            "owner": row.owner,
            "description": row.description,
            "is_current": row.id == bom.id,
        }
        for row in rows
    ]


@router.get("/api/baselines")
def baselines(page: int = 1, page_size: int = 20, keyword: str = "", db: Session = Depends(get_db)) -> dict:
    q = db.query(models.ProductBaseline).options(selectinload(models.ProductBaseline.product), selectinload(models.ProductBaseline.items))
    if keyword:
        kw = f"%{keyword}%"
        q = q.filter(models.ProductBaseline.baseline_no.ilike(kw) | models.ProductBaseline.name.ilike(kw))
    total = q.count()
    rows = q.order_by(models.ProductBaseline.id.desc()).offset((page - 1) * page_size).limit(page_size).all()
    return {
        "items": [
            {
                "id": row.id,
                "baseline_no": row.baseline_no,
                "name": row.name,
                "product_model": row.product.product_def_name if row.product else "",
                "version": row.version,
                "status": row.status,
                "created_by": row.created_by,
                "released_at": row.released_at,
                "items": [
                    {
                        "id": item.id,
                        "item_type": item.item_type,
                        "item_no": item.item_no,
                        "title": item.title,
                        "version": item.version,
                        "status": item.status,
                    }
                    for item in row.items
                ],
            }
            for row in rows
        ],
        "total": total,
        "page": page,
        "page_size": page_size,
    }
